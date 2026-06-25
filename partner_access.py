"""Partner real-time access routes.

This module intentionally keeps the new partner-access feature isolated from
the existing board code. It reads only from whitelisted local PostgreSQL views.
"""

import logging
import configparser
import io
import re
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

from flask import Blueprint, jsonify, render_template, request, send_file

from config.menu import MENU_CONFIG
from db_connection import get_db_connection
from permission_helpers import enforce_permission
from timezone_config import get_korean_time


partner_access_bp = Blueprint("partner_access", __name__)

PARTNER_ACCESS_MENU_CODE = "PARTNER_ACCESS"
DEFAULT_RESULT_LIMIT = 1000
DEFAULT_DISTINCT_LIMIT = 200
DEFAULT_LOCATION_LOOKUP_TABLE = "vw_card_reader_info_dedup"
DEFAULT_LOCATION_LOOKUP_COLUMNS = {
    "site_name": "site_nm",
    "building_name": "line_nm",
    "floor_name": "area_nm",
    "detail_location": "gubun",
}

DEFAULT_SITE_OPTIONS: Tuple[Dict[str, Optional[str]], ...] = (
    {"key": "giheung", "label": "기흥", "table": "vw_eventlog_gh", "site_filter": None},
    {"key": "hwaseong", "label": "화성", "table": "vw_eventlog_hs", "site_filter": None},
    {"key": "dsr", "label": "DSR", "table": "vw_eventlog_dsr", "site_filter": None},
    {"key": "pyeongtaek", "label": "평택", "table": "vw_eventlog_pt", "site_filter": None},
    {"key": "cheonan", "label": "천안", "table": "vw_eventlogcaoy", "site_filter": "천안"},
    {"key": "onyang", "label": "온양", "table": "vw_eventlogcaoy", "site_filter": "온양"},
)

DEFAULT_COLUMN_EXPRESSIONS = {
    "event_time": "event_time",
    "direction": "direction",
    "company_cd": "company_cd",
    "employee_type": "employee_type",
    "card_type": "card_type",
    "employee_no": "employee_no",
    "domain_id": "domain_id",
    "company_name": "company_name",
    "employee_name": "employee_name",
    "phone_number": "phone_number",
    "site_name": "site_name",
    "building_name": "building_name",
    "floor_name": "floor_name",
    "detail_location": "detail_location",
    "access_level": "access_level",
    "reader_name": "reader_name",
}

ROW_FIELDS = (
    "event_time",
    "direction",
    "employee_name",
    "phone_number",
    "company_name",
    "company_cd",
    "employee_no",
    "employee_type",
    "card_type",
    "domain_id",
    "site_name",
    "building_name",
    "floor_name",
    "detail_location",
    "access_level",
    "reader_name",
    "person_key",
)

EXPORT_COLUMNS = {
    "occupancy": (
        ("event_time", "최종출입시각"),
        ("direction", "최종상태"),
        ("employee_name", "성함"),
        ("phone_number", "연락처"),
        ("company_name", "협력사명"),
        ("company_cd", "사업자번호"),
        ("employee_no", "사번"),
        ("employee_type", "임직원구분"),
        ("card_type", "카드구분"),
        ("domain_id", "DomainID"),
        ("site_name", "사업장"),
        ("building_name", "건물"),
        ("floor_name", "층"),
        ("detail_location", "상세위치"),
        ("access_level", "구분"),
        ("reader_name", "리더기명"),
    ),
    "history": (
        ("event_time", "출입시각"),
        ("direction", "IN/OUT"),
        ("employee_name", "성함"),
        ("phone_number", "연락처"),
        ("company_name", "협력사명"),
        ("company_cd", "사업자번호"),
        ("employee_no", "사번"),
        ("employee_type", "임직원구분"),
        ("card_type", "카드구분"),
        ("domain_id", "DomainID"),
        ("site_name", "사업장"),
        ("building_name", "건물"),
        ("floor_name", "층"),
        ("detail_location", "상세위치"),
        ("access_level", "구분"),
        ("reader_name", "리더기명"),
    ),
}


def _load_config() -> configparser.ConfigParser:
    config = configparser.ConfigParser()
    config.read("config.ini", encoding="utf-8")
    return config


def _get_partner_access_config() -> configparser.SectionProxy:
    config = _load_config()
    if not config.has_section("PARTNER_ACCESS"):
        config.add_section("PARTNER_ACCESS")
    return config["PARTNER_ACCESS"]


def _get_int_config(option_name: str, fallback: int) -> int:
    section = _get_partner_access_config()
    try:
        value = int(section.get(option_name, fallback=fallback))
    except (TypeError, ValueError):
        value = fallback
    return max(1, value)


def _get_result_limit() -> int:
    return _get_int_config("result_limit", DEFAULT_RESULT_LIMIT)


def _get_distinct_limit() -> int:
    return _get_int_config("distinct_limit", DEFAULT_DISTINCT_LIMIT)


def _safe_table_name(table_name: str) -> str:
    table_name = (table_name or "").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*(\.[A-Za-z_][A-Za-z0-9_]*)?", table_name):
        raise ValueError(f"출입정보 테이블 설정이 올바르지 않습니다: {table_name}")
    return table_name


def _safe_column_name(column_name: str) -> str:
    column_name = (column_name or "").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", column_name):
        raise ValueError(f"출입정보 컬럼 설정이 올바르지 않습니다: {column_name}")
    return column_name


def _get_location_lookup_config() -> Dict[str, Any]:
    section = _get_partner_access_config()
    table_name = _safe_table_name(
        section.get("location_lookup_table", fallback=DEFAULT_LOCATION_LOOKUP_TABLE)
    )
    columns = {
        alias: _safe_column_name(
            section.get(f"location_{alias}_column", fallback=fallback)
        )
        for alias, fallback in DEFAULT_LOCATION_LOOKUP_COLUMNS.items()
    }
    return {"table": table_name, "columns": columns}


def _get_site_options() -> Tuple[Dict[str, Optional[str]], ...]:
    section = _get_partner_access_config()
    site_options = []
    for site in DEFAULT_SITE_OPTIONS:
        key = site["key"]
        table_name = _safe_table_name(section.get(f"{key}_table", fallback=site["table"]))
        label = section.get(f"{key}_label", fallback=site["label"]).strip() or site["label"]
        site_filter = section.get(f"{key}_site_filter", fallback=site.get("site_filter") or "").strip()
        site_options.append({
            "key": key,
            "label": label,
            "table": table_name,
            "site_filter": site_filter or None,
        })
    return tuple(site_options)


def _get_column_expressions() -> Dict[str, str]:
    section = _get_partner_access_config()
    expressions = {}
    for alias, fallback in DEFAULT_COLUMN_EXPRESSIONS.items():
        option_name = f"{alias}_column"
        expression = section.get(option_name, fallback=fallback).strip()
        expressions[alias] = expression or fallback
    return expressions


def _build_eventlog_subquery(table_name: str) -> str:
    columns = _get_column_expressions()
    normalized_select = ",\n                ".join(
        f"{expression} AS {alias}"
        for alias, expression in columns.items()
    )
    return f"""
        SELECT
            event_time,
            direction,
            company_name,
            company_cd,
            employee_name,
            phone_number,
            employee_no,
            employee_type,
            card_type,
            domain_id,
            site_name,
            building_name,
            floor_name,
            detail_location,
            access_level,
            reader_name,
            COALESCE(TRIM(company_cd::text), '') || '|' ||
            COALESCE(TRIM(employee_type::text), '') || '|' ||
            COALESCE(TRIM(card_type::text), '') || '|' ||
            COALESCE(TRIM(employee_no::text), '') || '|' ||
            COALESCE(TRIM(domain_id::text), '') AS person_key
        FROM (
            SELECT
                {normalized_select}
            FROM {table_name}
        ) normalized_eventlog
    """


def _get_site(site_key: str) -> Dict[str, Optional[str]]:
    site_map = {site["key"]: site for site in _get_site_options()}
    site = site_map.get((site_key or "").strip())
    if not site:
        raise ValueError("사업장을 선택해 주세요.")
    return site


def _parse_datetime(value: str, label: str) -> datetime:
    raw_value = (value or "").strip()
    if not raw_value:
        raise ValueError(f"{label}을(를) 입력해 주세요.")
    try:
        return datetime.fromisoformat(raw_value)
    except ValueError as exc:
        raise ValueError(f"{label} 형식이 올바르지 않습니다.") from exc


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_list(value: Any) -> List[str]:
    if not value:
        return []
    if isinstance(value, str):
        value = [value]
    normalized = []
    for item in value:
        text = _normalize_text(item)
        if text:
            normalized.append(text)
    return normalized


def _row_to_dict(row: Any) -> Dict[str, Any]:
    result = {}
    for field in ROW_FIELDS:
        try:
            value = row[field]
        except Exception:
            value = None
        if isinstance(value, datetime):
            value = value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, date):
            value = value.isoformat()
        result[field] = "" if value is None else value
    return result


def _append_site_filter(where: List[str], params: List[Any], site: Dict[str, Optional[str]]) -> None:
    site_filter = site.get("site_filter")
    if site_filter:
        where.append("site_name ILIKE %s")
        params.append(f"%{site_filter}%")


def _append_in_filter(
    where: List[str],
    params: List[Any],
    column_name: str,
    values: Iterable[str],
) -> None:
    cleaned = [value for value in values if value]
    if not cleaned:
        return
    placeholders = ", ".join(["%s"] * len(cleaned))
    where.append(f"{column_name} IN ({placeholders})")
    params.extend(cleaned)


def _append_person_filters(where: List[str], params: List[Any], filters: Dict[str, Any]) -> None:
    company_name = _normalize_text(filters.get("company_name"))
    employee_name = _normalize_text(filters.get("employee_name"))
    if company_name:
        where.append("company_name ILIKE %s")
        params.append(f"%{company_name}%")
    if employee_name:
        where.append("employee_name ILIKE %s")
        params.append(f"%{employee_name}%")


def _append_location_filters(where: List[str], params: List[Any], filters: Dict[str, Any]) -> None:
    building_name = _normalize_text(filters.get("building_name"))
    floor_names = _normalize_list(filters.get("floor_names"))
    detail_locations = _normalize_list(filters.get("detail_locations"))
    detail_text = _normalize_text(filters.get("detail_text"))

    if building_name:
        where.append("building_name = %s")
        params.append(building_name)
    _append_in_filter(where, params, "floor_name", floor_names)

    if detail_locations:
        _append_in_filter(where, params, "detail_location", detail_locations)
    elif detail_text:
        where.append("detail_location ILIKE %s")
        params.append(f"%{detail_text}%")


def _has_required_discriminator(filters: Dict[str, Any]) -> bool:
    return any(
        [
            _normalize_text(filters.get("company_name")),
            _normalize_text(filters.get("employee_name")),
            _normalize_text(filters.get("building_name")),
            _normalize_list(filters.get("floor_names")),
            _normalize_list(filters.get("detail_locations")),
            _normalize_text(filters.get("detail_text")),
        ]
    )


def _execute_rows(sql: str, params: List[Any]) -> List[Dict[str, Any]]:
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(sql, tuple(params))
        return [_row_to_dict(row) for row in cursor.fetchall()]
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        conn.close()


def _query_distinct_values(
    site_key: str,
    target_column: str,
    search_term: str = "",
    building_name: str = "",
    floor_names: Optional[List[str]] = None,
) -> List[str]:
    site = _get_site(site_key)
    lookup = _get_location_lookup_config()
    table_name = lookup["table"]
    columns = lookup["columns"]
    if target_column not in {"building_name", "floor_name", "detail_location"}:
        raise ValueError(f"지원하지 않는 출입정보 조회 컬럼입니다: {target_column}")

    target_lookup_column = columns[target_column]
    site_lookup_column = columns["site_name"]
    building_lookup_column = columns["building_name"]
    floor_lookup_column = columns["floor_name"]

    where = [
        f"{site_lookup_column} = %s",
        f"{target_lookup_column} IS NOT NULL",
        f"TRIM({target_lookup_column}::text) <> ''",
    ]
    params: List[Any] = [site.get("site_filter") or site["label"]]

    if building_name:
        where.append(f"{building_lookup_column} = %s")
        params.append(building_name)
    if floor_names:
        _append_in_filter(where, params, floor_lookup_column, floor_names)
    if search_term:
        where.append(f"{target_lookup_column} ILIKE %s")
        params.append(f"%{search_term}%")

    sql = f"""
        SELECT value
        FROM (
            SELECT DISTINCT TRIM({target_lookup_column}::text) AS value
            FROM {table_name}
            WHERE {" AND ".join(where)}
        ) distinct_values
        ORDER BY
            CASE WHEN value ~ '^[0-9]+$' THEN 0 ELSE 1 END,
            CASE WHEN value ~ '^[0-9]+$' THEN value::integer END NULLS LAST,
            value
        LIMIT %s
    """
    params.append(_get_distinct_limit())

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(sql, tuple(params))
        values = []
        for row in cursor.fetchall():
            try:
                value = row["value"]
            except Exception:
                value = row[0]
            text = _normalize_text(value)
            if text:
                values.append(text)
        return values
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        conn.close()


def _search_history(site: Dict[str, Optional[str]], payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    started_at = _parse_datetime(payload.get("start_at"), "시작시각")
    ended_at = _parse_datetime(payload.get("end_at"), "종료시각")
    if started_at > ended_at:
        raise ValueError("시작시각은 종료시각보다 늦을 수 없습니다.")

    where = ["event_time >= %s", "event_time <= %s"]
    params: List[Any] = [started_at, ended_at]
    _append_site_filter(where, params, site)
    _append_person_filters(where, params, payload)
    _append_location_filters(where, params, payload)

    sql = f"""
        SELECT *
        FROM ({_build_eventlog_subquery(site["table"])}) eventlog
        WHERE {" AND ".join(where)}
        ORDER BY event_time DESC
        LIMIT %s
    """
    params.append(_get_result_limit())
    return _execute_rows(sql, params)


def _search_occupancy(site: Dict[str, Optional[str]], payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    as_of = _parse_datetime(payload.get("as_of"), "기준시각")

    base_where = ["event_time <= %s"]
    base_params: List[Any] = [as_of]
    _append_site_filter(base_where, base_params, site)
    _append_person_filters(base_where, base_params, payload)

    latest_where = ["UPPER(TRIM(COALESCE(direction::text, ''))) = 'IN'"]
    latest_params: List[Any] = []
    _append_location_filters(latest_where, latest_params, payload)

    sql = f"""
        WITH base AS (
            SELECT *
            FROM ({_build_eventlog_subquery(site["table"])}) eventlog
            WHERE {" AND ".join(base_where)}
        ),
        latest AS (
            SELECT DISTINCT ON (person_key) *
            FROM base
            WHERE person_key <> '||||'
            ORDER BY person_key, event_time DESC
        )
        SELECT *
        FROM latest
        WHERE {" AND ".join(latest_where)}
        ORDER BY event_time DESC
        LIMIT %s
    """
    return _execute_rows(sql, [*base_params, *latest_params, _get_result_limit()])


def _validate_search_payload(payload: Dict[str, Any]) -> Tuple[str, Dict[str, Optional[str]]]:
    site = _get_site(payload.get("site_key", ""))
    mode = (payload.get("mode") or "occupancy").strip()

    if mode not in {"occupancy", "history"}:
        raise ValueError("조회 모드가 올바르지 않습니다.")
    if mode == "occupancy" and not _normalize_text(payload.get("as_of")):
        raise ValueError("기준시각을 입력해 주세요.")
    if mode == "history" and (
        not _normalize_text(payload.get("start_at")) or not _normalize_text(payload.get("end_at"))
    ):
        raise ValueError("조회기간을 입력해 주세요.")
    if not _has_required_discriminator(payload):
        raise ValueError("성명, 협력사명, 건물/층, 상세위치 중 하나 이상 입력해 주세요.")
    return mode, site


def _search_by_payload(payload: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]]]:
    mode, site = _validate_search_payload(payload)
    rows = _search_occupancy(site, payload) if mode == "occupancy" else _search_history(site, payload)
    return mode, rows


def _build_excel_response(mode: str, rows: List[Dict[str, Any]]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "재실인원" if mode == "occupancy" else "출입이력"

    columns = EXPORT_COLUMNS[mode]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (_, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 34)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    mode_label = "재실인원" if mode == "occupancy" else "출입이력"
    filename = f"협력사_출입정보_{mode_label}_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@partner_access_bp.route("/partner-access")
def partner_access_page():
    guard = enforce_permission(PARTNER_ACCESS_MENU_CODE, "view")
    if guard:
        return guard
    return render_template(
        "partner-access.html",
        menu=MENU_CONFIG,
        sites=list(_get_site_options()),
        result_limit=_get_result_limit(),
    )


@partner_access_bp.route("/api/partner-access/buildings")
def api_partner_access_buildings():
    guard = enforce_permission(PARTNER_ACCESS_MENU_CODE, "view", response_type="json")
    if guard:
        return guard
    try:
        values = _query_distinct_values(
            request.args.get("site", ""),
            "building_name",
            request.args.get("q", "").strip(),
        )
        return jsonify({"success": True, "data": values})
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("partner-access building lookup failed")
        return jsonify({"success": False, "message": str(exc)}), 500


@partner_access_bp.route("/api/partner-access/floors")
def api_partner_access_floors():
    guard = enforce_permission(PARTNER_ACCESS_MENU_CODE, "view", response_type="json")
    if guard:
        return guard
    try:
        building_name = request.args.get("building", "").strip()
        if not building_name:
            return jsonify({"success": False, "message": "건물을 선택해 주세요."}), 400
        values = _query_distinct_values(
            request.args.get("site", ""),
            "floor_name",
            request.args.get("q", "").strip(),
            building_name=building_name,
        )
        return jsonify({"success": True, "data": values})
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("partner-access floor lookup failed")
        return jsonify({"success": False, "message": str(exc)}), 500


@partner_access_bp.route("/api/partner-access/locations")
def api_partner_access_locations():
    guard = enforce_permission(PARTNER_ACCESS_MENU_CODE, "view", response_type="json")
    if guard:
        return guard
    try:
        floor_names = _normalize_list(request.args.getlist("floor"))
        values = _query_distinct_values(
            request.args.get("site", ""),
            "detail_location",
            request.args.get("q", "").strip(),
            building_name=request.args.get("building", "").strip(),
            floor_names=floor_names,
        )
        return jsonify({"success": True, "data": values})
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("partner-access location lookup failed")
        return jsonify({"success": False, "message": str(exc)}), 500


@partner_access_bp.route("/api/partner-access/search", methods=["POST"])
def api_partner_access_search():
    guard = enforce_permission(PARTNER_ACCESS_MENU_CODE, "view", response_type="json")
    if guard:
        return guard
    try:
        payload = request.get_json(silent=True) or {}
        mode, rows = _search_by_payload(payload)
        return jsonify({
            "success": True,
            "data": rows,
            "count": len(rows),
            "limit": _get_result_limit(),
            "limited": len(rows) >= _get_result_limit(),
        })
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("partner-access search failed")
        return jsonify({"success": False, "message": str(exc)}), 500


@partner_access_bp.route("/api/partner-access/export", methods=["POST"])
def api_partner_access_export():
    guard = enforce_permission(PARTNER_ACCESS_MENU_CODE, "view", response_type="json")
    if guard:
        return guard
    try:
        payload = request.get_json(silent=True) or {}
        mode, rows = _search_by_payload(payload)
        return _build_excel_response(mode, rows)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        logging.exception("partner-access export failed")
        return jsonify({"success": False, "message": str(exc)}), 500
