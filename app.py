import os
import logging
from datetime import datetime
import pytz
from pathlib import Path
import shutil
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, Response, send_from_directory, abort
import configparser
from timezone_config import KST, get_korean_time, get_korean_time_str
from config.menu import MENU_CONFIG
from database_config import db_config, partner_manager
import re
import base64
import sqlite3
import math
from types import SimpleNamespace
from typing import Any, Dict, List, Optional
from board_services import CodeService, ItemService
from repositories.common.column_config_repository import ColumnConfigRepository
from column_service import ColumnConfigService
from search_popup_service import SearchPopupService
from column_sync_service import ColumnSyncService
from db_connection import get_db_connection
from database_config import execute_SQL
from db.upsert import safe_upsert
from column_utils import normalize_column_types, determine_linked_type
from upload_utils import sanitize_filename, validate_uploaded_files
from permission_helpers import (
    is_super_admin,
    SUPER_ADMIN_USERS,
    build_user_menu_config,
    enforce_permission,
    resolve_menu_code,
)
from audit_logger import record_audit_log, record_board_action, record_menu_view, normalize_scope, normalize_action, normalize_result
from notification_service import get_notification_service, NotificationError
from permission_api import register_permission_routes
from add_page_routes import follow_sop_bp, full_process_bp, safe_workplace_bp, _response_info
from boards.safety_instruction import safety_instruction_bp
from controllers.boards.accident_controller import (
    AccidentController,
    build_accident_config,
)
from repositories.boards.accident_repository import AccidentRepository
from controllers.boards.safety_instruction_controller import (
    SafetyInstructionController,
    build_safety_instruction_config,
)
from repositories.boards.safety_instruction_repository import SafetyInstructionRepository
from utils.sql_filters import sql_is_active_true, sql_is_deleted_false
from typing import List
import schedule
import threading
import time
# SSO 관련 imports 추가
import jwt
import json
import uuid
import ssl
from cryptography import x509
from cryptography.hazmat.backends import default_backend


def generate_manual_accident_number(cursor):
    """수기입력 사고번호 자동 생성 (ACCYYMMDD00 형식)

    기준 테이블: accidents_cache (SOT)
    """
    today = get_korean_time()
    date_part = today.strftime('%y%m%d')  # 240822
    pattern = f'ACC{date_part}%'
    # accidents_cache에서 최신 번호 조회
    try:
        cursor.execute(
            """
            SELECT accident_number 
            FROM accidents_cache 
            WHERE accident_number LIKE %s
            ORDER BY accident_number DESC
            LIMIT 1
            """,
            (pattern,)
        )
        last = cursor.fetchone()
    except Exception:
        last = None
    if last:
        last_num = last[0] if not isinstance(last, dict) else last['accident_number']
    else:
        last_num = None
    if last_num and str(last_num).startswith('ACC'):
        try:
            seq = int(str(last_num)[-2:]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f'ACC{date_part}{seq:02d}'

app = Flask(__name__, static_folder='static')
app.register_blueprint(follow_sop_bp)
app.register_blueprint(full_process_bp)
app.register_blueprint(safety_instruction_bp)
app.register_blueprint(safe_workplace_bp)
register_permission_routes(app)


@app.context_processor
def inject_user_menu():
    try:
        return {'user_menu': build_user_menu_config()}
    except Exception as exc:
        logging.debug('inject_user_menu failed: %s', exc)
        return {'user_menu': MENU_CONFIG}

# Jinja2 템플릿 필터 정의
@app.template_filter('date_only')
def date_only_filter(datetime_str):
    """날짜시간 문자열에서 날짜만 추출 (YYYY-MM-DD)"""
    if not datetime_str:
        return ''
    # 2025-08-27 14:29:03 → 2025-08-27
    datetime_str = str(datetime_str)
    return datetime_str.split(' ')[0] if ' ' in datetime_str else datetime_str

@app.template_filter('date_korean')
def date_korean_filter(datetime_str):
    """날짜를 한국식 표기로 변환 (YYYY년 MM월 DD일)"""
    if not datetime_str:
        return ''
    datetime_str = str(datetime_str)
    date_part = datetime_str.split(' ')[0] if ' ' in datetime_str else datetime_str
    try:
        parts = date_part.split('-')
        if len(parts) == 3:
            return f"{parts[0]}년 {parts[1]}월 {parts[2]}일"
    except:
        pass
    return date_part

# 메뉴 설정
menu = MENU_CONFIG

# 설정 파일에서 환경 설정 로드
app.secret_key = db_config.config.get('DEFAULT', 'SECRET_KEY')

# 세션 쿠키 설정 (localhost와 127.0.0.1 호환)
app.config['SESSION_COOKIE_DOMAIN'] = None  # 도메인 제한 없음
app.config['SESSION_COOKIE_SECURE'] = False  # HTTP/HTTPS 둘 다 허용
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.debug = db_config.config.getboolean('DEFAULT', 'DEBUG')

# Jinja2 필터 추가 (JSON 파싱용)
import json as pyjson  # GPT 권고: json 모듈 별칭 사용으로 충돌 방지
def from_json_filter(value):
    """JSON 필터 - dict/list는 그대로, 문자열은 파싱"""
    # 이미 dict나 list면 그대로 반환
    if isinstance(value, (list, dict)):
        return value
    # 문자열이면 파싱 시도
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return []
        try:
            return pyjson.loads(v)
        except:
            return []
    # 그 외의 경우 빈 리스트
    return []
app.jinja_env.filters['from_json'] = from_json_filter

# 리스트 요약 필터: "첫번째이름 외 N명" 형태로 표시
@app.template_filter('list_summary')
def list_summary_filter(value):
    """value가 문자열(JSON) / 리스트 / 기타일 때 공통 요약 문자열 반환"""
    import json as _json
    # 1) normalize to list
    data = None
    try:
        if isinstance(value, (list, tuple)):
            data = list(value)
        elif isinstance(value, str):
            v = value.strip()
            if not v:
                data = []
            else:
                try:
                    parsed = _json.loads(v)
                    data = parsed if isinstance(parsed, list) else []
                except Exception:
                    data = []
        else:
            data = []
    except Exception:
        data = []

    if not data:
        return '-'

    # 2) pick first display name
    def extract_name(item):
        if isinstance(item, dict):
            for k in (
                'name', 'worker_name', 'violator_name', 'employee_name', 'person_name',
                'victim_name', 'value', 'label', 'title'
            ):
                if k in item and item[k]:
                    return str(item[k])
            # 아무 키도 못 찾으면 dict 전체를 문자열로
            return ''
        elif isinstance(item, str):
            return item
        else:
            return str(item)

    first = extract_name(data[0])
    others = max(0, len(data) - 1)

    if not first:
        # 이름을 못 뽑으면 개수만 표시
        return f"{len(data)}개 항목"

    return f"{first} 외 {others}명" if others > 0 else first

DB_PATH = db_config.local_db_path

_safety_instruction_controller = SafetyInstructionController(
    build_safety_instruction_config(),
    SafetyInstructionRepository(DB_PATH),
    menu_config=MENU_CONFIG,
)

_accident_controller = AccidentController(
    build_accident_config(),
    AccidentRepository(DB_PATH),
    menu_config=MENU_CONFIG,
)

PASSWORD = db_config.config.get('DEFAULT', 'EDIT_PASSWORD')
ADMIN_PASSWORD = db_config.config.get('DEFAULT', 'ADMIN_PASSWORD')

# 관리자 인증 함수
def require_admin_auth(f):
    """관리자 인증이 필요한 함수에 사용하는 decorator"""
    from functools import wraps
    
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # 세션에서 관리자 인증 확인
        if session.get('admin_authenticated') != True:
            # 인증이 안된 경우 비밀번호 입력 페이지로 리다이렉트
            target_path = request.full_path if request.query_string else request.path
            if target_path.endswith('?'):
                target_path = target_path[:-1]
            return render_template(
                'admin-login.html',
                redirect_url=target_path or '/',
                menu=MENU_CONFIG
            )
        return f(*args, **kwargs)
    return decorated_function

# 드롭다운 코드 매핑
DROPDOWN_MAPPINGS = {
    'column3': {
        'COLUMN3_001': '진행중',
        'COLUMN3_002': '완료대기',
        'COLUMN3_003': '보류',
        'COLUMN3_004': '보류3'
    }
}

def convert_dropdown_code(column_key, code):
    """드롭다운 코드를 실제 값으로 변환"""
    if column_key in DROPDOWN_MAPPINGS:
        return DROPDOWN_MAPPINGS[column_key].get(code, code)
    return code

# 로깅 설정
logging.basicConfig(
    level=getattr(logging, db_config.config.get('LOGGING', 'LOG_LEVEL')),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(db_config.config.get('LOGGING', 'LOG_FILE')),
        logging.StreamHandler()
    ]
)

# Flask 템플릿 자동 리로드 설정
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Jinja2 템플릿 캐시 비우기
app.jinja_env.cache = {}

# =====================
# Helper Functions for Linked Types
# =====================
# determine_linked_type function moved to column_utils.py for reuse across all boards

# =====================
# Scoring Service Import
# =====================
from scoring_service import calculate_score as _calc_score

# =====================
# Notification API
# =====================
@app.route('/api/notifications/send', methods=['POST'])
def api_send_notification():
    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({'success': False, 'message': 'JSON 본문이 필요합니다.'}), 400

    if not (session.get('admin_authenticated') or session.get('user_id')):
        return jsonify({'success': False, 'message': '인증이 필요합니다.'}), 401

    channel = (payload.get('channel') or 'chatbot').lower()
    event = payload.get('event')
    recipients = payload.get('recipients')
    context = payload.get('context') or {}
    metadata = payload.get('metadata') or None

    if not event:
        return jsonify({'success': False, 'message': 'event 값은 필수입니다.'}), 400
    if not recipients:
        return jsonify({'success': False, 'message': 'recipients 목록이 필요합니다.'}), 400

    service = get_notification_service()
    try:
        result = service.send_event_notification(
            channel=channel,
            event=event,
            recipients=recipients,
            context=context,
            metadata=metadata,
        )
        return jsonify({'success': True, 'result': result})
    except NotificationError as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception as exc:
        logging.exception('Notification API error: %s', exc)
        return jsonify({'success': False, 'message': '알림 전송 중 오류가 발생했습니다.'}), 500


# =====================
# Scoring API
# =====================
@app.route('/api/<board>/calculate-score', methods=['POST'])
def api_calculate_score(board):
    try:
        data = request.get_json(silent=True) or {}
        # Normalize board keys: allow hyphen
        board_norm = board.replace('-', '_')
        summary = _calc_score(board_norm, data, DB_PATH)
        return jsonify(summary)
    except Exception as e:
        logging.error(f"calculate-score error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 400

def init_db():
    """Run database migrations and seed static page definitions."""
    logging.info("[INIT] Running database migrations")
    try:
        from migrations.run_migrations import run_migrations
        run_migrations()
    except Exception as exc:
        logging.error("[INIT] Migration execution failed: %s", exc)
        raise

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            for category in MENU_CONFIG:
                for submenu in category['submenu']:
                    cursor.execute(
                        "SELECT COUNT(*) FROM pages WHERE url = %s",
                        (submenu['url'],),
                    )
                    if cursor.fetchone()[0] == 0:
                        cursor.execute(
                            "INSERT INTO pages (url, title, content) VALUES (%s, %s, %s)",
                            (
                                submenu['url'],
                                submenu['title'],
                                f"<h1>{submenu['title']}</h1><p>이 페이지의 내용을 편집하세요.</p>",
                            ),
                        )
            conn.commit()
        finally:
            cursor.close()
    except Exception as exc:
        if conn:
            conn.rollback()
        logging.error("[INIT] 기본 페이지 시드 중 오류: %s", exc)
        raise
    finally:
        if conn:
            conn.close()

    if db_config.config.getboolean('DEFAULT', 'SEED_DUMMY', fallback=False):
        logging.warning("[INIT] SEED_DUMMY 옵션은 Postgres 환경에서 지원되지 않아 건너뜁니다.")

    if db_config.external_db_enabled:
        logging.info("외부 DB 동기화 시작...")
        try:
            from database_config import maybe_daily_sync
            maybe_daily_sync(force=False)  # 하루에 한 번만 동기화
        except Exception as e:
            logging.error(f"외부 DB 동기화 실패: {e}")
    else:
        logging.info("EXTERNAL_DB_ENABLED=False - 외부 동기화를 건너뜁니다")


def sync_all_master_data():
    """모든 마스터 데이터 동기화 함수 (스케줄러용)"""
    logging.info("=" * 50)
    logging.info(f"스케줄 동기화 시작: {get_korean_time_str()}")
    logging.info("=" * 50)
    
    try:
        # 외부 DB 활성화 확인
        external_db_enabled = partner_manager.config.getboolean('DATABASE', 'EXTERNAL_DB_ENABLED', fallback=False)
        
        if not external_db_enabled:
            logging.info("외부 DB가 비활성화되어 있어 동기화를 건너뜁니다.")
            return
        
        # 각 데이터 동기화
        sync_results = {
            '협력사': False,
            '사고': False,
            '임직원': False,
            '부서': False,
            '건물': False,
            '협력사 근로자': False
        }
        
        # 1. 협력사 데이터 동기화
        try:
            logging.info("협력사 데이터 동기화 중...")
            sync_results['협력사'] = partner_manager.sync_partners_from_external_db()
        except Exception as e:
            logging.error(f"협력사 동기화 오류: {e}")
        
        # 2. 사고 데이터 동기화
        try:
            if partner_manager.config.has_option('SQL_QUERIES', 'ACCIDENTS_QUERY'):
                logging.info("사고 데이터 동기화 중...")
                sync_results['사고'] = partner_manager.sync_accidents_from_external_db()
        except Exception as e:
            logging.error(f"사고 동기화 오류: {e}")
        
        # 3. 임직원 데이터 동기화
        try:
            if partner_manager.config.has_option('MASTER_DATA_QUERIES', 'EMPLOYEE_QUERY'):
                logging.info("임직원 데이터 동기화 중...")
                sync_results['임직원'] = partner_manager.sync_employees_from_external_db()
        except Exception as e:
            logging.error(f"임직원 동기화 오류: {e}")
        
        # 4. 부서 데이터 동기화
        try:
            if partner_manager.config.has_option('MASTER_DATA_QUERIES', 'DEPARTMENT_QUERY'):
                logging.info("부서 데이터 동기화 중...")
                sync_results['부서'] = partner_manager.sync_departments_from_external_db()
        except Exception as e:
            logging.error(f"부서 동기화 오류: {e}")
        
        # 5. 건물 데이터 동기화
        try:
            if partner_manager.config.has_option('MASTER_DATA_QUERIES', 'BUILDING_QUERY'):
                logging.info("건물 데이터 동기화 중...")
                sync_results['건물'] = partner_manager.sync_buildings_from_external_db()
        except Exception as e:
            logging.error(f"건물 동기화 오류: {e}")
        
        # 6. 협력사 근로자 데이터 동기화
        try:
            if partner_manager.config.has_option('MASTER_DATA_QUERIES', 'CONTRACTOR_QUERY'):
                logging.info("협력사 근로자 데이터 동기화 중...")
                sync_results['협력사 근로자'] = partner_manager.sync_contractors_from_external_db()
        except Exception as e:
            logging.error(f"협력사 근로자 동기화 오류: {e}")
        
        # 결과 로깅
        logging.info("=" * 50)
        logging.info("동기화 결과:")
        for name, result in sync_results.items():
            status = "✅ 성공" if result else "❌ 실패"
            logging.info(f"  {name}: {status}")
        logging.info("=" * 50)
        logging.info(f"스케줄 동기화 완료: {get_korean_time_str()}")
        logging.info("=" * 50)
        
    except Exception as e:
        logging.error(f"스케줄 동기화 중 전체 오류: {e}")

def run_scheduler():
    """스케줄러 실행 함수"""
    while True:
        schedule.run_pending()
        time.sleep(60)  # 1분마다 스케줄 체크

def init_sample_data():
    """외부 DB 없을 때 샘플 데이터 생성"""
    conn = partner_manager.db_config.get_connection()
    cursor = conn.cursor()
    
    # 이미 데이터가 있는지 확인
    try:
        cursor.execute("SELECT COUNT(*) FROM partners_cache")
        count = cursor.fetchone()[0]
        
        # permanent_workers 컬럼이 있는지 확인
        # PostgreSQL: information_schema를 통해 컬럼 정보 조회
        cursor.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'partners_cache'
        """)
        columns = [col[0] for col in cursor.fetchall()]
        
        # permanent_workers 컬럼이 없으면 기존 데이터에 랜덤값 추가
        if 'permanent_workers' not in columns:
            logging.info("permanent_workers 컬럼이 없어서 기존 데이터에 값을 추가합니다")
            import random
            cursor.execute("SELECT business_number FROM partners_cache")
            existing_partners = cursor.fetchall()
            for partner in existing_partners:
                permanent_workers = random.randint(5, 500)
                cursor.execute("UPDATE partners_cache SET permanent_workers = %s WHERE business_number = %s", 
                             (permanent_workers, partner[0]))
            conn.commit()
            logging.info(f"기존 {len(existing_partners)}개 협력사에 상시근로자 데이터 추가 완료")
        
        # 데이터가 충분히 있으면 종료
        if count > 0:
            conn.close()
            return
    except Exception as e:
        logging.warning(f"데이터 확인 중 오류: {e}")
    
    logging.info("샘플 데이터 생성 중...")
    
    import random
    random.seed(42)  # 고정된 시드
    
    # 샘플 데이터 생성 로직 (기존과 동일)
    base_companies = [
        '삼성전자', 'LG전자', '현대자동차', 'SK하이닉스', 'POSCO홀딩스',
        '네이버', '카카오', '신한금융지주', '한국전력공사', 'KT',
        'LG화학', '현대중공업', '한화솔루션', 'SK텔레콤', '기아',
        '롯데케미칼', 'S-Oil', 'GS칼텍스', '두산에너빌리티', 'HD현대중공업'
    ]
    
    business_types_data = {
        '제조업': ['전자제품', '자동차', '기계', '화학', '섬유', '식품', '의약품', '철강', '플라스틱', '기타제조'],
        '건설업': ['건축공사', '토목공사', '전기공사', '통신공사', '설비공사', '조경공사', '인테리어', '기타건설'],
        'IT업': ['소프트웨어개발', '시스템통합', '데이터베이스', '네트워크', '보안', '게임개발', '웹개발', '모바일앱'],
        '서비스업': ['컨설팅', '교육', '의료', '법률', '회계', '인사', '마케팅', '디자인', '청소', '보안서비스'],
        '운수업': ['육상운송', '해상운송', '항공운송', '물류', '창고', '택배', '렌터카', '기타운송'],
        '유통업': ['도매', '소매', '전자상거래', '백화점', '마트', '편의점', '온라인쇼핑몰', '기타유통'],
        '금융업': ['은행', '증권', '보험', '카드', '리스', '투자', '자산관리', '핀테크'],
        '에너지업': ['전력', '가스', '석유', '신재생에너지', '원자력', '석탄', '기타에너지']
    }
    
    business_types = list(business_types_data.keys())
    certifications = ['ISO 9001', 'ISO 14001', 'ISO 45001', 'KS인증', 'GMP', 'HACCP', '없음']
    safety_ratings = ['A등급', 'B등급', 'C등급', 'D등급']
    products = ['전자부품', '자동차부품', '화학원료', '기계부품', '소프트웨어', '통신장비', '건설자재', '의료기기', '식품', '기타']
    
    # 203개 샘플 데이터 생성
    for i in range(203):
        if i < 20:
            company_name = f"{base_companies[i % len(base_companies)]}(주)"
            business_number = f"{100 + i:03d}81{random.randint(10000, 99999):05d}"
        else:
            company_name = f"협력업체{i-19:03d}(주)"
            business_number = f"{random.randint(100, 999)}81{random.randint(10000, 99999):05d}"
        
        representative = f"대표자{i+1:03d}"
        permanent_workers = random.randint(5, 500)  # 상시근로자 수 (5명~500명)
        partner_class = random.choice(['-', 'A', 'B', 'C'])
        
        business_type_major = random.choice(business_types)
        minor_count = random.randint(1, 2)
        selected_minors = random.sample(business_types_data[business_type_major], min(minor_count, len(business_types_data[business_type_major])))
        business_type_minor = ', '.join(selected_minors)
        
        hazard_work_flag = random.choice(['O', 'X', ''])  # O: 위험작업, X: 비위험작업, '': 미분류
        address = f"서울특별시 {random.choice(['강남구', '서초구', '송파구', '영등포구', '마포구', '종로구', '중구', '용산구'])} 샘플로{random.randint(1, 999)}"
        average_age = random.randint(25, 55)  # 평균 연령
        annual_revenue = random.randint(1, 1000) * 100000000  # 연매출 (억원 단위)
        transaction_count = random.randint(1, 50)  # 거래 차수
        
        cursor.execute('''
            INSERT INTO partners_cache (
                business_number, company_name, partner_class, business_type_major, 
                business_type_minor, hazard_work_flag, representative, address,
                average_age, annual_revenue, transaction_count, permanent_workers
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            business_number, company_name, partner_class, business_type_major,
            business_type_minor, hazard_work_flag, representative, address,
            average_age, annual_revenue, transaction_count, permanent_workers
        ))
        
        # 일부 협력사에 실제 존재하는 샘플 첨부파일 추가
        if i < 5:  # 처음 5개 협력사에만 추가
            # uploads 폴더에 실제 존재하는 파일들 사용
            real_files = [
                ('sample_for_you_1755663410.xlsx', 'uploads/sample_for_you_1755663410.xlsx', '샘플 엑셀 파일'),
                ('[Quick Guide] DS사업장 내방신청 v.03_1755666638.pdf', 'uploads/[Quick Guide] DS사업장 내방신청 v.03_1755666638.pdf', 'DS사업장 가이드'),
                ('[Quick Guide] 상생협력포털(PCMS) 회원가입 v.02_1755674307.pdf', 'uploads/[Quick Guide] 상생협력포털(PCMS) 회원가입 v.02_1755674307.pdf', 'PCMS 가입 가이드')
            ]
            
            # 하나의 실제 파일만 추가 (확실히 다운로드되도록)
            file_info = real_files[i % len(real_files)]
            file_path = os.path.join(os.getcwd(), file_info[1])
            
            # 파일이 실제로 존재하는지 확인
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                cursor.execute('''
                    INSERT INTO partner_attachments (
                        business_number, file_name, file_path, file_size, description
                    ) VALUES (%s, %s, %s, %s, %s)
                ''', (
                    business_number,
                    file_info[0],
                    file_path,  # 전체 경로로 저장
                    file_size,
                    file_info[2]
                ))
    
    conn.commit()
    conn.close()
    logging.info("샘플 데이터 생성 완료")


# ======================================================================
# 부트 동기화 훅 - 첫 요청시 한번만 실행
# ======================================================================
from flask import current_app

boot_sync_done = False

def boot_sync_once():
    """
    서버 프로세스가 어떤 방식으로 떠도(WSGI/리로더/워커), 첫 요청 들어올 때 1회 실행.
    - 마스터 데이터: 매일 1회
    - 컨텐츠 데이터: 최초 1회(또는 force)
    """
    global boot_sync_done
    if boot_sync_done:
        return
    boot_sync_done = True

    try:
        # 0) 항상 코어 스키마 보장 (pages 등 기본 테이블 생성)
        #    WSGI로 실행 시 __main__ 블록이 호출되지 않아 Postgres에서
        #    pages 등 기본 테이블이 없어지는 문제가 발생할 수 있음.
        #    idempotent하므로 안전하게 1회 실행한다.
        try:
            init_db()
        except Exception as _e:
            try:
                current_app.logger.error(f"[BOOT] init_db failed: {_e}")
            except Exception:
                pass

        from database_config import maybe_daily_sync_master, maybe_one_time_sync_content, db_config

        ext_on = db_config.config.getboolean('DATABASE', 'EXTERNAL_DB_ENABLED', fallback=False)
        init_on = db_config.config.getboolean('DATABASE', 'INITIAL_SYNC_ON_FIRST_REQUEST', fallback=False)
        current_app.logger.info(f"[BOOT] EXTERNAL_DB_ENABLED={ext_on}, INITIAL_SYNC_ON_FIRST_REQUEST={init_on}")

        # 필수 키 점검(빠르게 조기 경보)
        req_master = ['PARTNERS_QUERY','ACCIDENTS_QUERY','EMPLOYEE_QUERY','DEPARTMENT_QUERY','BUILDING_QUERY','CONTRACTOR_QUERY']
        missing_master = [k for k in req_master if not db_config.config.has_option('MASTER_DATA_QUERIES', k)]
        if missing_master:
            current_app.logger.warning(f"[BOOT] MASTER_DATA_QUERIES missing keys: {missing_master}")

        # 컨텐츠 쿼리는 옵션(있으면 사용)
        if not db_config.config.has_section('CONTENT_DATA_QUERIES'):
            current_app.logger.info("[BOOT] CONTENT_DATA_QUERIES section not found (skip content one-time sync)")

        if ext_on and init_on:
            # 1) 마스터: 매일 체크 → 필요 시 동기화
            if db_config.config.getboolean('DATABASE', 'MASTER_DATA_DAILY', fallback=True):
                current_app.logger.info("[BOOT] Master daily sync check...")
                maybe_daily_sync_master(force=False)  # 24시간 경과 시에만
                current_app.logger.info("[BOOT] Master daily sync done/kept")

            # 2) 컨텐츠: 최초 1회만
            if db_config.config.getboolean('DATABASE', 'CONTENT_DATA_ONCE', fallback=True):
                current_app.logger.info("[BOOT] Content one-time sync check...")
                maybe_one_time_sync_content(force=False)  # 최초 1회만
                current_app.logger.info("[BOOT] Content one-time sync done/kept")
        else:
            current_app.logger.info("[BOOT] Initial sync skipped (flag off or external off)")
    except Exception as e:
        current_app.logger.error(f"[BOOT] Initial sync error: {e}")

# Flask 2.3+ 호환 방식으로 첫 요청 훅 등록
@app.before_request
def check_first_request():
    if not hasattr(app, '_first_request_done'):
        app._first_request_done = True
        boot_sync_once()

@app.route("/api/test-simple")
def test_simple():
    return jsonify({"status": "ok"})

# ===== 검색 팝업 공통 API =====
@app.route("/api/search-popup/search", methods=["GET"])
def api_search_popup():
    """공통 검색 팝업 API"""
    search_type = request.args.get('type', 'company')
    search_field = request.args.get('value', '')  # 검색할 필드
    query = request.args.get('q', '')  # 검색어
    limit = int(request.args.get('limit', '50'))
    
    try:
        # SearchPopupService 사용
        search_service = SearchPopupService(DB_PATH)
        result = search_service.search(search_type, query, search_field, limit)
        
        return jsonify({
            'success': True,
            'results': result['results'],
            'total': result.get('total', len(result['results'])),
            'config': result.get('config', {})
        })
    except Exception as e:
        logging.error(f"Search popup API error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route("/api/search-popup/autocomplete", methods=["GET"])
def api_search_popup_autocomplete():
    """검색 팝업 자동완성 API"""
    search_type = request.args.get('type', 'company')
    query = request.args.get('query', '')
    
    try:
        # SearchPopupService 사용 (자동완성은 최대 10개만)
        search_service = SearchPopupService(DB_PATH)
        result = search_service.search(search_type, query, limit=10)
        
        # 자동완성 형식으로 변환
        config = search_service.search_configs.get(search_type, {})
        display_fields = config.get('display_fields', [])
        
        items = []
        for row in result['results'][:10]:
            # 첫 번째 필드를 메인으로, 나머지를 서브로
            if display_fields:
                main = row.get(display_fields[0], '')
                sub = ' | '.join([str(row.get(field, '')) for field in display_fields[1:] if row.get(field)])
                items.append({
                    'main': main,
                    'sub': sub,
                    'data': row
                })
        
        return jsonify({
            'success': True,
            'items': items
        })
    except Exception as e:
        logging.error(f"Autocomplete API error: {e}")
        return jsonify({
            'success': False,
            'items': []
        }), 500

@app.route("/")
def index():
    # 대시보드 설정 가져오기 (단순화)
    dashboard_config = {
        'url': db_config.config.get('DASHBOARD', 'DASHBOARD_URL', 
                                   fallback='https://your-dashboard.com'),
        'enabled': db_config.config.getboolean('DASHBOARD', 'DASHBOARD_ENABLED', 
                                              fallback=True)
    }
    return render_template("index.html", menu=MENU_CONFIG, dashboard_config=dashboard_config)

# ===== 검색 팝업 라우트 =====
@app.route("/search-popup")
def search_popup():
    """공통 검색 팝업"""
    search_type = request.args.get('type', 'company')
    callback = request.args.get('callback', 'handleSelection')
    
    # SearchPopupService 사용
    search_service = SearchPopupService(DB_PATH)
    
    # 설정 정보 가져오기
    config = search_service.search_configs.get(search_type, {})
    
    # search_fields를 search_options로 변환
    search_options = []
    if 'search_fields' in config:
        for field_info in config['search_fields']:
            search_options.append({
                'value': field_info['field'],
                'label': field_info['label']
            })
    
    # config에 필요한 URL 추가
    display_labels = config.get('display_labels', {})
    enhanced_config = {
        **config,
        'searchUrl': '/api/search-popup/search',
        'autocompleteUrl': '/api/search-popup/autocomplete',
        'callback': callback,
        'columns': [{'field': field, 'label': display_labels.get(field, field)} for field in config.get('display_fields', [])]
    }
    
    # 템플릿 렌더링을 시도하고, 실패하면 간단한 HTML 반환
    try:
        return render_template('search-popup.html',
                             search_type=search_type,
                             search_title=config.get('title', '검색'),
                             search_options=search_options,
                             placeholder=config.get('placeholder', '검색어를 입력하세요'),
                             config=enhanced_config,
                             menu=MENU_CONFIG)
    except:
        # 템플릿이 없으면 간단한 HTML 반환
        return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{result['config'].get('title', '검색')}</title>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
            .search-container {{ background: white; border-radius: 8px; padding: 20px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            h2 {{ margin-top: 0; color: #333; }}
            .search-box {{ display: flex; gap: 10px; margin-bottom: 20px; }}
            input[type="text"] {{ flex: 1; padding: 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 14px; }}
            button {{ padding: 10px 20px; background: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; }}
            button:hover {{ background: #0056b3; }}
            .results {{ max-height: 400px; overflow-y: auto; }}
            .result-item {{ padding: 12px; border-bottom: 1px solid #eee; cursor: pointer; transition: background 0.2s; }}
            .result-item:hover {{ background: #f8f9fa; }}
            .result-item:last-child {{ border-bottom: none; }}
            .no-results {{ text-align: center; color: #666; padding: 20px; }}
            .field-label {{ font-size: 12px; color: #666; }}
            .field-value {{ font-size: 14px; color: #333; margin-top: 2px; }}
        </style>
    </head>
    <body>
        <div class="search-container">
            <h2>{result['config'].get('title', '검색')}</h2>
            <form class="search-box" method="get">
                <input type="hidden" name="type" value="{search_type}">
                <input type="hidden" name="callback" value="{callback}">
                <select name="search_field" style="padding: 10px; border: 1px solid #ddd; border-radius: 4px; margin-right: 5px;">
                    {"".join([f'<option value="{field["field"] if isinstance(field, dict) else field}" {"selected" if (search_field == (field["field"] if isinstance(field, dict) else field)) or (not search_field and (field["field"] if isinstance(field, dict) else field) == result["config"].get("default_search_field")) else ""}>{field["label"] if isinstance(field, dict) else field}</option>' for field in result["config"].get("search_fields", [])])}
                </select>
                <input type="text" name="q" value="{query}" placeholder="{result['config'].get('placeholder', '검색어를 입력하세요')}" autofocus>
                <button type="submit">검색</button>
            </form>
            <div class="results">
                {"".join([f'''
                <div class="result-item" onclick="selectItem('{row.get(result['config']['id_field'])}', '{row.get(result['config']['display_fields'][0])}')">
                    {"".join([f'<div><span class="field-label">{field}:</span> <div class="field-value">{row.get(field, "")}</div></div>' for field in result['config'].get('display_fields', [])])}
                </div>
                ''' for row in result['results']]) if result['results'] else '<div class="no-results">검색 결과가 없습니다.</div>'}
            </div>
        </div>
        <script>
            function selectItem(id, name) {{
                if (window.opener && !window.opener.closed) {{
                    if (window.opener.{callback}) {{
                        window.opener.{callback}(id, name);
                    }}
                    window.close();
                }} else {{
                    alert('부모 창을 찾을 수 없습니다.');
                }}
            }}
        </script>
    </body>
    </html>
    """

# 개별 라우트들을 catch-all 라우트보다 먼저 정의
@app.route("/partner-standards")
def partner_standards_route():
    """협력사 기준정보 페이지 라우트"""
    guard = enforce_permission('VENDOR_MGT', 'view')
    if guard:
        return guard
    return partner_standards()

@app.route("/partner-change-request")
def partner_change_request_route():
    """기준정보 변경요청 페이지 라우트"""
    guard = enforce_permission('REFERENCE_CHANGE', 'view')
    if guard:
        return guard
    return partner_change_request()

@app.route("/change-request-detail/<int:request_id>")
def change_request_detail_route(request_id):
    """변경요청 상세정보 페이지 라우트"""
    guard = enforce_permission('REFERENCE_CHANGE', 'view')
    if guard:
        return guard
    return change_request_detail(request_id)

@app.route("/accident")
def accident_route():
    """사고 메인 페이지 라우트"""
    guard = enforce_permission('ACCIDENT_MGT', 'view')
    if guard:
        return guard
    response = _accident_controller.list_view(request)
    record_menu_view('ACCIDENT_MGT')
    return response


def _table_has_column(conn, table_name: str, column_name: str) -> bool:
    """Return True if the given column exists on the table for the active backend."""
    cursor = conn.cursor()
    try:
        if getattr(conn, 'is_postgres', False):
            cursor.execute(
                """
                SELECT 1
                FROM information_schema.columns
                WHERE table_name = %s AND column_name = %s
                """,
                (table_name, column_name),
            )
            return cursor.fetchone() is not None
        cursor.execute(f"PRAGMA table_info({table_name})")
        for row in cursor.fetchall():
            try:
                name = row['name']
            except Exception:
                name = row[1] if len(row) > 1 else None
            if (name or '').lower() == column_name.lower():
                return True
        return False
    finally:
        try:
            cursor.close()
        except Exception:
            pass


def partner_standards():
    """협력사 기준정보 페이지"""
    guard = enforce_permission('VENDOR_MGT', 'view')
    if guard:
        return guard
    page = max(1, request.args.get('page', 1, type=int))
    per_page = request.args.get('per_page', 10, type=int)
    per_page = max(1, min(per_page, 200))

    filters = {
        'company_name': (request.args.get('company_name') or '').strip(),
        'business_number': (request.args.get('business_number') or '').strip(),
        'business_type_major': (request.args.get('business_type_major') or '').strip(),
        'business_type_minor': (request.args.get('business_type_minor') or '').strip(),
        'workers_min': request.args.get('workers_min', type=int),
        'workers_max': request.args.get('workers_max', type=int),
    }

    try:
        partner_rows, total_count = partner_manager.get_all_partners(
            page=page,
            per_page=per_page,
            filters=filters,
        )
    except Exception as exc:
        logging.error("partner_standards: partner fetch failed: %s", exc)
        partner_rows = []
        total_count = 0

    offset = (page - 1) * per_page
    partners = []
    for idx, row in enumerate(partner_rows):
        row_dict = dict(row)
        row_dict['no'] = total_count - offset - idx if total_count else len(partner_rows) - idx
        partners.append(row_dict)

    dynamic_columns = []
    conn = None
    try:
        conn = get_db_connection(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        config_rows = cursor.execute(
            """
            SELECT column_key, column_name, column_type, column_order, COALESCE(is_active, 1) AS is_active
            FROM partner_standards_column_config
            WHERE COALESCE(is_active, 1) = 1
            ORDER BY column_order, column_name
            """
        ).fetchall()
        dynamic_columns = [
            {
                'key': row['column_key'],
                'name': row['column_name'],
                'type': row.get('column_type') or 'text',
                'table_display': True,
            }
            for row in config_rows
            if row.get('column_key')
        ]
    except Exception as exc:
        logging.debug("partner_standards: column config load skipped: %s", exc)
    finally:
        if conn:
            conn.close()

    if not dynamic_columns:
        dynamic_columns = [
            {'key': 'company_name', 'name': '협력사명', 'type': 'text', 'table_display': True},
            {'key': 'business_number', 'name': '사업자번호', 'type': 'text', 'table_display': True},
            {'key': 'partner_class', 'name': '구분', 'type': 'text', 'table_display': True},
            {'key': 'business_type_major', 'name': '업종(대)', 'type': 'text', 'table_display': True},
            {'key': 'business_type_minor', 'name': '업종(소)', 'type': 'text', 'table_display': True},
            {'key': 'hazard_work_flag', 'name': '위험작업 여부', 'type': 'text', 'table_display': True},
            {'key': 'permanent_workers', 'name': '상시근로자 수', 'type': 'number', 'table_display': True},
        ]

    class Pagination:
        def __init__(self, page_num: int, per_page_num: int, total: int) -> None:
            self.page = page_num
            self.per_page = per_page_num
            self.total_count = total
            self.pages = math.ceil(total / per_page_num) if total > 0 else 1
            self.has_prev = page_num > 1
            self.prev_num = page_num - 1 if self.has_prev else None
            self.has_next = page_num < self.pages
            self.next_num = page_num + 1 if self.has_next else None

        def iter_pages(self, window_size: int = 10):
            start = ((self.page - 1) // window_size) * window_size + 1
            end = min(start + window_size - 1, self.pages)
            for num in range(start, end + 1):
                yield num

        def get_window_info(self, window_size: int = 10):
            start = ((self.page - 1) // window_size) * window_size + 1
            end = min(start + window_size - 1, self.pages)
            return {
                'start': start,
                'end': end,
                'has_prev_window': start > 1,
                'has_next_window': end < self.pages,
                'prev_window_start': max(1, start - window_size),
                'next_window_start': min(end + 1, self.pages),
            }

    pagination = Pagination(page, per_page, total_count)

    return render_template(
        'partner-standards.html',
        partners=partners,
        total_count=total_count,
        pagination=pagination,
        dynamic_columns=dynamic_columns,
        menu=MENU_CONFIG,
    )


def partner_change_request():
    """기준정보 변경요청 페이지"""
    guard = enforce_permission('REFERENCE_CHANGE', 'view')
    if guard:
        return guard
    page = max(1, request.args.get('page', 1, type=int))
    per_page = request.args.get('per_page', 10, type=int)
    per_page = max(1, min(per_page, 200))

    filters = {
        'requester_name': (request.args.get('requester_name') or '').strip(),
        'company_name': (request.args.get('company_name') or '').strip(),
        'business_number': (request.args.get('business_number') or '').strip(),
    }

    change_requests: List[Dict[str, Any]] = []
    total_count = 0

    conn = None
    try:
        conn = get_db_connection(DB_PATH)
        conn.row_factory = sqlite3.Row
        is_postgres = getattr(conn, 'is_postgres', False)

        has_is_deleted = _table_has_column(conn, 'partner_change_requests', 'is_deleted')

        where_clauses: List[str] = []
        params: List[Any] = []

        if has_is_deleted:
            if is_postgres:
                where_clauses.append("COALESCE(is_deleted::text, '0') NOT IN ('1','t','true')")
            else:
                where_clauses.append("COALESCE(is_deleted, 0) = 0")

        if filters['requester_name']:
            where_clauses.append("LOWER(COALESCE(requester_name, '')) LIKE LOWER(%s)")
            params.append(f"%{filters['requester_name']}%")

        if filters['company_name']:
            where_clauses.append("LOWER(COALESCE(company_name, '')) LIKE LOWER(%s)")
            params.append(f"%{filters['company_name']}%")

        if filters['business_number']:
            where_clauses.append("LOWER(COALESCE(business_number, '')) LIKE LOWER(%s)")
            params.append(f"%{filters['business_number']}%")

        where_clause = ''
        if where_clauses:
            where_clause = ' WHERE ' + ' AND '.join(where_clauses)

        cursor = conn.cursor()
        count_sql = f"SELECT COUNT(*) FROM partner_change_requests{where_clause}"
        count_row = cursor.execute(count_sql, params).fetchone()
        if count_row is not None:
            try:
                total_count = int(count_row[0])
            except Exception:
                try:
                    total_count = int(next(iter(count_row.values())))
                except Exception:
                    total_count = 0

        offset = (page - 1) * per_page
        if is_postgres:
            order_clause = " ORDER BY created_at DESC NULLS LAST, id DESC"
        else:
            order_clause = " ORDER BY (created_at IS NULL), created_at DESC, id DESC"

        data_sql = (
            "SELECT id, request_number, created_at, company_name, business_number, "
            "change_type, current_value, new_value, change_reason, status, requester_name, requester_department, custom_data "
            "FROM partner_change_requests"
            f"{where_clause}{order_clause} LIMIT %s OFFSET %s"
        )
        rows = cursor.execute(data_sql, params + [per_page, offset]).fetchall()

        type_codes: Dict[str, str] = {}
        status_codes: Dict[str, str] = {}
        try:
            code_service = CodeService('change_request', DB_PATH)
            for code in code_service.list('change_type') or []:
                key = code.get('option_code') or code.get('code')
                value = code.get('option_value') or code.get('value')
                if key:
                    type_codes[key] = value
            for code in code_service.list('status') or []:
                key = code.get('option_code') or code.get('code')
                value = code.get('option_value') or code.get('value')
                if key:
                    status_codes[key] = value
        except Exception as exc:
            logging.debug("partner_change_request: code lookup failed: %s", exc)

        for idx, row in enumerate(rows):
            item = dict(row)
            item['no'] = total_count - offset - idx if total_count else len(rows) - idx
            change_type_key = item.get('change_type')
            status_key = item.get('status')
            item['change_type_label'] = type_codes.get(change_type_key, change_type_key)
            item['status_label'] = status_codes.get(status_key, status_key)
            change_requests.append(item)

    except Exception as exc:
        logging.error("partner_change_request: query failed: %s", exc)
    finally:
        if conn:
            conn.close()

    class Pagination:
        def __init__(self, page_num: int, per_page_num: int, total: int) -> None:
            self.page = page_num
            self.per_page = per_page_num
            self.total_count = total
            self.pages = math.ceil(total / per_page_num) if total > 0 else 1
            self.has_prev = page_num > 1
            self.prev_num = page_num - 1 if self.has_prev else None
            self.has_next = page_num < self.pages
            self.next_num = page_num + 1 if self.has_next else None

        def iter_pages(self, window_size: int = 10):
            start = ((self.page - 1) // window_size) * window_size + 1
            end = min(start + window_size - 1, self.pages)
            for num in range(start, end + 1):
                yield num

        def get_window_info(self, window_size: int = 10):
            start = ((self.page - 1) // window_size) * window_size + 1
            end = min(start + window_size - 1, self.pages)
            return {
                'start': start,
                'end': end,
                'has_prev_window': start > 1,
                'has_next_window': end < self.pages,
                'prev_window_start': max(1, start - window_size),
                'next_window_start': min(end + 1, self.pages),
            }

    pagination = Pagination(page, per_page, total_count)

    return render_template(
        'partner-change-request.html',
        change_requests=change_requests,
        total_count=total_count,
        pagination=pagination,
        menu=MENU_CONFIG,
    )


def get_dropdown_options_for_display(board_type: str, column_key: str) -> List[Dict[str, Any]]:
    """Return dropdown options for a board/column using v2 option codes."""
    if not board_type or not column_key:
        return []

    try:
        conn = get_db_connection(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT option_code, option_value
            FROM dropdown_option_codes_v2
            WHERE board_type = %s AND column_key = %s AND COALESCE(is_active, 1) = 1
            ORDER BY display_order, option_value
            """,
            (board_type, column_key),
        )
        rows = cursor.fetchall()
    except Exception as exc:
        logging.debug(
            "get_dropdown_options_for_display: lookup failed for %s.%s: %s",
            board_type,
            column_key,
            exc,
        )
        return []
    finally:
        try:
            cursor.close()
            conn.close()
        except Exception:
            pass

    if not rows:
        return []

    if len(rows) == 1:
        value = rows[0]['option_value']
        if isinstance(value, str):
            stripped = value.strip()
            if stripped.startswith('[') and stripped.endswith(']'):
                try:
                    parsed = pyjson.loads(stripped)
                    if isinstance(parsed, list):
                        return [
                            {
                                'code': f"{column_key.upper()}_{index + 1:03d}",
                                'value': str(item),
                            }
                            for index, item in enumerate(parsed)
                        ]
                except Exception:
                    logging.debug(
                        "get_dropdown_options_for_display: array parse failed for %s.%s",
                        board_type,
                        column_key,
                    )

    options = []
    for row in rows:
        code = row['option_code']
        value = row['option_value']
        if code is None and value is None:
            continue
        options.append({'code': code, 'value': value})
    return options


@app.route("/partner/<business_number>")
@app.route("/partner-detail/<business_number>")
def partner_detail(business_number: str):
    """협력사 상세정보 페이지."""
    partner = partner_manager.get_partner_by_business_number(business_number)
    if not partner:
        logging.warning("partner_detail: partner not found for %s", business_number)
        return "협력사 정보를 찾을 수 없습니다.", 404

    attachments: List[Dict[str, Any]] = []
    try:
        conn = partner_manager.db_config.get_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT *
            FROM partner_attachments
            WHERE business_number = %s
            ORDER BY upload_date DESC
            """,
            (business_number,),
        )
        attachments = [dict(row) for row in cursor.fetchall()]
    except Exception as exc:
        logging.debug("partner_detail: attachment lookup failed: %s", exc)
    finally:
        try:
            cursor.close()
            conn.close()
        except Exception:
            pass

    is_popup = request.args.get('popup') == '1'
    return render_template(
        'partner-detail.html',
        partner=partner,
        attachments=attachments,
        menu=MENU_CONFIG,
        is_popup=is_popup,
        board_type='partner',
    )


def change_request_detail(request_id: int):
    """변경요청 상세정보 페이지."""
    record: Dict[str, Any] | None = None
    conn = None
    try:
        conn = get_db_connection(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT *
            FROM partner_change_requests
            WHERE id = %s
            """,
            (request_id,),
        )
        row = cursor.fetchone()
        if not row:
            cursor.execute(
                """
                SELECT *
                FROM change_requests
                WHERE id = %s
                """,
                (request_id,),
            )
            row = cursor.fetchone()
        if row:
            record = dict(row)
    except Exception as exc:
        logging.error("change_request_detail: primary lookup failed: %s", exc)
    finally:
        try:
            cursor.close()
            conn.close()
        except Exception:
            pass

    if not record:
        return "변경요청 정보를 찾을 수 없습니다.", 404

    detail_text = ''
    try:
        conn = get_db_connection(DB_PATH)
        cursor = conn.cursor()
        request_number = record.get('request_number')
        if request_number:
            cursor.execute(
                """
                SELECT detailed_content
                FROM change_request_details
                WHERE request_number = %s
                """,
                (request_number,),
            )
            detail_row = cursor.fetchone()
            if detail_row and 'detailed_content' in detail_row:
                detail_text = detail_row['detailed_content']
    except Exception as exc:
        logging.debug("change_request_detail: detail lookup failed: %s", exc)
    finally:
        try:
            cursor.close()
            conn.close()
        except Exception:
            pass

    dynamic_columns: List[Dict[str, Any]] = []
    try:
        conn = get_db_connection(DB_PATH)
        cursor = conn.cursor()
        rows = cursor.execute(
            """
            SELECT column_key, column_name, column_type, column_order,
                   dropdown_options, tab, column_span, linked_columns,
                   is_active, is_deleted
            FROM change_request_column_config
            ORDER BY column_order
            """
        ).fetchall()
        for row in rows:
            is_active_val = str(row['is_active'] if 'is_active' in row else '1').lower()
            is_deleted_val = str(row['is_deleted'] if 'is_deleted' in row else '0').lower()
            if is_active_val not in ('1', 'true', 't', 'y'):
                continue
            if is_deleted_val in ('1', 'true', 't', 'y'):
                continue

            column_key = row['column_key']
            column_name = row['column_name']
            column_type = (row.get('column_type') or 'text') if isinstance(row, dict) else row['column_type']
            tab_key = row.get('tab') or 'additional' if isinstance(row, dict) else row['tab'] or 'additional'

            col = {
                'column_key': column_key,
                'column_name': column_name,
                'column_type': column_type,
                'tab': tab_key,
                'column_span': row.get('column_span') if isinstance(row, dict) else row['column_span'],
                'linked_columns': row.get('linked_columns') if isinstance(row, dict) else row['linked_columns'],
                'table_display': True,
                'dropdown_options_mapped': get_dropdown_options_for_display('change_request', column_key)
                if column_type == 'dropdown'
                else [],
            }
            dynamic_columns.append(col)
    except Exception as exc:
        logging.debug("change_request_detail: dynamic column load failed: %s", exc)
    finally:
        try:
            cursor.close()
            conn.close()
        except Exception:
            pass

    custom_data = {}
    raw_custom = record.get('custom_data')
    if raw_custom:
        try:
            custom_data = raw_custom if isinstance(raw_custom, dict) else pyjson.loads(raw_custom)
        except Exception as exc:
            logging.debug("change_request_detail: custom_data parse failed: %s", exc)
            custom_data = {}

    if not detail_text:
        detail_text = custom_data.get('detailed_content', '')

    request_dict = dict(record)
    fallback_detail = detail_text or custom_data.get('detailed_content', '')
    request_dict['detailed_content'] = fallback_detail
    request_data = SimpleNamespace(**request_dict)

    section_columns: Dict[str, List[Dict[str, Any]]] = {}
    for col in dynamic_columns:
        section_columns.setdefault(col.get('tab') or 'additional', []).append(col)

    attachments: List[Dict[str, Any]] = []
    try:
        from board_services import AttachmentService

        request_number = request_data.request_number or f"CR-{request_id}"
        attachment_service = AttachmentService('change_request', DB_PATH)
        attachments = [dict(item) for item in attachment_service.list(request_number)]
    except Exception as exc:
        logging.debug("change_request_detail: attachment lookup failed: %s", exc)

    is_popup = request.args.get('popup') == '1'

    return render_template(
        'change-request-detail.html',
        request_data=request_data,
        dynamic_columns=dynamic_columns,
        custom_data=custom_data,
        attachments=attachments,
        section_columns=section_columns,
        is_popup=is_popup,
        menu=MENU_CONFIG,
    )


@app.route("/register-change-request", methods=["POST"])
def register_change_request():
    """새 변경요청 등록."""
    conn = None
    try:
        from board_services import AttachmentService
        from timezone_config import get_korean_time, get_korean_time_str

        data = pyjson.loads(request.form.get('data', '{}'))
        attachment_data = pyjson.loads(request.form.get('attachment_data', '[]'))
        files = request.files.getlist('files')
        detailed_content = data.get('detailed_content', '')

        today = get_korean_time()
        year_month = today.strftime('%Y%m')
        prefix = f"CR-{year_month}-"

        conn = get_db_connection(DB_PATH, timeout=30.0)
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT request_number
            FROM partner_change_requests
            WHERE request_number LIKE %s
            ORDER BY request_number DESC
            LIMIT 1
            """,
            (f"{prefix}%",),
        )
        last_row = cursor.fetchone()
        if last_row:
            try:
                last_seq = int(last_row[0][-2:])
            except Exception:
                last_seq = 0
        else:
            last_seq = 0
        request_number = f"{prefix}{last_seq + 1:02d}"

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS partner_change_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_number TEXT UNIQUE,
                requester_name TEXT,
                requester_department TEXT,
                company_name TEXT,
                business_number TEXT,
                change_type TEXT,
                current_value TEXT,
                new_value TEXT,
                change_reason TEXT,
                status TEXT DEFAULT 'requested',
                created_at TIMESTAMP,
                updated_at TIMESTAMP,
                custom_data TEXT,
                is_deleted INTEGER DEFAULT 0
            )
            """
        )

        data['status'] = 'requested'
        cursor.execute(
            """
            INSERT INTO partner_change_requests (
                request_number,
                requester_name,
                requester_department,
                company_name,
                business_number,
                change_reason,
                status,
                created_at,
                updated_at,
                custom_data,
                change_type,
                current_value,
                new_value
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                request_number,
                data.get('requester_name', data.get('req_name', '')),
                data.get('requester_department', data.get('req_name_dept', '')),
                data.get('company_name', data.get('compname', '')),
                data.get('business_number', data.get('compname_bizno', '')),
                data.get('change_reason', ''),
                'requested',
                get_korean_time_str(),
                get_korean_time_str(),
                pyjson.dumps(data, ensure_ascii=False),
                data.get('change_type', ''),
                data.get('current_value', ''),
                data.get('new_value', ''),
            ),
        )
        request_id = cursor.fetchone()[0]

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS change_request_details (
                request_number TEXT PRIMARY KEY,
                detailed_content TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            INSERT INTO change_request_details (request_number, detailed_content)
            VALUES (%s, %s)
            ON CONFLICT (request_number)
            DO UPDATE SET detailed_content = EXCLUDED.detailed_content,
                          updated_at = CURRENT_TIMESTAMP
            """,
            (request_number, detailed_content),
        )

        if files:
            attachment_service = AttachmentService('change_request', DB_PATH, conn)
            descriptions = [
                item.get('description', '')
                for item in attachment_data
                if isinstance(item, dict)
            ]
        for index, file in enumerate(files):
            desc = descriptions[index] if index < len(descriptions) else ''
            attachment_service.add(
                request_number,
                file,
                {
                    'description': desc,
                    'uploaded_by': session.get('user_id', 'user'),
                },
            )

        conn.commit()
        return jsonify(
            {
                'success': True,
                'request_id': request_id,
                'request_number': request_number,
                'message': '변경요청이 성공적으로 등록되었습니다.',
            }
        )
    except Exception as exc:
        logging.error("register_change_request failed: %s", exc)
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 500
    finally:
        if conn:
            conn.close()


@app.route('/update-change-request', methods=['POST'])
def update_change_request():
    """변경요청 수정."""
    from board_services import AttachmentService

    conn = None
    try:
        request_id = request.form.get('request_id')
        request_number = request.form.get('request_number')
        change_reason = request.form.get('change_reason', '')
        detailed_content = request.form.get('detailed_content', '')
        custom_data_raw = request.form.get('custom_data', '{}')
        deleted_raw = request.form.get('deleted_attachments', '[]')
        attachment_data_raw = request.form.get('attachment_data', '[]')
        files = request.files.getlist('files')

        if not request_id or not request_number:
            return jsonify({'success': False, 'message': '요청번호가 필요합니다.'}), 400

        try:
            custom_data = (
                custom_data_raw
                if isinstance(custom_data_raw, dict)
                else pyjson.loads(custom_data_raw)
            )
        except Exception:
            return jsonify({'success': False, 'message': '잘못된 데이터 형식입니다.'}), 400

        try:
            deleted_attachments = (
                deleted_raw if isinstance(deleted_raw, list) else pyjson.loads(deleted_raw)
            )
        except Exception:
            deleted_attachments = []

        try:
            attachment_meta = (
                attachment_data_raw
                if isinstance(attachment_data_raw, list)
                else pyjson.loads(attachment_data_raw)
            )
        except Exception:
            attachment_meta = []

        conn = get_db_connection(DB_PATH, timeout=30.0)
        cursor = conn.cursor()

        cursor.execute(
            "SELECT id, request_number, status FROM partner_change_requests WHERE id = %s",
            (request_id,),
        )
        existing = cursor.fetchone()
        if not existing:
            return jsonify({'success': False, 'message': '변경요청을 찾을 수 없습니다.'}), 404

        actual_request_number = existing[1] or request_number
        status_current = existing[2]

        status_value = custom_data.get('status', status_current)
        if status_value not in ('requested', 'approved', 'rejected'):
            status_value = 'requested'

        custom_data['detailed_content'] = detailed_content

        update_sql = [
            'status = %s',
            'change_reason = %s',
            'custom_data = %s',
            'updated_at = CURRENT_TIMESTAMP',
        ]
        params: List[Any] = [
            status_value,
            change_reason,
            pyjson.dumps(custom_data, ensure_ascii=False),
        ]

        for key, value in custom_data.items():
            if key in ('status', 'custom_data', 'detailed_content'):
                continue
            cursor.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'partner_change_requests' AND column_name = %s
                """,
                (key,),
            )
            if cursor.fetchone():
                update_sql.append(f"{key} = %s")
                params.append(value)

        params.append(request_id)
        cursor.execute(
            f"UPDATE partner_change_requests SET {', '.join(update_sql)} WHERE id = %s",
            params,
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS change_request_details (
                request_number TEXT PRIMARY KEY,
                detailed_content TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            INSERT INTO change_request_details (request_number, detailed_content, updated_at)
            VALUES (%s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (request_number)
            DO UPDATE SET detailed_content = EXCLUDED.detailed_content,
                          updated_at = CURRENT_TIMESTAMP
            """,
            (actual_request_number, detailed_content),
        )

        attachment_service = AttachmentService('change_request', DB_PATH, conn)
        if deleted_attachments:
            attachment_service.delete(deleted_attachments)

        for item in attachment_meta:
            attachment_id = item.get('id')
            if attachment_id and not item.get('isNew'):
                attachment_service.update_meta(
                    int(attachment_id),
                    {
                        'description': item.get('description', ''),
                        'uploaded_by': session.get('user_id', 'user'),
                    },
                )

        if files:
            new_descriptions = [
                item.get('description', '')
                for item in attachment_meta
                if item.get('isNew')
            ]
            for index, file in enumerate(files):
                desc = new_descriptions[index] if index < len(new_descriptions) else ''
                attachment_service.add(
                    actual_request_number,
                    file,
                    {
                        'description': desc,
                        'uploaded_by': session.get('user_id', 'user'),
                    },
                )

        conn.commit()
        return jsonify({'success': True, 'message': '수정이 완료되었습니다.'})
    except Exception as exc:
        logging.error('update_change_request failed: %s', exc)
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 500
    finally:
        if conn:
            conn.close()
@app.route("/accident-register")
def accident_register():
    """사고 등록 페이지 (컨트롤러 위임)"""
    guard = enforce_permission('ACCIDENT_MGT', 'write')
    if guard:
        return guard
    response = _accident_controller.register_view(request)
    record_board_action('ACCIDENT_MGT', 'VIEW', object_type='ACCIDENT', object_name='register')
    return response


@app.route("/accident-detail/<accident_id>")
def accident_detail(accident_id):
    """사고 상세 페이지 (컨트롤러 위임)"""
    guard = enforce_permission('ACCIDENT_MGT', 'view')
    if guard:
        return guard
    response = _accident_controller.detail_view(request, accident_id)
    success, _ = _response_info(response)
    record_board_action(
        'ACCIDENT_MGT',
        'VIEW',
        object_type='ACCIDENT',
        object_id=accident_id,
        success=success,
    )
    return response


@app.route("/register-accident", methods=["POST"])
def register_accident():
    """사고 신규 등록 처리"""
    guard = enforce_permission('ACCIDENT_MGT', 'write', response_type='json')
    if guard:
        return guard
    response = _accident_controller.save(request)
    success, payload = _response_info(response)
    record_board_action(
        'ACCIDENT_MGT',
        'CREATE',
        object_type='ACCIDENT',
        object_id=(payload.get('accident_number') if isinstance(payload, dict) else None),
        success=success,
        details=payload if isinstance(payload, dict) else None,
        error_message=(payload.get('message') if isinstance(payload, dict) and not success else None),
    )
    return response


@app.route("/partner-accident")
def partner_accident_route():
    """구경로 호환: /accident로 리다이렉트"""
    return redirect(url_for('accident_route'))

def partner_accident():
    """협력사 사고 목록 페이지"""
    from common_mapping import smart_apply_mappings
    import math
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    # 검색 조건
    filters = {
        'accident_date_start': request.args.get('accident_date_start'),
        'accident_date_end': request.args.get('accident_date_end'),
        'workplace': request.args.get('workplace', '').strip(),
        'accident_grade': request.args.get('accident_grade', '').strip()
    }
    
    conn = get_db_connection()
    
    # 섹션 정보 가져오기 (단일 소스: section_config via SectionConfigService)
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('accident', DB_PATH)
        sections = section_service.get_sections() or []
    except Exception as _e:
        logging.error(f"섹션 로드 실패, 기본값 사용: {_e}")
        sections = []
    
    # 동적 컬럼 설정 가져오기 (활성화되고 삭제되지 않은 것만)
    _wa3 = sql_is_active_true('is_active', conn)
    _wd3 = sql_is_deleted_false('is_deleted', conn)
    dynamic_columns_rows = conn.execute(
        f"""
        SELECT * FROM accident_column_config 
        WHERE {_wa3} AND {_wd3}
        ORDER BY column_order
        """
    ).fetchall()
    dynamic_columns = [dict(row) for row in dynamic_columns_rows]

    # 전역 키(활성/비활성 포함) 수집 - 상세 화면 팝업 타입 보정에 사용
    try:
        _wd4 = sql_is_deleted_false('is_deleted', conn)
        _all_keys_rows = conn.execute(
            f"SELECT column_key FROM accident_column_config WHERE {_wd4}"
        ).fetchall()
        all_keys = set()
        for r in _all_keys_rows:
            try:
                all_keys.add(r['column_key'])
            except Exception:
                try:
                    all_keys.add(r[0])
                except Exception:
                    pass
        all_keys = {k for k in all_keys if k}
    except Exception:
        all_keys = {c.get('column_key') for c in dynamic_columns if c.get('column_key')}
    
    # 섹션별로 컬럼 그룹핑 (detailed_content 제외)
    section_columns = {}
    for section in sections:
        section_columns[section['section_key']] = [
            col for col in dynamic_columns 
            if col.get('tab') == section['section_key'] 
            and col['column_key'] not in ['detailed_content']
        ]
    
    # 드롭다운 컬럼에 대해 코드-값 매핑 정보 추가
    for col in dynamic_columns:
        if col['column_type'] == 'dropdown':
            col['code_mapping'] = get_dropdown_options_for_display('accident', col['column_key'])
    
    # 사고 목록 조회
    _wd5 = sql_is_deleted_false('is_deleted', conn)
    query = f"""
        SELECT * FROM accidents_cache 
        WHERE {_wd5}
    """
    params = []
    
    # 필터링 적용
    if filters['accident_date_start']:
        query += " AND accident_date >= %s"
        params.append(filters['accident_date_start'])
    
    if filters['accident_date_end']:
        query += " AND accident_date <= %s"
        params.append(filters['accident_date_end'])
    
    if filters['workplace']:
        query += " AND workplace LIKE %s"
        params.append(f"%{filters['workplace']}%")
    
    if filters['accident_grade']:
        query += " AND accident_grade LIKE %s"
        params.append(f"%{filters['accident_grade']}%")
    
    # 전체 개수 조회 (ORDER BY 제거 후 COUNT)
    import re as _re
    count_query = _re.sub(r"ORDER BY[\s\S]*$", "", query, flags=_re.IGNORECASE)
    count_query = count_query.replace("SELECT *", "SELECT COUNT(*)")
    total_count = conn.execute(count_query, params).fetchone()[0]
    
    # ORDER BY는 데이터 조회시에만 추가 - created_at 기준으로 최신순 정렬
    # 정렬 컬럼(report_date)은 초기화 단계에서 보장되어야 함. 요청 중 DDL 금지.

    query += " ORDER BY (report_date IS NULL) ASC, report_date DESC, created_at DESC, accident_number DESC"
    
    # 페이지네이션 적용
    query += f" LIMIT {per_page} OFFSET {(page - 1) * per_page}"
    accidents = conn.execute(query, params).fetchall()
    accidents = [dict(row) for row in accidents]
    
    # No 컬럼 추가 (역순 번호) 및 안전 병합/표시 보정
    offset = (page - 1) * per_page
    for i, accident in enumerate(accidents):
        accident['no'] = total_count - offset - i

        # custom_data 파싱 (dict 보장)
        custom_data = {}
        if accident.get('custom_data'):
            try:
                import json as pyjson
                # PostgreSQL JSONB는 이미 dict로 반환됨
                if isinstance(accident['custom_data'], dict):
                    custom_data = accident['custom_data']
                else:
                    custom_data = pyjson.loads(accident['custom_data']) if accident['custom_data'] else {}
                accident['custom_data'] = custom_data
            except Exception as e:
                logging.error(f"custom_data 파싱 오류: {e}")
                custom_data = {}

        # 안전 병합: K사고 기본키 보호 + 빈값 미덮어쓰기
        def _is_empty(v):
            try:
                if v is None:
                    return True
                if isinstance(v, str) and v.strip() == '':
                    return True
                return False
            except Exception:
                return False

        protected_keys_for_k = {
            'accident_number','accident_name','workplace','accident_grade','major_category',
            'injury_form','injury_type','building','floor','location_category','location_detail',
            'accident_date','created_at','report_date','day_of_week',
            'responsible_company1','responsible_company1_no','responsible_company2','responsible_company2_no'
        }

        acc_no = str(accident.get('accident_number') or '')
        is_direct = acc_no.startswith('ACC')
        if custom_data:
            safe_updates = {}
            for k, v in custom_data.items():
                if _is_empty(v):
                    continue
                if k in protected_keys_for_k and not is_direct:
                    # K사고 보호
                    continue
                # 허용: 상위가 비어있거나 일반키
                if _is_empty(accident.get(k)) or k not in protected_keys_for_k:
                    safe_updates[k] = v
            if safe_updates:
                accident.update(safe_updates)

        # 등록일 표기 필드
        if acc_no.startswith('K'):
            accident['display_created_at'] = accident.get('report_date', accident.get('created_at', '-'))
        else:
            accident['display_created_at'] = accident.get('created_at', '-')

        # 사고명 최종 폴백 (상위 없으면 custom_data에서)
        if not accident.get('accident_name'):
            nm = None
            if isinstance(custom_data, dict):
                nm = custom_data.get('accident_name')
            accident['accident_name'] = (nm if (nm and str(nm).strip()) else '-')
    
    conn.close()
    
    # smart_apply_mappings 적용 (드롭다운 코드를 라벨로 변환)
    if accidents:
        from common_mapping import smart_apply_mappings
        accidents = smart_apply_mappings(accidents, 'accident', dynamic_columns, DB_PATH)
    
    # 페이지네이션 객체 생성
    import math
    class Pagination:
        def __init__(self, page, per_page, total_count):
            self.page = page
            self.per_page = per_page
            self.total_count = total_count
            self.pages = math.ceil(total_count / per_page) if total_count > 0 else 1
            self.has_prev = page > 1
            self.prev_num = page - 1 if self.has_prev else None
            self.has_next = page < self.pages
            self.next_num = page + 1 if self.has_next else None
        
        def iter_pages(self, window_size=10):
            start = ((self.page - 1) // window_size) * window_size + 1
            end = min(start + window_size - 1, self.pages)
            for num in range(start, end + 1):
                yield num
        
        def get_window_info(self, window_size=10):
            start = ((self.page - 1) // window_size) * window_size + 1
            end = min(start + window_size - 1, self.pages)
            has_prev_window = start > 1
            has_next_window = end < self.pages
            prev_window_start = max(1, start - window_size)
            next_window_start = min(end + 1, self.pages)
            return {
                'start': start,
                'end': end,
                'has_prev_window': has_prev_window,
                'has_next_window': has_next_window,
                'prev_window_start': prev_window_start,
                'next_window_start': next_window_start
            }
    
    pagination = Pagination(page, per_page, total_count)
    
    return render_template('partner-accident.html',
                         accidents=accidents,
                         total_count=total_count,
                         pagination=pagination,
                         sections=sections,
                         section_columns=section_columns,
                         dynamic_columns=dynamic_columns,
                         menu=MENU_CONFIG)

def safety_instruction_route_logic():
    """환경안전 지시서 페이지 라우트"""
    return _safety_instruction_controller.list_view(request)

def safety_instruction_register_logic():
    """환경안전 지시서 등록 페이지"""
    return _safety_instruction_controller.register_view(request)

def safety_instruction_detail_logic(issue_number):
    """환경안전 지시서 상세정보 페이지"""
    return _safety_instruction_controller.detail_view(request, issue_number)

def register_safety_instruction_logic():
    """새 환경안전 지시서 등록"""
    return _safety_instruction_controller.save(request)

def update_safety_instruction_logic():
    """환경안전 지시서 수정"""
    return _safety_instruction_controller.update(request)


@app.route("/verify-password", methods=["POST"])
def verify_password():
    """게시판별 비밀번호 검증"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "No data received"}), 400
            
        password = data.get('password')
        board_type = data.get('board_type', 'default')  # partner, accident, 또는 default
        
        if not password:
            return jsonify({"success": False, "message": "Password not provided"}), 400
        
        # 게시판 타입별 비밀번호 확인
        correct_password = None
        
        if board_type == 'partner':
            # 협력사 게시판 비밀번호
            correct_password = db_config.config.get('PASSWORDS', 'PARTNER_EDIT_PASSWORD', fallback=None)
        elif board_type == 'accident':
            # 사고 게시판 비밀번호
            correct_password = db_config.config.get('PASSWORDS', 'ACCIDENT_EDIT_PASSWORD', fallback=None)
        else:
            # 기본 비밀번호 (기존 호환성)
            correct_password = db_config.config.get('DEFAULT', 'EDIT_PASSWORD')
        
        # 비밀번호가 설정되지 않은 경우 기본 비밀번호 사용
        if not correct_password:
            correct_password = db_config.config.get('DEFAULT', 'EDIT_PASSWORD')
        
        logging.info(f"비밀번호 검증 요청: board_type={board_type}")
        
        if password == correct_password:
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "message": "비밀번호가 올바르지 않습니다."})
    except Exception as e:
        logging.error(f"비밀번호 검증 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/update-partner", methods=["POST"])
def update_partner():
    """협력사 정보 업데이트"""
    conn = None
    try:
        # json already imported globally
        
        business_number = request.form.get('business_number')
        detailed_content = request.form.get('detailed_content')
        
        # 안전하게 JSON 파싱
        try:
            deleted_attachments = pyjson.loads(request.form.get('deleted_attachments', '[]'))
        except:
            deleted_attachments = []
        
        try:
            attachment_data_raw = request.form.get('attachment_data', '[]')
            if isinstance(attachment_data_raw, str):
                attachment_data = pyjson.loads(attachment_data_raw)
            else:
                attachment_data = attachment_data_raw
            # 리스트가 아닌 경우 빈 리스트로
            if not isinstance(attachment_data, list):
                attachment_data = []
        except Exception as e:
            logging.warning(f"attachment_data 파싱 실패: {e}")
            attachment_data = []
        
        files = request.files.getlist('files')
        
        print(f"Business Number: {business_number}")
        print(f"Files count: {len(files)}")
        print(f"Attachment data: {attachment_data}")
        
        # 협력사 존재 여부 확인 (먼저 확인)
        partner = partner_manager.get_partner_by_business_number(business_number)
        if not partner:
            from flask import jsonify
            return jsonify({"success": False, "message": "협력사를 찾을 수 없습니다."})
        
        print(f"Connecting to database (unified backend)")
        conn = get_db_connection(timeout=30.0)
        cursor = conn.cursor()
        
        logging.info(f"업데이트 대상 협력사: {business_number}")
        
        # 1. 협력사 상세내용 업데이트 (partner_details 테이블)
        logging.info(f"상세내용 업데이트: {detailed_content[:50]}...")
        # partner_details: business_number를 고유키로 사용하도록 인덱스 보장 (PostgreSQL ON CONFLICT 요구)
        try:
            cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_partner_details_business_number ON partner_details (business_number)")
        except Exception as _e_idx:
            logging.debug(f"Ensure unique index skipped/failed: {_e_idx}")
        # partner_details safe_upsert 사용
        detail_data = {
            'business_number': business_number,
            'detailed_content': detailed_content,
            'updated_at': None  # 자동으로 처리됨
        }
        safe_upsert(conn, 'partner_details', detail_data)
        logging.info("상세내용 업데이트 완료")
        
        # 2. 삭제된 첨부파일 처리
        for attachment_id in deleted_attachments:
            cursor.execute("DELETE FROM partner_attachments WHERE id = %s", (attachment_id,))
        
        # 3. 기존 첨부파일 정보 업데이트
        for attachment in attachment_data:
            # attachment가 딕셔너리인지 확인
            if isinstance(attachment, dict):
                if attachment.get('id') and not attachment.get('isNew'):
                    cursor.execute("""
                        UPDATE partner_attachments 
                        SET description = %s 
                        WHERE id = %s
                    """, (attachment.get('description', ''), attachment['id']))
            else:
                logging.warning(f"attachment가 딕셔너리가 아님: {type(attachment)}")
        
        # 4. 새 파일 업로드 처리
        import os
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
            
        # 새 파일들과 새 첨부파일 데이터 매칭
        new_attachments = [a for a in attachment_data if isinstance(a, dict) and a.get('isNew')]
        print(f"New attachments: {new_attachments}")
        
        for i, file in enumerate(files):
            if file and file.filename and i < len(new_attachments):
                # 파일명에 타임스탬프 추가하여 중복 방지
                import time
                original_filename = sanitize_filename(file.filename, fallback_prefix='upload')
                timestamp = str(int(time.time()))
                name, ext = os.path.splitext(original_filename)
                unique_filename = f"{name}_{timestamp}{ext}"
                file_path = os.path.join(upload_folder, unique_filename)

                print(f"Saving file: {original_filename} as {unique_filename}")
                file.save(file_path)

                attachment_info = new_attachments[i]
                cursor.execute("""
                    INSERT INTO partner_attachments 
                    (business_number, file_name, file_path, file_size, description)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    business_number,
                    original_filename,
                    file_path,
                    os.path.getsize(file_path),
                    attachment_info.get('description', '')
                ))
                logging.info(f"첨부파일 추가: {original_filename} - {attachment_info.get('description', '')}")
        
        # 커밋 전 확인
        check_result = cursor.execute("SELECT COUNT(*) FROM partner_attachments WHERE business_number = %s", (business_number,)).fetchone()
        logging.info(f"커밋 전 {business_number} 협력사 첨부파일 개수: {check_result[0]}개")
        
        try:
            conn.commit()
            logging.info("데이터베이스 커밋 성공")
            
            # 커밋 후 다시 확인
            check_result2 = cursor.execute("SELECT COUNT(*) FROM partner_attachments WHERE business_number = %s", (business_number,)).fetchone()
            logging.info(f"커밋 후 {business_number} 협력사 첨부파일 개수: {check_result2[0]}개")
            
            conn.close()
            
            # 새로운 연결로 다시 확인
            logging.info("새 연결로 데이터 지속성 확인...")
            verify_conn = get_db_connection()
            verify_result = verify_conn.execute("SELECT COUNT(*) FROM partner_attachments WHERE business_number = %s", (business_number,)).fetchone()
            logging.info(f"새 연결 확인: {business_number} 협력사 첨부파일 개수: {verify_result[0]}개")
            verify_conn.close()
            
            from flask import jsonify
            return jsonify({"success": True})
        except Exception as commit_error:
            print(f"Commit failed: {commit_error}")
            conn.rollback()
            conn.close()
            from flask import jsonify
            return jsonify({"success": False, "message": f"Commit failed: {str(commit_error)}"})
        
    except Exception as e:
        if conn:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        from flask import jsonify
        logging.error(f"업데이트 중 오류 발생: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500

# Legacy update function retained for reference (no longer routed)
def _legacy_update_accident():
    """사고 정보 업데이트"""
    conn = None
    try:
        # json already imported globally
        
        accident_number = request.form.get('accident_number')
        
        # K 사고도 수정 가능하되, 기본정보는 업데이트하지 않고 custom_data/첨부만 반영
        
        detailed_content = request.form.get('detailed_content', '')
        custom_data = request.form.get('custom_data', '{}')  # Phase 2: 동적 컬럼 데이터
        base_fields = request.form.get('base_fields', '{}')  # ACC 사고일 때 기본정보
        
        # 상세내용 디버깅
        # detailed_content만 사용하도록 통일
        final_content = detailed_content
        
        # 안전하게 JSON 파싱
        try:
            deleted_attachments = pyjson.loads(request.form.get('deleted_attachments', '[]'))
        except:
            deleted_attachments = []
        
        try:
            attachment_data_raw = request.form.get('attachment_data', '[]')
            if isinstance(attachment_data_raw, str):
                attachment_data = pyjson.loads(attachment_data_raw)
            else:
                attachment_data = attachment_data_raw
            # 리스트가 아닌 경우 빈 리스트로
            if not isinstance(attachment_data, list):
                attachment_data = []
        except Exception as e:
            logging.warning(f"attachment_data 파싱 실패: {e}")
            attachment_data = []
        files = request.files.getlist('files')
        
        print(f"=== UPDATE_ACCIDENT FULL DEBUG START ===")
        print(f"Accident Number: {accident_number}")
        print(f"Custom Data received: {custom_data}")
        print(f"Custom Data type: {type(custom_data)}")
        print(f"Detailed Content received: '{detailed_content}'")
        print(f"Detailed Content length: {len(detailed_content) if detailed_content else 0}")

        print(f"Request form keys: {list(request.form.keys())}")
        print(f"Attachment data received: {attachment_data}")
        print(f"Files received: {len(files)} files")
        
        # 모든 form 데이터 출력
        print(f"=== ALL FORM DATA ===")
        for key in request.form.keys():
            value = request.form.get(key)
            if key == 'custom_data':
                print(f"  {key}: {value[:100]}... (length: {len(value)})")
            else:
                print(f"  {key}: '{value}'")
        
        # custom_data 파싱
        if isinstance(custom_data, str):
            try:
                custom_data = pyjson.loads(custom_data)
                print(f"Custom Data parsed: {custom_data}")
                print(f"Custom Data keys: {list(custom_data.keys())}")
                if 'injured_person' in custom_data:
                    injured_person = custom_data['injured_person']
                    print(f"injured_person type: {type(injured_person)}, value: {injured_person}")
            except Exception as e:
                print(f"Custom Data parsing failed: {e}")
                custom_data = {}
        
        # 리스트 필드 정규화 (이중 JSON 인코딩 방지)
        for key, value in custom_data.items():
            if isinstance(value, str) and value.startswith('[') and value.endswith(']'):
                try:
                    # 이미 JSON 문자열인 리스트 필드를 배열로 변환
                    custom_data[key] = pyjson.loads(value)
                    print(f"List field {key} normalized from string to array")
                except:
                    pass
        
        print(f"Files count: {len(files)}")
        print(f"Attachment data: {attachment_data}")
        
        print(f"Connecting to database: {DB_PATH}")
        conn = get_db_connection(timeout=30.0)
        cursor = conn.cursor()
        
        # 사고번호가 없으면 자동 생성 (수기입력용)
        if not accident_number:
            accident_number = generate_manual_accident_number(cursor)
            logging.info(f"자동 생성된 사고번호: {accident_number}")
        
        # 사고 형식 검증 (K로 시작하는 외부시스템 사고 또는 ACC로 시작하는 수기입력 사고)
        if not (accident_number.startswith('K') or accident_number.startswith('ACC')):
            from flask import jsonify
            return jsonify({"success": False, "message": "잘못된 사고번호 형식입니다."})
        
        logging.info(f"업데이트 대상 사고: {accident_number}")
        
        # 기본정보 업데이트 정책
        # - K사고: 기본정보 미수정 (custom_data/첨부만 반영)
        # - ACC사고: 기본정보 수정 허용하되, 빈값('' 또는 None)은 덮어쓰지 않음
        is_direct_entry = accident_number.startswith('ACC')
        if is_direct_entry and base_fields != '{}':
            try:
                base_data = pyjson.loads(base_fields)
                logging.info(f"ACC 사고 기본정보 업데이트 요청: {base_data}")

                # 허용된 필드만 업데이트 (SQL injection 방지)
                allowed_fields = [
                    'accident_name', 'accident_date', 'workplace', 'accident_grade',
                    'major_category', 'injury_form', 'injury_type', 'building', 'floor',
                    'location_category', 'location_detail', 'created_at', 'day_of_week'
                ]

                def _is_empty(v):
                    try:
                        if v is None:
                            return True
                        if isinstance(v, str) and v.strip() == '':
                            return True
                        return False
                    except Exception:
                        return False

                # 업데이트할 필드와 값 준비 (빈값은 스킵)
                update_fields = []
                update_values = []
                for field in allowed_fields:
                    if field in base_data:
                        value = base_data[field]
                        if _is_empty(value):
                            # 빈값은 기존 데이터를 덮어쓰지 않음
                            continue
                        update_fields.append(f"{field} = %s")
                        update_values.append(value)

                if update_fields:
                    update_values.append(accident_number)  # WHERE 절용
                    update_query = f"""
                        UPDATE accidents_cache 
                        SET {', '.join(update_fields)}
                        WHERE accident_number = %s
                    """
                    cursor.execute(update_query, update_values)
                    logging.info(f"ACC 사고 기본정보 업데이트 완료: {len(update_fields)}개 필드")
                else:
                    logging.info("ACC 사고 기본정보 변경 없음(모든 값이 빈값/미포함)")
            except Exception as e:
                logging.error(f"ACC 사고 기본정보 업데이트 중 오류: {e}")
        
        # 먼저 해당 사고가 accidents_cache에 있는지 확인
        cursor.execute("SELECT id FROM accidents_cache WHERE accident_number = %s", (accident_number,))
        accident_row = cursor.fetchone()
        
        if accident_row:
            # 기존 custom_data를 먼저 가져와서 병합
            cursor.execute("SELECT custom_data FROM accidents_cache WHERE accident_number = %s", (accident_number,))
            existing_row = cursor.fetchone()
            existing_custom_data = {}
            
            if existing_row and existing_row[0]:
                if isinstance(existing_row[0], dict):
                    # PostgreSQL JSONB - 깊은 복사로 안전하게 처리
                    existing_custom_data = pyjson.loads(pyjson.dumps(existing_row[0]))
                elif isinstance(existing_row[0], str):
                    # SQLite JSON 문자열
                    try:
                        existing_custom_data = pyjson.loads(existing_row[0])
                    except:
                        existing_custom_data = {}
            
            
            # 특별히 injured_person은 덮어쓰지 않고 보존
            if 'injured_person' in existing_custom_data and 'injured_person' not in custom_data:
                # injured_person이 새 데이터에 없으면 기존 것 보존
                pass
            elif 'injured_person' in custom_data:
                # injured_person이 새 데이터에 있으면 그것 사용
                pass
            
            # K사고 보호 및 안전 병합 규칙
            # - K사고: IQADB(원본) 기본 필드는 custom_data로 덮지 않음
            # - 공통: 빈값('' 또는 None)은 기존 값을 덮지 않음
            protected_keys_for_k = {
                'accident_number','accident_name','workplace','accident_grade','major_category',
                'injury_form','injury_type','building','floor','location_category','location_detail',
                'accident_date','created_at','report_date','day_of_week',
                # 확장: 책임회사/번호(원본 보호)
                'responsible_company1','responsible_company1_no','responsible_company2','responsible_company2_no'
            }

            # 주의: 화면 렌더 단계에서 안전 병합을 적용하므로 custom_data 키를 삭제하지 않는다

            def _is_empty_value(v):
                try:
                    if v is None:
                        return True
                    if isinstance(v, str) and v.strip() == '':
                        return True
                    return False
                except Exception:
                    return False

            # 안전한 병합: 리스트 타입 필드는 병합 처리
            def is_list_field(field_value):
                """리스트 필드인지 확인"""
                if isinstance(field_value, list):
                    return True
                if isinstance(field_value, str) and field_value.strip():
                    return field_value.startswith('[') and field_value.endswith(']')
                return False

            # 사용자 편집 가능한 섹션 키를 조회하여 '덮어쓰기 허용' 집합 구성
            additional_keys = set()
            try:
                _tab_rows = cursor.execute(
                    "SELECT column_key, tab FROM accident_column_config"
                ).fetchall()
                # 시스템 보호 탭 vs 사용자 편집 가능 탭 구분
                protected_tabs = {'basic_info', 'accident_info', 'location_info'}

                for _r in _tab_rows:
                    try:
                        k = _r['column_key'] if hasattr(_r, 'keys') else _r[0]
                        t = _r['tab'] if hasattr(_r, 'keys') else (_r[1] if len(_r) > 1 else None)
                    except Exception:
                        k, t = None, None
                    # 시스템 보호 탭이 아닌 모든 탭의 컬럼을 덮어쓰기 허용
                    if k and t and t not in protected_tabs:
                        additional_keys.add(k)
            except Exception:
                pass

            # 명시적 우선 키(요청된 필드들): 추가정보와 동일하게 덮어쓰기 허용
            override_always = {
                'company_1cha','company_1cha_bizno',
                'accident_company','accident_company_bizno',
                'accident_time'
            }
            overwrite_keys = additional_keys | override_always
            
            for key, value in custom_data.items():
                # 리스트 필드 병합 우선 처리
                if is_list_field(value) or is_list_field(existing_custom_data.get(key)):
                    
                    # 기존 데이터 파싱
                    existing_list = existing_custom_data.get(key, [])
                    if isinstance(existing_list, str) and existing_list.strip():
                        try:
                            existing_list = pyjson.loads(existing_list)
                        except:
                            existing_list = []
                    if not isinstance(existing_list, list):
                        existing_list = []
                    
                    # 새 데이터 파싱
                    new_list = []
                    if isinstance(value, list):
                        new_list = value
                    elif isinstance(value, str) and value.strip():
                        if value.startswith('[') and value.endswith(']'):
                            try:
                                new_list = pyjson.loads(value)
                                if not isinstance(new_list, list):
                                    new_list = []
                            except:
                                new_list = []
                    
                    
                    # 추가정보/오버라이드 키는 우선 대체하되, 완전 빈 배열이면 기존값 보존
                    if key in overwrite_keys:
                        if len(new_list) == 0 and len(existing_list) > 0:
                            existing_custom_data[key] = existing_list
                        else:
                            existing_custom_data[key] = new_list
                    # 그 외: 프론트에서 전체 배열을 보냈다면 그대로 사용, 아니면 병합
                    elif len(new_list) > 0 and len(existing_list) > 0:
                        # 새로운 데이터에 기존 데이터의 첫 번째 항목이 포함되어 있다면 전체 교체로 간주
                        first_existing_id = existing_list[0].get('id', '') if existing_list and isinstance(existing_list[0], dict) else ''
                        has_existing_data = any(
                            isinstance(item, dict) and item.get('id') == first_existing_id 
                            for item in new_list
                        ) if first_existing_id else False
                        
                        if has_existing_data:
                            # 전체 데이터를 보냈으므로 그대로 사용
                            existing_custom_data[key] = new_list
                        else:
                            # 새 항목만 추가하므로 병합 처리
                            merged_list = list(existing_list)
                            existing_ids = {item.get('id', '') for item in existing_list if isinstance(item, dict)}
                            
                            for new_item in new_list:
                                if isinstance(new_item, dict) and new_item.get('id') not in existing_ids:
                                    merged_list.append(new_item)
                                    existing_ids.add(new_item.get('id'))
                            
                            existing_custom_data[key] = merged_list
                    else:
                        # 하나가 비어있으면 비어있지 않은 것을 사용 (일반 키)
                        existing_custom_data[key] = new_list if len(new_list) > 0 else existing_list
                else:
                    # 일반 필드
                    # 값이 비어있으면 공통적으로 보존
                    if _is_empty_value(value):
                        continue
                    existing_custom_data[key] = value
            
            # detailed_content를 custom_data에 항상 반영(빈 문자열 포함)
            existing_custom_data['detailed_content'] = final_content
            logging.info(f"detailed_content를 custom_data에 반영: {len(final_content) if final_content else 0}자")

            # Details 테이블 저장
            try:
                from db.upsert import safe_upsert as _su
                _su(conn, 'accident_details', {
                    'accident_number': accident_number,
                    'detailed_content': final_content,
                    'updated_at': None
                })
                logging.info("[ACC] details upsert applied")
            except Exception as _e_det:
                logging.warning(f"[ACC] details upsert warning: {_e_det}")
            
            # custom_data 업데이트 (JSONB 캐스팅 제거)
            cursor.execute("""
                UPDATE accidents_cache 
                SET custom_data = %s
                WHERE accident_number = %s
            """, (existing_custom_data, accident_number))
            logging.info(f"custom_data 업데이트 완료: {accident_number}")
        else:
            # custom_data를 JSON 문자열로 변환
            if not isinstance(custom_data, str):
                custom_data_str = pyjson.dumps(custom_data)
            else:
                custom_data_str = custom_data
            
            # detailed_content를 custom_data에 항상 반영(빈 문자열 포함)
            custom_data_dict = pyjson.loads(custom_data_str) if isinstance(custom_data_str, str) else (custom_data_str or {})
            if not isinstance(custom_data_dict, dict):
                custom_data_dict = {}
            custom_data_dict['detailed_content'] = final_content
            custom_data_str = pyjson.dumps(custom_data_dict)
            
            # FormData에서 직접 받은 필드들 사용 (하드코딩 제거)
            # 새 레코드 생성 (업체 정보는 선택적)
            # 비공식/직접등록 사고
            korean_date = get_korean_time().strftime('%Y-%m-%d')
            cursor.execute("""
                INSERT INTO accidents_cache (
                    accident_number, accident_name, custom_data, accident_date,
                    workplace, accident_grade, major_category, injury_form, 
                    injury_type, created_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                accident_number, 
                request.form.get('accident_name', f"사고_{accident_number}"),
                custom_data_str, 
                request.form.get('accident_date', korean_date) or None,
                request.form.get('workplace', ''),
                request.form.get('accident_grade', ''),
                request.form.get('major_category', ''),
                request.form.get('injury_form', ''),
                request.form.get('injury_type', ''),
                korean_date
            ))
            logging.info(f"새 사고 레코드 생성 (직접등록) 및 동적 컬럼 데이터 저장: {accident_number}")
        
        # 2. 사고 첨부파일 처리 - item_id 사용
        
        # 3. 삭제된 첨부파일 처리
        for attachment_id in deleted_attachments:
            cursor.execute("DELETE FROM accident_attachments WHERE id = %s", (attachment_id,))
        
        # 4. 기존 첨부파일 정보 업데이트
        for attachment in attachment_data:
            # attachment가 딕셔너리인지 확인
            if isinstance(attachment, dict):
                print(f"[ATTACH DEBUG] 처리할 첨부파일: {attachment}")
                
                # 기존 첨부파일 업데이트 (id가 있는 경우)
                if attachment.get('id') and not attachment.get('isNew'):
                    print(f"[ATTACH DEBUG] 기존 첨부파일 업데이트: ID={attachment['id']}")
                    cursor.execute("""
                        UPDATE accident_attachments 
                        SET description = %s 
                        WHERE id = %s
                    """, (attachment.get('description', ''), attachment['id']))
                
                # id가 None이지만 isNew가 False인 경우 - 기존 파일인데 ID가 없는 상황
                elif attachment.get('id') is None and not attachment.get('isNew', True):
                    print(f"[ATTACH DEBUG] ID가 None인 기존 첨부파일 - accident_number로 찾아서 업데이트")
                    # 이 사고에 속한 첨부파일 중에서 description을 업데이트
                    cursor.execute("""
                        SELECT id FROM accident_attachments 
                        WHERE accident_number = %s 
                        ORDER BY id ASC
                    """, (accident_number,))
                    existing_ids = [row[0] for row in cursor.fetchall()]
                    print(f"[ATTACH DEBUG] 기존 첨부파일 IDs: {existing_ids}")
                    
                    if existing_ids:
                        # 첫 번째 첨부파일의 description을 업데이트 (간단한 매칭)
                        cursor.execute("""
                            UPDATE accident_attachments 
                            SET description = %s 
                            WHERE id = %s
                        """, (attachment.get('description', ''), existing_ids[0]))
                        print(f"[ATTACH DEBUG] 첨부파일 {existing_ids[0]}의 설명을 '{attachment.get('description', '')}' 로 업데이트")
            else:
                logging.warning(f"attachment가 딕셔너리가 아님: {type(attachment)}")
        
        # 5. 새 파일 업로드 처리
        import os
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
            
        # 새 파일들과 새 첨부파일 데이터 매칭
        new_attachments = [a for a in attachment_data if isinstance(a, dict) and a.get('isNew')]
        print(f"New attachments: {new_attachments}")
        
        for i, file in enumerate(files):
            if file and file.filename and i < len(new_attachments):
                # 파일명에 타임스탬프 추가하여 중복 방지
                import time
                original_filename = sanitize_filename(file.filename, fallback_prefix='upload')
                timestamp = str(int(time.time()))
                name, ext = os.path.splitext(original_filename)
                unique_filename = f"{name}_{timestamp}{ext}"
                file_path = os.path.join(upload_folder, unique_filename)

                print(f"Saving file: {original_filename} as {unique_filename}")
                file.save(file_path)

                attachment_info = new_attachments[i]
                cursor.execute("""
                    INSERT INTO accident_attachments 
                    (accident_number, file_name, file_path, file_size, description)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    accident_number,
                    original_filename,
                    file_path,
                    os.path.getsize(file_path),
                    attachment_info.get('description', '')
                ))
                logging.info(f"첨부파일 추가: {original_filename} - {attachment_info.get('description', '')}")
        
        # 커밋 전 확인
        check_result = cursor.execute("SELECT COUNT(*) FROM accident_attachments WHERE accident_number = %s", (accident_number,)).fetchone()
        logging.info(f"커밋 전 {accident_number} 사고 첨부파일 개수: {check_result[0]}개")
        
        try:
            conn.commit()
            logging.info("데이터베이스 커밋 성공")
            
            # 커밋 후 다시 확인
            check_result2 = cursor.execute("SELECT COUNT(*) FROM accident_attachments WHERE accident_number = %s", (accident_number,)).fetchone()
            logging.info(f"커밋 후 {accident_number} 사고 첨부파일 개수: {check_result2[0]}개")
            
            conn.close()
            
            # 새로운 연결로 다시 확인
            logging.info("새 연결로 데이터 지속성 확인...")
            verify_conn = get_db_connection()
            verify_result = verify_conn.execute("SELECT COUNT(*) FROM accident_attachments WHERE accident_number = %s", (accident_number,)).fetchone()
            logging.info(f"새 연결 확인: {accident_number} 사고 첨부파일 개수: {verify_result[0]}개")
            verify_conn.close()
            
            from flask import jsonify
            return jsonify({"success": True})
        except Exception as commit_error:
            print(f"Commit failed: {commit_error}")
            conn.rollback()
            conn.close()
            from flask import jsonify
            return jsonify({"success": False, "message": f"Commit failed: {str(commit_error)}"})
        
    except Exception as e:
        if conn:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        from flask import jsonify
        logging.error(f"사고 업데이트 중 오류 발생: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500



@app.route("/update-accident", methods=["POST"])
def update_accident():
    guard = enforce_permission('ACCIDENT_MGT', 'write', response_type='json')
    if guard:
        return guard
    response = _accident_controller.update(request)
    success, payload = _response_info(response)
    record_board_action(
        'ACCIDENT_MGT',
        'UPDATE',
        object_type='ACCIDENT',
        object_id=(payload.get('accident_number') if isinstance(payload, dict) else None),
        success=success,
        details=payload if isinstance(payload, dict) else None,
        error_message=(payload.get('message') if isinstance(payload, dict) and not success else None),
    )
    return response


def _resolve_storage_path(file_path: str) -> Optional[str]:
    """DB에 저장된 파일 경로를 실제 파일 시스템 경로로 매핑한다."""
    if not file_path:
        return None

    raw_path = str(file_path).strip().strip('\"\'')
    if not raw_path:
        return None

    import os
    import re
    from flask import current_app

    candidates: List[str] = []

    def add_candidate(path: Optional[str]) -> None:
        if not path:
            return
        normalized = path
        if normalized not in candidates:
            candidates.append(normalized)

    # 1. 원본 경로 그대로
    add_candidate(raw_path)

    # 2. OS 기본 구분자로 정규화한 경로
    normalized_sep = raw_path.replace('\\', os.sep)
    add_candidate(normalized_sep)

    # 3. 절대 경로라면 그대로 검사
    if os.path.isabs(normalized_sep):
        add_candidate(os.path.normpath(normalized_sep))

    # 4. 현재 작업 디렉터리 기준 상대 경로
    rel_fragment = normalized_sep.lstrip('/\\')
    add_candidate(os.path.join(os.getcwd(), rel_fragment))

    # 5. Flask 애플리케이션 루트 기준 상대 경로
    try:
        app_root = current_app.root_path
    except Exception:
        app_root = None
    if app_root:
        add_candidate(os.path.join(app_root, rel_fragment))

    basename = os.path.basename(os.path.normpath(normalized_sep))
    potential_dirs: List[str] = []

    default_upload = os.path.join(os.getcwd(), 'uploads')
    potential_dirs.append(default_upload)

    if app_root:
        potential_dirs.append(os.path.join(app_root, 'uploads'))

    try:
        cfg_upload = current_app.config.get('UPLOAD_FOLDER')
    except Exception:
        cfg_upload = None
    if cfg_upload:
        if not os.path.isabs(cfg_upload):
            cfg_path = os.path.join(app_root or os.getcwd(), cfg_upload)
        else:
            cfg_path = cfg_upload
        potential_dirs.append(cfg_path)

    for folder in potential_dirs:
        if not folder:
            continue
        normalized_folder = os.path.normpath(folder)
        if basename:
            add_candidate(os.path.join(normalized_folder, basename))

    # 6. Windows 드라이브 경로(C:\..) → WSL 경로(/mnt/c/..)
    drive_match = re.match(r'^[A-Za-z]:[\\/]', raw_path)
    if drive_match:
        drive_letter = raw_path[0].lower()
        remainder = raw_path[2:].lstrip('\\/')
        remainder_for_wsl = remainder.replace('\\', '/')
        wsl_candidate = '/mnt/{}/{}'.format(drive_letter, remainder_for_wsl)
        add_candidate(wsl_candidate)

    # 7. WSL 경로(/mnt/c/..) → Windows 경로(C:\..)
    if raw_path.startswith('/mnt/'):
        parts = raw_path.split('/')
        if len(parts) > 3:
            drive_letter = parts[2]
            remainder_windows = '\\'.join(parts[3:])
            windows_candidate = '{}:\\{}'.format(drive_letter.upper(), remainder_windows)
            add_candidate(windows_candidate)

    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate

    logging.error('첨부파일 경로를 찾을 수 없습니다: raw=%s candidates=%s', raw_path, candidates)
    return None


def _send_attachment_response(file_path: str, original_name: Optional[str], mime_type: Optional[str] = None):
    """첨부파일을 안전하게 전송하는 공통 헬퍼"""
    from flask import send_file
    import os
    from urllib.parse import quote

    actual_file_path = _resolve_storage_path(file_path)
    if not actual_file_path:
        logging.error(f"파일을 찾을 수 없습니다: {file_path}")
        return "File not found on disk", 404

    original_name = original_name or os.path.basename(actual_file_path) or 'download'
    safe_name = sanitize_filename(original_name, fallback_prefix='download')
    if not safe_name:
        safe_name = 'download.file'

    encoded_name = quote(safe_name)
    ascii_fallback = safe_name.encode('ascii', 'ignore').decode('ascii', 'ignore')
    if not ascii_fallback:
        ascii_fallback = 'download.file'

    logging.debug(
        "[download] sending attachment path=%s cwd=%s safe_name=%s",
        actual_file_path,
        os.getcwd(),
        safe_name
    )

    response = send_file(
        actual_file_path,
        mimetype=mime_type or 'application/octet-stream',
        as_attachment=True,
        download_name=safe_name
    )
    response.headers['Content-Disposition'] = (
        f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded_name}"
    )
    return response


def _get_attachment_info(board_type: str, attachment_id: int, conn=None) -> Optional[Dict[str, Any]]:
    """보드 타입 별 첨부파일 정보를 반환"""
    normalized = board_type.replace('-', '_').lower()

    if normalized == 'partner':
        close_conn = False
        if conn is None:
            conn = get_db_connection()
            close_conn = True
        cursor = conn.cursor()
        cursor.execute(
            "SELECT file_path, file_name, COALESCE(mime_type, '') FROM partner_attachments WHERE id = %s",
            (attachment_id,)
        )
        row = cursor.fetchone()
        if close_conn:
            conn.close()
        if row:
            try:
                file_path = row['file_path']
                file_name = row['file_name']
                mime_type = row.get('mime_type') if hasattr(row, 'get') else row['mime_type']
            except (TypeError, KeyError):
                file_path = row[0] if len(row) > 0 else None
                file_name = row[1] if len(row) > 1 else None
                mime_type = row[2] if len(row) > 2 else None
            return {'path': file_path, 'name': file_name, 'mime_type': mime_type or None}
        return None

    from board_services import AttachmentService

    valid_boards = AttachmentService.ID_COLUMN_MAP.keys()
    if normalized not in valid_boards:
        return None

    service = AttachmentService(normalized, DB_PATH, conn)
    return service.download(attachment_id)


@app.route("/download/<board_type>/<int:attachment_id>")
def download_attachment_for_board(board_type, attachment_id):
    """보드 타입별 첨부파일 다운로드 엔드포인트"""
    info = _get_attachment_info(board_type, attachment_id)
    if not info:
        return "File not found", 404

    logging.info(
        "다운로드 요청: board=%s, id=%s, file=%s",
        board_type,
        attachment_id,
        info.get('name')
    )

    return _send_attachment_response(info['path'], info.get('name'), info.get('mime_type'))


@app.route("/download/<int:attachment_id>")
def download_attachment(attachment_id):
    """레거시 첨부파일 다운로드 - 모든 게시판을 순회 탐색"""
    conn = get_db_connection()

    try:
        info = _get_attachment_info('partner', attachment_id, conn)
        if info:
            logging.info("다운로드 요청(legacy): board=partner, id=%s, file=%s", attachment_id, info.get('name'))
            return _send_attachment_response(info['path'], info.get('name'), info.get('mime_type'))

        from board_services import AttachmentService

        for board in AttachmentService.ID_COLUMN_MAP.keys():
            info = _get_attachment_info(board, attachment_id, conn)
            if info:
                logging.info(
                    "다운로드 요청(legacy): board=%s, id=%s, file=%s",
                    board,
                    attachment_id,
                    info.get('name')
                )
                return _send_attachment_response(info['path'], info.get('name'), info.get('mime_type'))
    finally:
        conn.close()

    return "File not found", 404


@app.route("/partner-attachments/<business_number>")
def get_partner_attachments(business_number):
    """협력사 첨부파일 목록 가져오기"""
    conn = get_db_connection()
    attachments = conn.execute("""
        SELECT * FROM partner_attachments 
        WHERE business_number = %s 
        ORDER BY upload_date DESC
    """, (business_number,)).fetchall()
    conn.close()
    
    # 첨부파일 목록을 딕셔너리로 변환
    result = [dict(attachment) for attachment in attachments]
    
    from flask import jsonify
    return jsonify(result)

@app.route("/api/auto-upload-partner-files", methods=['POST'])
def auto_upload_partner_files():
    """협력사 사업자번호에 대한 HTML 파일 자동 생성 및 업로드"""
    conn = None
    cursor = None
    try:
        data = request.get_json() or {}
        business_number = (data.get('business_number') or '').strip()
        file_paths = data.get('file_paths') or []
        description_input = (data.get('description') or '').strip()

        if not business_number:
            return jsonify({"error": "business_number is required"}), 400
        if not file_paths:
            return jsonify({"error": "file_paths is required"}), 400

        # 협력사 정보 확인
        partner = partner_manager.get_partner_by_business_number(business_number)
        if not partner:
            return jsonify({"error": f"Partner not found: {business_number}"}), 404

        # 업로드 폴더 (현재 작업 디렉토리 기준 상대경로)
        upload_folder = Path(os.getcwd()) / "uploads"
        upload_folder.mkdir(parents=True, exist_ok=True)

        uploaded_files = []
        skipped = []  # 업로드 실패한 파일 추적
        conn = get_db_connection()
        cursor = conn.cursor()

        from timezone_config import get_korean_time, get_korean_time_str

        korean_time = get_korean_time()
        year = korean_time.strftime("%Y")
        month = korean_time.strftime("%m")
        company_name = (partner.get('company_name') if isinstance(partner, dict) else None) or '협력사'

        default_suffix = '통합레포트'
        if description_input:
            base_name = description_input
        else:
            base_name = f"{company_name}_{year}년_{month}월_{default_suffix}"
        base_name = re.sub(r'[<>:"/\|?*%]', '_', str(base_name)).strip() or default_suffix
        description_value = base_name

        display_input = data.get('file_name') or data.get('display_name')
        display_list = None
        display_single = None
        if isinstance(display_input, list):
            display_list = [re.sub(r'[<>:"/\|?*%]', '_', str(item)).strip() for item in display_input]
        elif display_input:
            display_single = re.sub(r'[<>:"/\|?*%]', '_', str(display_input)).strip()

        # 기존 파일 삭제 ( 동일 description )
        deleted_count = 0
        try:
            existing_files = cursor.execute(
                """
                SELECT file_path FROM partner_attachments
                WHERE business_number = %s AND description = %s
                """,
                (business_number, description_value)
            ).fetchall()

            for file_row in existing_files:
                old_file_path = Path(file_row['file_path'])
                if old_file_path.exists():
                    try:
                        old_file_path.unlink()
                        logging.info(f"Deleted old file: {old_file_path}")
                    except Exception as exc:
                        logging.warning(f"Failed to delete old file {old_file_path}: {exc}")

            cursor.execute(
                """
                DELETE FROM partner_attachments
                WHERE business_number = %s AND description = %s
                """,
                (business_number, description_value)
            )
            conn.commit()
            deleted_count = len(existing_files)
        except Exception as exc:
            logging.error(f"Error deleting old files: {exc}")
            conn.rollback()
            deleted_count = 0

        for idx, file_path in enumerate(file_paths):
            try:
                file_path = Path(file_path).expanduser().resolve()
                if not file_path.exists():
                    logging.warning(f"File not found: {file_path}")
                    skipped.append(str(file_path))
                    continue

                original_name = file_path.name
                safe_name = re.sub(r'[<>:"/\|?*%]', '_', original_name) if original_name else "file"
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                new_filename = f"{business_number}_{timestamp}_{safe_name}"
                dest_path = upload_folder / new_filename

                shutil.copy2(file_path, dest_path)

                if display_list:
                    try:
                        idx = file_paths.index(str(file_path))
                    except ValueError:
                        idx = 0
                    base_candidate = display_list[idx] if idx < len(display_list) else display_list[-1]
                elif display_single:
                    base_candidate = display_single
                else:
                    base_candidate = description_value

                base_candidate = re.sub(r'[<>:"/\|?*%]', '_', str(base_candidate)).strip() or description_value

                original_ext = Path(base_candidate).suffix
                fallback_ext = Path(safe_name).suffix or '.html'
                ext = original_ext or fallback_ext

                if original_ext and base_candidate.lower().endswith(original_ext.lower()):
                    base_without_ext = base_candidate[:-len(original_ext)]
                else:
                    base_without_ext = base_candidate

                base_without_ext = base_without_ext.rstrip('.') or description_value
                display_name = f"{base_without_ext}{ext}"

                cursor.execute(
                    """
                    INSERT INTO partner_attachments
                    (business_number, file_name, file_path, file_size, upload_date, description)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        business_number,
                        display_name,
                        str(dest_path),
                        dest_path.stat().st_size,
                        get_korean_time_str(),
                        description_value,
                    ),
                )

                uploaded_files.append({
                    "original_path": str(file_path),
                    "uploaded_filename": new_filename,
                    "display_name": display_name,
                    "file_size": dest_path.stat().st_size,
                })

                logging.info(f"File uploaded: {file_path.name} → {new_filename}")

            except Exception as exc:
                logging.error(f"Error processing file {file_path}: {exc}")
                skipped.append(str(file_path))
                continue

        conn.commit()

        status = 200 if uploaded_files and not skipped else (207 if uploaded_files and skipped else 400)

        return jsonify({
            "success": bool(uploaded_files),
            "business_number": business_number,
            "description": description_value,
            "uploaded_files": uploaded_files,
            "skipped": skipped,
            "deleted_count": deleted_count,
            "total_uploaded": len(uploaded_files),
            "total_skipped": len(skipped),
            "message": f"기존 {deleted_count}개 파일 삭제 후 {len(uploaded_files)}개 새 파일 업로드"
        }), status

    except Exception as exc:
        logging.error(f"Error in auto_upload_partner_files: {exc}")
        return jsonify({"error": str(exc)}), 500

    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@app.route("/api/accident-columns", methods=["GET"])
def get_accident_columns():
    """사고 페이지 동적 컬럼 설정 조회"""
    try:
        column_service = ColumnConfigService('accident', DB_PATH)
        columns = column_service.list_columns()
        return jsonify(columns)
    except Exception as e:
        logging.error(f"컬럼 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-columns", methods=["POST"])
def add_accident_column():
    """사고 페이지 동적 컬럼 추가"""
    try:
        column_service = ColumnConfigService('accident', DB_PATH)
        result = column_service.add_column(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"컬럼 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-columns/<int:column_id>", methods=["PUT"])
def update_accident_column(column_id):
    """사고 페이지 동적 컬럼 수정"""
    try:
        column_service = ColumnConfigService('accident', DB_PATH)
        result = column_service.update_column(column_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"컬럼 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-columns/<int:column_id>", methods=["DELETE"])
def delete_accident_column(column_id):
    """사고 페이지 동적 컬럼 삭제 (비활성화)"""
    try:
        column_service = ColumnConfigService('accident', DB_PATH)
        result = column_service.delete_column(column_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"컬럼 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-columns/order", methods=["PUT"])
def update_accident_columns_order():
    """사고 페이지 동적 컬럼 순서 변경"""
    try:
        column_service = ColumnConfigService('accident', DB_PATH)
        result = column_service.update_columns_order(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"컬럼 순서 변경 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# 관리자 인증 라우트
@app.route("/admin/login", methods=["POST"])
def admin_login():
    """관리자 비밀번호 인증 처리"""
    from urllib.parse import urlparse

    password = request.form.get('password')
    redirect_url = request.form.get('redirect_url') or '/admin/menu-settings'

    parsed = urlparse(redirect_url)
    safe_redirect = parsed.path or '/admin/menu-settings'
    if parsed.query:
        safe_redirect = f"{safe_redirect}?{parsed.query}"

    if password == ADMIN_PASSWORD:
        session['admin_authenticated'] = True
        return redirect(safe_redirect)
    else:
        return render_template(
            'admin-login.html',
            error='비밀번호가 틀렸습니다.',
            redirect_url=safe_redirect,
            menu=MENU_CONFIG
        )

@app.route("/admin/logout")
def admin_logout():
    """관리자 로그아웃"""
    session.pop('admin_authenticated', None)
    return redirect(url_for('index'))

# ======================================================================
# Admin 동기화 관리 엔드포인트
# ======================================================================

@app.route('/admin/sync-now', methods=['POST','GET'])
def admin_sync_now():
    """수동 강제 동기화 엔드포인트"""
    try:
        from database_config import maybe_daily_sync_master, maybe_one_time_sync_content

        # 간단한 토큰 보호 (선택): config.ini [ADMIN] SYNC_TOKEN과 헤더 'X-Sync-Token' 매칭 시에만 허용
        try:
            cfg = configparser.ConfigParser()
            cfg.read('config.ini', encoding='utf-8')
            expected = cfg.get('ADMIN', 'SYNC_TOKEN', fallback='')
        except Exception:
            expected = ''
        provided = request.headers.get('X-Sync-Token', '')
        if expected:
            if provided != expected:
                return jsonify({'success': False, 'message': 'Unauthorized (sync token required)'}), 401

        # 디버그 로깅(요청 메타)
        try:
            print(f"[SYNC] method={request.method} ct={request.headers.get('Content-Type')} host={request.host}")
        except Exception:
            pass

        # type 파라미터는 JSON/폼/쿼리스트링 모두 허용
        sync_type = None
        if request.is_json:
            try:
                sync_type = (request.get_json(silent=True) or {}).get('type')
            except Exception:
                sync_type = None
        if not sync_type:
            sync_type = request.form.get('type') or request.args.get('type') or 'all'
        
        if sync_type == 'master' or sync_type == 'all':
            maybe_daily_sync_master(force=True)
        
        if sync_type == 'content' or sync_type == 'all':
            maybe_one_time_sync_content(force=True)
            
        return jsonify({'success': True, 'message': f'Sync completed ({sync_type})'})
    except Exception as e:
        logging.error(f"Manual sync failed: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/cache-counts', methods=['GET'])
def admin_cache_counts():
    """캐시 테이블 레코드 수 확인 엔드포인트"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        tables = ['partners_cache','accidents_cache','safety_instructions',
                 'departments_cache','buildings_cache','contractors_cache','employees_cache']
        counts = {}
        for t in tables:
            try:
                cur.execute(f"SELECT COUNT(*) FROM {t}")
                counts[t] = cur.fetchone()[0]
            except Exception:
                counts[t] = 'table not found'
        conn.close()
        return jsonify({'success': True, 'counts': counts})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# 이 라우트는 아래에 더 완전한 버전이 있으므로 제거됨

@app.route("/admin/accident-columns")
@require_admin_auth
def admin_accident_columns():
    """사고 컬럼 관리 페이지 (표준 경로) - simplified 템플릿 사용"""
    # 기존 simplified 구현을 표준 경로에서 사용
    conn = get_db_connection()
    _wa = sql_is_active_true('is_active', conn)
    _wd = sql_is_deleted_false('is_deleted', conn)
    sections = conn.execute(
        f"SELECT * FROM section_config WHERE board_type = 'accident' AND {_wa} AND {_wd} ORDER BY section_order"
    ).fetchall()
    sections = [dict(row) for row in sections]
    # 컬럼 관리 페이지에서는 사고 데이터 매핑이 필요 없음 (불필요 코드 제거)
    conn.close()
    return render_template('admin-accident-columns.html', sections=sections, menu=MENU_CONFIG)

@app.route("/admin/accident-columns-simplified")
@require_admin_auth
def admin_accident_columns_simplified():
    """구 경로 호환: 표준 경로로 리다이렉트"""
    return redirect(url_for('admin_accident_columns'))

@app.route("/admin/accident-basic-codes")
@require_admin_auth
def admin_accident_basic_codes():
    """사고 기본정보 코드 관리 페이지"""
    return render_template('admin-accident-basic-codes.html', menu=MENU_CONFIG)

@app.route("/admin/accident-codes")
@require_admin_auth
def admin_accident_codes():
    """사고 코드 관리 임베디드 페이지"""
    column_key = request.args.get('column_key', '')
    embedded = request.args.get('embedded', 'false') == 'true'
    
    return render_template('admin-accident-codes.html', 
                         column_key=column_key,
                         embedded=embedded,
                         menu=MENU_CONFIG)

# ===== 사고 데이터 캐시 이관 (accidents -> accidents_cache) =====
@app.route('/admin/migrate-accidents-to-cache', methods=['POST'])
@require_admin_auth
def migrate_accidents_to_cache():
    """외부 accidents 테이블 데이터를 accidents_cache로 일괄 업서트 이관

    - 기준: accident_number(UNIQUE)
    - 이미 존재하면 업데이트, 없으면 신규 생성
    - 소스의 잔여 필드는 custom_data로 병합
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # accidents 테이블 존재 여부 확인
        try:
            # PostgreSQL: information_schema를 통해 컬럼 정보 조회
            cur.execute("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'accidents'
            """)
            cols = cur.fetchall()
            if not cols:
                return jsonify({'success': False, 'message': 'accidents 테이블이 없습니다.'}), 404
            acc_columns = [c[1] for c in cols]
        except Exception:
            try:
                cur.execute(
                    "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
                    ('accidents',)
                )
                acc_columns = [r[0] if not isinstance(r, dict) else r['column_name'] for r in cur.fetchall()]
                if not acc_columns:
                    return jsonify({'success': False, 'message': 'accidents 테이블이 없습니다.'}), 404
            except Exception:
                return jsonify({'success': False, 'message': 'accidents 테이블 확인 실패'}), 500

        # cache 필수 컬럼 보강
        try:
            # PostgreSQL: information_schema를 통해 컬럼 정보 조회
            cur.execute("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'accidents_cache'
            """)
            cache_cols = [c[0] for c in cur.fetchall()]
            ensure_cols = [
                ('accident_number','TEXT'),('accident_name','TEXT'),('workplace','TEXT'),
                ('accident_grade','TEXT'),('major_category','TEXT'),('injury_form','TEXT'),('injury_type','TEXT'),
                ('accident_date','TEXT'),('day_of_week','TEXT'),('report_date','TEXT'),('created_at','TEXT'),
                ('building','TEXT'),('floor','TEXT'),('location_category','TEXT'),('location_detail','TEXT'),
                ('custom_data','TEXT'),('is_deleted','INTEGER')
            ]
            for cn, ct in ensure_cols:
                if cn not in cache_cols:
                    try:
                        cur.execute(f"ALTER TABLE accidents_cache ADD COLUMN {cn} {ct}")
                    except Exception:
                        pass
        except Exception:
            pass

        # 소스 데이터 조회 (삭제 제외)
        where_notdel = sql_is_deleted_false('is_deleted', conn) if 'is_deleted' in acc_columns else '1=1'
        src_rows = conn.execute(f"SELECT * FROM accidents WHERE {where_notdel}").fetchall()

        migrated = 0
        updated = 0
        skipped = 0

        from db.upsert import safe_upsert
        for r in src_rows:
            row = dict(r)
            acc_no = row.get('accident_number')
            if not acc_no:
                skipped += 1
                continue

            top_keys = ['accident_number','accident_name','workplace','accident_grade','major_category',
                        'injury_form','injury_type','accident_date','day_of_week','report_date','created_at',
                        'building','floor','location_category','location_detail']
            data = {k: row.get(k) for k in top_keys if k in row}
            for dk in ('accident_date','report_date','created_at'):
                if dk in data and data[dk] is not None:
                    data[dk] = str(data[dk])

            # custom_data 병합
            src_cd = row.get('custom_data')
            if isinstance(src_cd, str):
                try:
                    import json as _json
                    src_cd = _json.loads(src_cd) if src_cd else {}
                except Exception:
                    src_cd = {}
            elif not isinstance(src_cd, dict):
                src_cd = {}
            leftovers = {k: v for k, v in row.items() if (k not in data and k not in ('custom_data','id','is_deleted'))}
            merged_cd = {}
            merged_cd.update(src_cd)
            merged_cd.update(leftovers)
            data['custom_data'] = merged_cd
            if 'is_deleted' in row:
                data['is_deleted'] = row['is_deleted']

            try:
                exist = conn.execute("SELECT 1 FROM accidents_cache WHERE accident_number = %s", (acc_no,)).fetchone()
                safe_upsert(conn, 'accidents_cache', data, conflict_cols=['accident_number'])
                if exist:
                    updated += 1
                else:
                    migrated += 1
            except Exception as _e:
                logging.error(f"업서트 실패: {acc_no}: {_e}")
                skipped += 1

        conn.commit(); conn.close()
        return jsonify({'success': True, 'migrated': migrated, 'updated': updated, 'skipped': skipped, 'total_source': len(src_rows)})
    except Exception as e:
        logging.error(f"사고 캐시 이관 실패: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route("/admin/safety-instruction-codes")
@require_admin_auth
def admin_safety_instruction_codes():
    """환경안전 지시서 코드 관리 페이지"""
    return render_template('admin-safety-instruction-codes.html', menu=MENU_CONFIG)

@app.route("/admin/full-process-codes")
@require_admin_auth
def admin_full_process_codes():
    """Full Process 코드 관리 임베디드 페이지"""
    column_key = request.args.get('column_key', '')
    embedded = request.args.get('embedded', 'false') == 'true'
    
    return render_template('admin-full-process-codes.html', 
                         column_key=column_key,
                         embedded=embedded,
                         menu=MENU_CONFIG)

@app.route("/admin/follow-sop-codes")
@require_admin_auth
def admin_follow_sop_codes():
    """Follow SOP 코드 관리 임베디드 페이지"""
    column_key = request.args.get('column_key', '')
    embedded = request.args.get('embedded', 'false') == 'true'

    return render_template('admin-follow-sop-codes.html', 
                         column_key=column_key,
                         embedded=embedded,
                         menu=MENU_CONFIG)


@app.route("/admin/safe-workplace-codes")
@require_admin_auth
def admin_safe_workplace_codes():
    """안전한 일터 코드 관리 임베디드 페이지"""
    column_key = request.args.get('column_key', '')
    embedded = request.args.get('embedded', 'false') == 'true'

    return render_template('admin-safe-workplace-codes.html',
                         column_key=column_key,
                         embedded=embedded,
                         menu=MENU_CONFIG)

@app.route("/admin/person-master")
@require_admin_auth
def admin_person_master():
    """담당자 마스터 관리 페이지"""
    return render_template('admin-person-master.html', menu=MENU_CONFIG)

# Follow SOP API 엔드포인트
@app.route("/api/follow-sop-columns", methods=["GET"])
def get_follow_sop_columns():
    """Follow SOP 페이지 동적 컬럼 설정 조회"""
    try:
        column_service = ColumnConfigService('follow_sop', DB_PATH)
        columns = column_service.list_columns()
        return jsonify(columns)
    except Exception as e:
        logging.error(f"Follow SOP 컬럼 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/follow-sop-columns", methods=["POST"])
def add_follow_sop_column():
    """Follow SOP 페이지 동적 컬럼 추가"""
    try:
        if not request.json:
            return jsonify({"success": False, "message": "JSON data required"}), 400
            
        column_service = ColumnConfigService('follow_sop', DB_PATH)
        result = column_service.add_column(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Follow SOP 컬럼 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/follow-sop-columns/<int:column_id>", methods=["PUT"])
def update_follow_sop_column(column_id):
    """Follow SOP 페이지 동적 컬럼 수정"""
    try:
        column_service = ColumnConfigService('follow_sop', DB_PATH)
        result = column_service.update_column(column_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Follow SOP 컬럼 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/follow-sop-columns/<int:column_id>", methods=["DELETE"])
def delete_follow_sop_column(column_id):
    """Follow SOP 페이지 동적 컬럼 삭제 (비활성화)"""
    try:
        column_service = ColumnConfigService('follow_sop', DB_PATH)
        result = column_service.delete_column(column_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Follow SOP 컬럼 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# Follow SOP Sections API 엔드포인트
@app.route("/api/follow-sop-sections", methods=["GET"])
def get_follow_sop_sections():
    """Follow SOP 섹션 목록 조회"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('follow_sop', DB_PATH)
        sections = section_service.get_sections()
        return jsonify({"success": True, "sections": sections})
    except Exception as e:
        logging.error(f"Follow SOP 섹션 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/follow-sop-sections", methods=["POST"])
def add_follow_sop_section():
    """Follow SOP 섹션 추가"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('follow_sop', DB_PATH)
        result = section_service.add_section(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Follow SOP 섹션 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/follow-sop-sections/<int:section_id>", methods=["PUT"])
def update_follow_sop_section(section_id):
    """Follow SOP 섹션 수정"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('follow_sop', DB_PATH)
        result = section_service.update_section(section_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Follow SOP 섹션 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/follow-sop-sections/<int:section_id>", methods=["DELETE"])
def delete_follow_sop_section(section_id):
    """Follow SOP 섹션 삭제"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('follow_sop', DB_PATH)
        result = section_service.delete_section(section_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Follow SOP 섹션 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/follow-sop-sections/reorder", methods=["POST"])
def reorder_follow_sop_sections():
    """Follow SOP 섹션 순서 변경"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('follow_sop', DB_PATH)
        result = section_service.reorder_sections(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Follow SOP 섹션 순서 변경 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# Safe Workplace API 엔드포인트
@app.route("/api/safe-workplace-columns", methods=["GET"])
def get_safe_workplace_columns():
    """Safe Workplace 페이지 동적 컬럼 설정 조회"""
    try:
        column_service = ColumnConfigService('safe_workplace', DB_PATH)
        columns = column_service.list_columns()
        return jsonify(columns)
    except Exception as e:
        logging.error(f"Safe Workplace 컬럼 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-columns", methods=["POST"])
def add_safe_workplace_column():
    """Safe Workplace 페이지 동적 컬럼 추가"""
    try:
        if not request.json:
            return jsonify({"success": False, "message": "JSON data required"}), 400

        column_service = ColumnConfigService('safe_workplace', DB_PATH)
        result = column_service.add_column(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Safe Workplace 컬럼 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-columns/<int:column_id>", methods=["PUT"])
def update_safe_workplace_column(column_id):
    """Safe Workplace 페이지 동적 컬럼 수정"""
    try:
        column_service = ColumnConfigService('safe_workplace', DB_PATH)
        result = column_service.update_column(column_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Safe Workplace 컬럼 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-columns/<int:column_id>", methods=["DELETE"])
def delete_safe_workplace_column(column_id):
    """Safe Workplace 페이지 동적 컬럼 삭제"""
    try:
        column_service = ColumnConfigService('safe_workplace', DB_PATH)
        result = column_service.delete_column(column_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Safe Workplace 컬럼 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-sections", methods=["GET"])
def get_safe_workplace_sections():
    """Safe Workplace 섹션 목록 조회"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safe_workplace', DB_PATH)
        sections = section_service.get_sections()
        return jsonify({"success": True, "sections": sections})
    except Exception as e:
        logging.error(f"Safe Workplace 섹션 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-sections", methods=["POST"])
def add_safe_workplace_section():
    """Safe Workplace 섹션 추가"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safe_workplace', DB_PATH)
        result = section_service.add_section(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Safe Workplace 섹션 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-sections/<int:section_id>", methods=["PUT"])
def update_safe_workplace_section(section_id):
    """Safe Workplace 섹션 수정"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safe_workplace', DB_PATH)
        result = section_service.update_section(section_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Safe Workplace 섹션 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-sections/<int:section_id>", methods=["DELETE"])
def delete_safe_workplace_section(section_id):
    """Safe Workplace 섹션 삭제"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safe_workplace', DB_PATH)
        result = section_service.delete_section(section_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Safe Workplace 섹션 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace-sections/reorder", methods=["POST"])
def reorder_safe_workplace_sections():
    """Safe Workplace 섹션 순서 변경"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safe_workplace', DB_PATH)
        result = section_service.reorder_sections(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Safe Workplace 섹션 순서 변경 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# Full Process API 엔드포인트
@app.route("/api/full-process-columns", methods=["GET"])
def get_full_process_columns():
    """Full Process 페이지 동적 컬럼 설정 조회"""
    try:
        column_service = ColumnConfigService('full_process', DB_PATH)
        columns = column_service.list_columns()
        return jsonify(columns)
    except Exception as e:
        logging.error(f"Full Process 컬럼 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process-columns", methods=["POST"])
def add_full_process_column():
    """Full Process 페이지 동적 컬럼 추가"""
    try:
        column_service = ColumnConfigService('full_process', DB_PATH)
        result = column_service.add_column(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Full Process 컬럼 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process-columns/<int:column_id>", methods=["PUT"])
def update_full_process_column(column_id):
    """Full Process 페이지 동적 컬럼 수정"""
    try:
        column_service = ColumnConfigService('full_process', DB_PATH)
        result = column_service.update_column(column_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Full Process 컬럼 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process-columns/<int:column_id>", methods=["DELETE"])
def delete_full_process_column(column_id):
    """Full Process 페이지 동적 컬럼 삭제 (비활성화)"""
    try:
        column_service = ColumnConfigService('full_process', DB_PATH)
        result = column_service.delete_column(column_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Full Process 컬럼 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# Full Process Sections API 엔드포인트
@app.route("/api/full-process-sections", methods=["GET"])
def get_full_process_sections():
    """Full Process 섹션 목록 조회"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('full_process', DB_PATH)
        sections = section_service.get_sections()
        return jsonify({"success": True, "sections": sections})
    except Exception as e:
        logging.error(f"Full Process 섹션 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process-sections", methods=["POST"])
def add_full_process_section():
    """Full Process 섹션 추가"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('full_process', DB_PATH)
        result = section_service.add_section(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Full Process 섹션 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process-sections/<int:section_id>", methods=["PUT"])
def update_full_process_section(section_id):
    """Full Process 섹션 수정"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('full_process', DB_PATH)
        result = section_service.update_section(section_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Full Process 섹션 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process-sections/<int:section_id>", methods=["DELETE"])
def delete_full_process_section(section_id):
    """Full Process 섹션 삭제"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('full_process', DB_PATH)
        result = section_service.delete_section(section_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Full Process 섹션 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process-sections/reorder", methods=["POST"])
def reorder_full_process_sections():
    """Full Process 섹션 순서 변경"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('full_process', DB_PATH)
        result = section_service.reorder_sections(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"Full Process 섹션 순서 변경 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/admin/safety-instruction-columns")
@require_admin_auth  
def admin_safety_instruction_columns():
    """환경안전 지시서 컬럼 관리 페이지"""
    # 섹션 정보 로드
    from section_service import SectionConfigService
    section_service = SectionConfigService('safety_instruction', DB_PATH)
    sections = section_service.get_sections()
    
    return render_template('admin-safety-instruction-columns.html', 
                         menu=MENU_CONFIG,
                         sections=sections)

@app.route("/admin/follow-sop-columns")
@require_admin_auth
def admin_follow_sop_columns():
    """Follow SOP 컬럼 관리 페이지"""
    # 섹션 정보 로드
    from section_service import SectionConfigService
    section_service = SectionConfigService('follow_sop', DB_PATH)
    section_columns = section_service.get_sections_with_columns()
    
    return render_template('admin-follow-sop-columns.html',
                         menu=MENU_CONFIG,
                         section_columns=section_columns,
                         sections=section_columns)  # 하위 호환성

@app.route("/admin/fullprocess-columns")
def redirect_fullprocess_columns():
    """구식 URL 리다이렉션"""
    return redirect("/admin/full-process-columns", code=301)

@app.route("/admin/followsop-columns")
def redirect_followsop_columns():
    """구식 URL 리다이렉션"""
    return redirect("/admin/follow-sop-columns", code=301)

@app.route("/admin/full-process-columns")
@require_admin_auth
def admin_full_process_columns():
    """Full Process 컬럼 관리 페이지"""
    # 섹션 정보 로드
    from section_service import SectionConfigService
    section_service = SectionConfigService('full_process', DB_PATH)
    section_columns = section_service.get_sections_with_columns()
    
    return render_template('admin-full-process-columns.html',
                         menu=MENU_CONFIG,
                         section_columns=section_columns,
                         sections=section_columns)  # 하위 호환성


@app.route("/admin/safe-workplace-columns")
@require_admin_auth
def admin_safe_workplace_columns():
    """Safe Workplace 컬럼 관리 페이지"""
    from section_service import SectionConfigService
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        default_sections = [
            ('basic_info', '기본정보', 1),
            ('workplace_info', '작업장정보', 2),
            ('safety_info', '안전정보', 3)
        ]

        for key, name, order in default_sections:
            cursor.execute(
                """
                INSERT INTO safe_workplace_sections (section_key, section_name, section_order, is_active, is_deleted)
                VALUES (%s, %s, %s, 1, 0)
                ON CONFLICT (section_key) DO NOTHING
                """,
                (key, name, order)
            )

            cursor.execute(
                """
                UPDATE safe_workplace_sections
                SET section_name = %s,
                    section_order = %s,
                    is_active = 1
                WHERE section_key = %s
                  AND (is_deleted = 0 OR is_deleted IS NULL)
                """,
                (name, order, key)
            )

        default_columns = [
            ('safeplace_no', '점검번호', 'text', 1, 'basic_info', 2, 1),
            ('created_at', '등록일', 'datetime', 2, 'basic_info', 2, 1)
        ]

        for col_key, col_name, col_type, order, tab, span, is_active in default_columns:
            cursor.execute(
                """
                INSERT INTO safe_workplace_column_config
                    (column_key, column_name, column_type, column_order, tab, column_span, is_active)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (column_key) DO UPDATE
                SET column_name = EXCLUDED.column_name,
                    column_type = EXCLUDED.column_type,
                    column_order = EXCLUDED.column_order,
                    tab = EXCLUDED.tab,
                    column_span = EXCLUDED.column_span,
                    is_active = EXCLUDED.is_active
                """,
                (col_key, col_name, col_type, order, tab, span, is_active)
            )

        conn.commit()

    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"Safe Workplace 기본 섹션/컬럼 초기화 실패: {e}")
    finally:
        if conn:
            conn.close()

    section_service = SectionConfigService('safe_workplace', DB_PATH)
    section_columns = section_service.get_sections_with_columns()

    return render_template(
        'admin-safe-workplace-columns.html',
        menu=MENU_CONFIG,
        section_columns=section_columns,
        sections=section_columns
    )

@app.route("/admin/safety-instruction-columns-simplified")
@require_admin_auth
def admin_safety_instruction_columns_simplified():
    """환경안전 지시서 컬럼 관리 페이지 Simplified - 간소화 버전"""
    return render_template('admin-safety-instruction-columns-simplified.html', menu=MENU_CONFIG)

@app.route("/admin/change-request-columns")
@require_admin_auth
def admin_change_request_columns():
    """기준정보 변경요청 컬럼 관리 페이지"""
    return render_template('admin-change-request-columns.html', menu=MENU_CONFIG)

# 중복 라우트 제거됨 - 위에서 이미 처리

@app.route("/admin/change-request-columns-simplified")
@require_admin_auth
def admin_change_request_columns_simplified():
    """기준정보 변경요청 컬럼 관리 페이지 Simplified - 간소화 버전"""
    return render_template('admin-change-request-columns-simplified.html', menu=MENU_CONFIG)

# ===== 기준정보 변경요청 컬럼 관리 API =====

@app.route("/api/change-request/columns", methods=["GET"])
def get_change_request_columns():
    """기준정보 변경요청 페이지 동적 컬럼 설정 조회"""
    try:
        column_service = ColumnConfigService('change_request', DB_PATH)
        columns = column_service.list_columns()
        return jsonify(columns)
    except Exception as e:
        logging.error(f"변경요청 컬럼 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request/columns", methods=["POST"])
def add_change_request_column():
    """기준정보 변경요청 페이지 동적 컬럼 추가"""
    try:
        column_service = ColumnConfigService('change_request', DB_PATH)
        result = column_service.add_column(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"변경요청 컬럼 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request/columns/<int:column_id>", methods=["PUT"])
def update_change_request_column(column_id):
    """기준정보 변경요청 페이지 동적 컬럼 수정"""
    try:
        column_service = ColumnConfigService('change_request', DB_PATH)
        result = column_service.update_column(column_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"변경요청 컬럼 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request/columns/<int:column_id>", methods=["DELETE"])
def delete_change_request_column(column_id):
    """기준정보 변경요청 페이지 동적 컬럼 삭제 (비활성화)"""
    try:
        column_service = ColumnConfigService('change_request', DB_PATH)
        result = column_service.delete_column(column_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"변경요청 컬럼 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request/columns/order", methods=["PUT"])
def update_change_request_columns_order():
    """변경요청 페이지 동적 컬럼 순서 변경"""
    try:
        column_service = ColumnConfigService('change_request', DB_PATH)
        result = column_service.update_columns_order(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"컬럼 순서 변경 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request/dropdown-codes", methods=["GET"])
def get_change_request_dropdown_codes():
    column_key = request.args.get('column_key')
    if not column_key:
        return jsonify({"success": False, "message": "column_key is required"}), 400
    """특정 컬럼의 드롭다운 코드 조회 (v2 통일)"""
    try:
        conn = get_db_connection()
        
        # v2 테이블 사용
        codes = conn.execute("""
            SELECT option_code as code, option_value as value, display_order, is_active
            FROM dropdown_option_codes_v2
            WHERE board_type = 'change_request' AND column_key = %s AND is_active = 1
            ORDER BY display_order, id
        """, (column_key,)).fetchall()
        
        conn.close()
        
        return jsonify({
            "success": True,
            "codes": [dict(code) for code in codes],
            "column_key": column_key
        })
    except Exception as e:
        logging.error(f"변경요청 드롭다운 코드 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request/dropdown-codes", methods=["POST"])
def save_change_request_dropdown_codes():
    """드롭다운 코드 일괄 저장"""
    try:
        column_key = request.args.get('column_key')
        codes = request.json  # 바디는 배열만
        
        if not column_key:
            return jsonify({"success": False, "message": "column_key is required"}), 400
        
        if not isinstance(codes, list):
            return jsonify({"success": False, "message": "Body must be an array"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 기존 코드 비활성화 (v2 테이블)
        cursor.execute("""
            UPDATE dropdown_option_codes_v2
            SET is_active = 0, updated_at = CURRENT_TIMESTAMP
            WHERE board_type = 'change_request' AND column_key = %s
        """, (column_key,))
        
        # 새 코드 삽입 또는 업데이트 (v2 테이블)
        for idx, code_data in enumerate(codes):
            cursor.execute("""
                INSERT INTO dropdown_option_codes_v2
                (board_type, column_key, option_code, option_value, display_order, is_active)
                VALUES ('change_request', %s, %s, %s, %s, 1)
                ON CONFLICT(board_type, column_key, option_code) DO UPDATE SET
                    option_value = excluded.option_value,
                    display_order = excluded.display_order,
                    is_active = 1,
                    updated_at = CURRENT_TIMESTAMP
            """, (column_key, code_data['code'], code_data['value'], idx))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "코드가 저장되었습니다."})
    except Exception as e:
        logging.error(f"변경요청 드롭다운 코드 저장 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request-dropdown-codes/<int:code_id>", methods=["DELETE"])
def delete_change_request_dropdown_code(code_id):
    """드롭다운 코드 삭제 (v2 통일)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE dropdown_option_codes_v2
            SET is_active = 0, updated_at = CURRENT_TIMESTAMP 
            WHERE id = %s
        """, (code_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "코드가 삭제되었습니다."})
    except Exception as e:
        logging.error(f"변경요청 드롭다운 코드 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/change-request/save", methods=["POST"])
def save_change_request():
    """변경요청 저장"""
    try:
        data = request.json
        request_number = data.get('request_number')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 테이블이 없으면 생성
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS change_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_number TEXT UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # 동적 컬럼들을 위한 ALTER TABLE
        for key, value in data.items():
            if key != 'request_number':
                try:
                    cursor.execute(f"ALTER TABLE change_requests ADD COLUMN {key} TEXT")
                except:
                    pass  # 컬럼이 이미 존재하면 무시
        
        # 상태를 강제로 'requested'로 설정
        if 'status' in data:
            data['status'] = 'requested'  # 등록 시에는 항상 '요청' 상태
        
        # 데이터 삽입
        columns = list(data.keys())
        values = list(data.values())
        placeholders = ', '.join(['%s' for _ in values])
        column_names = ', '.join(columns)
        
        cursor.execute(f"""
            INSERT INTO change_requests_cache ({column_names})
            VALUES ({placeholders})
        """, values)
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "변경요청이 저장되었습니다."})
    except Exception as e:
        logging.error(f"변경요청 저장 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/change-request-register")
def change_request_register():
    """변경요청 등록 팝업 페이지"""
    from timezone_config import get_korean_time
    
    # 요청번호 자동 생성 (CR-YYYYMM-NN 형식으로 통일)
    today = get_korean_time()
    base_number = f"CR-{today.strftime('%Y%m')}-"
    
    conn = None
    try:
        # 오늘 날짜의 마지막 번호 찾기
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 변경요청 캐시 테이블이 없으면 생성 (미니멀 필드)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS change_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_number TEXT UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            SELECT MAX(CAST(SUBSTR(request_number, -2) AS INTEGER))
            FROM change_requests
            WHERE request_number LIKE %s
        """, (f"{base_number}%",))
        
        last_number = cursor.fetchone()[0]
        if last_number is None:
            last_number = 0
        
        request_number = f"{base_number}{str(last_number + 1).zfill(2)}"
        conn.commit()
        
        # 동적 컬럼 설정 가져오기 (딕셔너리로 직접 조회)
        _wa = sql_is_active_true('is_active', conn)
        _wd = sql_is_deleted_false('is_deleted', conn)
        dynamic_columns_rows = conn.execute(f"""
            SELECT 
                column_key, column_name, column_type, column_order, is_active,
                dropdown_options, tab, column_span, linked_columns
            FROM change_request_column_config 
            WHERE {_wa} AND {_wd}
            ORDER BY column_order
        """).fetchall()
        
        # 딕셔너리 형태로 변환
        dynamic_columns = []
        for row in dynamic_columns_rows:
            dynamic_columns.append({
                'column_key': row[0],
                'column_name': row[1], 
                'column_type': row[2],
                'column_order': row[3],
                'is_active': row[4],
                'dropdown_options': row[5],
                'tab': row[6],
                'column_span': row[7],
                'linked_columns': row[8]
            })
        
    except Exception as e:
        logging.error(f"요청번호 생성 중 오류: {e}")
        request_number = f"{base_number}01"  # 오류 시 기본값
        dynamic_columns = []
    finally:
        if conn:
            conn.close()
    
    # 드롭다운 컬럼에 대해 코드-값 매핑 적용 (v2 통일)
    for col in dynamic_columns:
        if col['column_type'] == 'dropdown':
            # v2 공통 헬퍼 사용
            code_options = get_dropdown_options_for_display('change_request', col['column_key'])
            
            # 프런트엔드 호환성을 위해 형식 변환
            if code_options:
                col['dropdown_options_mapped'] = [{"code": opt["code"], "value": opt["value"]} for opt in code_options]
            else:
                col['dropdown_options_mapped'] = []
    
    # detailed_content는 별도 섹션으로 처리하므로 dynamic_columns에서 제외
    dynamic_columns = [col for col in dynamic_columns if col['column_key'] != 'detailed_content']
    
    logging.info(f"변경요청 동적 컬럼 {len(dynamic_columns)}개 로드됨 (detailed_content 제외)")
    
    # 팝업 모드 확인
    is_popup = request.args.get('popup', '0') == '1'
    
    return render_template('change-request-register.html', 
                         menu=MENU_CONFIG, 
                         request_number=request_number,
                         today=today,
                         is_popup=is_popup,
                         dynamic_columns=dynamic_columns)

@app.route("/admin/menu-settings")
@require_admin_auth
def admin_menu_settings():
    """메뉴 설정 페이지"""
    return render_template('admin-menu-settings.html', menu=MENU_CONFIG)

@app.route('/permission-request')
def permission_request_page():
    """일반 사용자의 권한 신청 페이지"""
    if not session.get('user_name'):
        session['next_url'] = request.url
        return redirect('/SSO')

    menu_options = []
    for section in MENU_CONFIG:
        submenu_items = section.get('submenu', [])
        if not submenu_items:
            continue
        menu_options.append({
            'group': section.get('title', ''),
            'menus': [
                {
                    'code': resolve_menu_code(item.get('url')),
                    'slug': item.get('url'),
                    'title': item.get('title')
                }
                for item in submenu_items if item.get('url')
            ]
        })

    return render_template(
        'permission-request.html',
        menu=MENU_CONFIG,
        menu_options=menu_options
    )

@app.route("/admin/permission-settings")
@require_admin_auth
def admin_permission_settings():
    """권한 설정 페이지 - 세밀한 CRUD 권한 관리"""
    try:
        super_flag = is_super_admin()
    except Exception:
        login_id = session.get('user_id')
        super_flag = login_id in SUPER_ADMIN_USERS if login_id else False
    return render_template(
        'admin/permission_management.html',
        menu=MENU_CONFIG,
        is_super_admin=super_flag
    )

@app.route("/admin/usage-dashboard")
@require_admin_auth
def admin_usage_dashboard():
    """사용 현황 대시보드"""
    return render_template('admin/usage_dashboard.html', menu=MENU_CONFIG)

@app.route("/admin/data-management")
@require_admin_auth
def admin_data_management():
    """데이터 관리 페이지"""
    return render_template('admin-data-management.html', menu=MENU_CONFIG)


@app.route('/api/table-search')
@require_admin_auth
def api_table_search():
    group = request.args.get('group', '').strip()
    query = request.args.get('q', '').strip()
    if not group or not query:
        return jsonify([])

    cfg = configparser.ConfigParser()
    try:
        cfg.read('config.ini', encoding='utf-8')
    except Exception as exc:
        logging.warning(f"table-search config read error: {exc}")

    def _config_query(key: str, fallback: str) -> str:
        try:
            base = cfg.get('MASTER_DATA_QUERIES', key, fallback=fallback)
        except Exception:
            base = fallback
        base = base.strip().rstrip(';')
        return base

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if group == 'building':
            base_query = _config_query('building_query', 'SELECT building_code, building_name, site, site_type FROM buildings')
            wrapped = (
                "SELECT * FROM (" + base_query + ") AS src "
                "WHERE (src.building_name ILIKE %s OR src.building_code ILIKE %s) "
                "ORDER BY src.building_name LIMIT 20"
            )
            cursor.execute(wrapped, (f'%{query}%', f'%{query}%'))
        elif group == 'department':
            base_query = _config_query('department_query', 'SELECT dept_id, dept_name, parent_dept_code FROM departments')
            wrapped = (
                "SELECT * FROM (" + base_query + ") AS src "
                "WHERE (src.dept_name ILIKE %s OR src.dept_id ILIKE %s) "
                "ORDER BY src.dept_name LIMIT 20"
            )
            cursor.execute(wrapped, (f'%{query}%', f'%{query}%'))
        else:
            return jsonify([])

        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        def _value(row_dict, *candidates):
            for key in candidates:
                if key in row_dict:
                    return row_dict[key]
                lower = key.lower()
                if lower in row_dict:
                    return row_dict[lower]
                upper = key.upper()
                if upper in row_dict:
                    return row_dict[upper]
            return None

        payload = []
        for row in rows:
            row_dict = {columns[idx]: row[idx] for idx in range(len(columns))}
            if group == 'building':
                code = _value(row_dict, 'building_code', 'code')
                name = _value(row_dict, 'building_name', 'name', 'title')
                site = _value(row_dict, 'site', 'site_name')
                site_type = _value(row_dict, 'site_type', 'zone', 'region')
                display = f"{name} ({code})" if name and code else (name or code)
                payload.append({
                    'id': code,
                    'code': code,
                    'name': name,
                    'display_name': display,
                    'site': site,
                    'site_type': site_type,
                })
            else:
                code = _value(row_dict, 'dept_id', 'dept_code', 'code')
                name = _value(row_dict, 'dept_name', 'name', 'title')
                parent = _value(row_dict, 'parent_dept_code', 'parent_code', 'parent_id')
                display = f"{name} ({code})" if name and code else (name or code)
                payload.append({
                    'id': code,
                    'code': code,
                    'name': name,
                    'display_name': display,
                    'parent_dept_code': parent,
                })
        return jsonify(payload)
    except Exception as exc:
        logging.error(f"Table search error: {exc}")
        return jsonify([]), 500
    finally:
        cursor.close()
        conn.close()



@app.route("/api/accidents/deleted")
def get_deleted_accidents():
    """삭제된 사고 목록 API"""
    conn = get_db_connection()
    
    # 삭제된 사고만 조회
    deleted_accidents_rows = conn.execute("""
        SELECT * FROM accidents_cache 
        WHERE is_deleted = 1
        ORDER BY accident_date DESC, accident_number DESC
    """).fetchall()
    
    deleted_accidents = [dict(row) for row in deleted_accidents_rows]
    conn.close()
    
    return jsonify({"success": True, "accidents": deleted_accidents})

@app.route("/api/partners/deleted")
def get_deleted_partners():
    """삭제된 협력사 목록 API"""
    conn = get_db_connection()
    
    # 삭제된 협력사만 조회
    deleted_partners_rows = conn.execute("""
        SELECT * FROM partners_cache 
        WHERE is_deleted = 1
        ORDER BY company_name
    """).fetchall()
    
    deleted_partners = [dict(row) for row in deleted_partners_rows]
    conn.close()
    
    return jsonify({"success": True, "partners": deleted_partners})

@app.route("/api/safety-instruction/deleted")
def get_deleted_safety_instructions():
    """삭제된 안전교육 목록 API"""
    conn = get_db_connection()
    
    deleted_items = conn.execute("""
        SELECT * FROM safety_instructions 
        WHERE is_deleted = 1
        ORDER BY created_at DESC
    """).fetchall()
    
    conn.close()
    return jsonify({"success": True, "items": [dict(row) for row in deleted_items]})


@app.route("/api/follow-sop/deleted") 
def get_deleted_follow_sop():
    """삭제된 Follow SOP 목록 API"""
    conn = get_db_connection()
    
    deleted_items = conn.execute("""
        SELECT * FROM follow_sop 
        WHERE is_deleted = 1
        ORDER BY created_at DESC
    """).fetchall()
    
    conn.close()
    return jsonify({"success": True, "items": [dict(row) for row in deleted_items]})


@app.route("/api/safe-workplace/deleted")
def get_deleted_safe_workplace():
    """삭제된 Safe Workplace 목록 API"""
    conn = get_db_connection()

    deleted_items = conn.execute("""
        SELECT * FROM safe_workplace
        WHERE is_deleted = 1
        ORDER BY created_at DESC
    """).fetchall()

    conn.close()
    return jsonify({"success": True, "items": [dict(row) for row in deleted_items]})


@app.route("/api/full-process")
def get_full_process():
    """Full Process 목록 API (일반 데이터)"""
    try:
        conn = get_db_connection()
        
        # 삭제되지 않은 Full Process 목록 조회
        _wd = sql_is_deleted_false('is_deleted', conn)
        items = conn.execute(f"""
            SELECT * FROM full_process 
            WHERE {_wd}
            ORDER BY created_at DESC
        """).fetchall()
        
        conn.close()
        
        # 결과를 딕셔너리 리스트로 변환
        result = []
        for item in items:
            item_dict = dict(item)
            # custom_data를 JSON으로 파싱
            if item_dict.get('custom_data'):
                try:
                    import json
                    custom_data = json.loads(item_dict['custom_data'])
                    item_dict.update(custom_data)  # custom_data의 내용을 최상위로 병합
                except json.JSONDecodeError:
                    pass
            result.append(item_dict)
        
        return jsonify({
            "success": True,
            "data": result,
            "total": len(result)
        })
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/full-process/deleted")
def get_deleted_full_process():
    """삭제된 Full Process 목록 API"""
    conn = get_db_connection()
    
    deleted_items = conn.execute("""
        SELECT * FROM full_process 
        WHERE is_deleted = 1
        ORDER BY created_at DESC
    """).fetchall()
    
    conn.close()
    return jsonify({"success": True, "items": [dict(row) for row in deleted_items]})


@app.route("/api/safety-instruction/restore", methods=['POST'])
def restore_safety_instructions():
    """안전교육 복구 API"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        for item_id in ids:
            cursor.execute("UPDATE safety_instructions SET is_deleted = 0 WHERE id = %s", (item_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": f"복구 완료: {len(ids)}개 항목"})
        
    except Exception as e:
        logging.error(f"Error restoring safety instructions: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/follow-sop/restore", methods=['POST'])
def restore_follow_sop():
    """Follow SOP 복구 API"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        for item_id in ids:
            cursor.execute("UPDATE follow_sop SET is_deleted = 0 WHERE work_req_no = %s", (item_id,))
            # Cache table is no longer used for display, only update main table
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": f"복구 완료: {len(ids)}개 항목"})
        
    except Exception as e:
        logging.error(f"Error restoring follow SOP: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safe-workplace/restore", methods=['POST'])
def restore_safe_workplace():
    """Safe Workplace 복구 API"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])

        conn = get_db_connection()
        cursor = conn.cursor()

        for item_id in ids:
            cursor.execute("UPDATE safe_workplace SET is_deleted = 0 WHERE safeplace_no = %s", (item_id,))

        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": f"복구 완료: {len(ids)}개 항목"})

    except Exception as e:
        logging.error(f"Error restoring safe workplace: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/full-process/restore", methods=['POST'])
def restore_full_process():
    """Full Process 복구 API"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        for item_id in ids:
            cursor.execute("UPDATE full_process SET is_deleted = 0 WHERE fullprocess_number = %s", (item_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": f"복구 완료: {len(ids)}개 항목"})
        
    except Exception as e:
        logging.error(f"Error restoring full process: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/<board_type>/delete', methods=['POST'])
def delete_items(board_type):
    """범용 소프트 삭제 API - 각 게시판의 primary key를 사용"""
    try:
        data = request.json
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400
        
        # 게시판별 테이블 및 primary key 매핑
        board_config = {
            'accidents': {'table': 'accidents_cache', 'pk': 'id'},
            'safety-instructions': {'table': 'safety_instructions', 'pk': 'id'},
            'change-requests': {'table': 'change_requests_cache', 'pk': 'id'},
            'partners': {'table': 'partners_cache', 'pk': 'id'}
        }
        
        if board_type not in board_config:
            return jsonify({"success": False, "message": "잘못된 게시판 타입입니다."}), 400
        
        config = board_config[board_type]
        table_name = config['table']
        pk_column = config['pk']
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 소프트 삭제 수행
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE {table_name} 
            SET is_deleted = 1 
            WHERE {pk_column} IN ({placeholders})
        """, ids)
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"{board_type} 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/accidents/delete', methods=['POST'])
def delete_accidents():
    """선택한 사고들을 소프트 삭제"""
    try:
        data = request.json
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 모든 사고 삭제 가능 (ACC, K 모두)
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE accidents_cache 
            SET is_deleted = 1 
            WHERE id IN ({placeholders})
        """, ids)
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"사고 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/safety-instructions/delete', methods=['POST'])
def delete_safety_instructions():
    """선택한 환경안전지시서들을 소프트 삭제"""
    try:
        data = request.json
        ids = data.get('ids', [])  # 실제로는 issue_number들
        
        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 메인 테이블에서 소프트 삭제 (issue_number 기준)
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE safety_instructions 
            SET is_deleted = 1 
            WHERE issue_number IN ({placeholders})
        """, ids)
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"환경안전지시서 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/follow-sop/delete', methods=['POST'])
def delete_follow_sop():
    """선택한 Follow SOP들을 소프트 삭제"""
    try:
        data = request.json
        ids = data.get('ids', [])  # 실제로는 work_req_no들
        
        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # follow_sop 테이블에서 소프트 삭제 (work_req_no 기준)
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE follow_sop 
            SET is_deleted = 1 
            WHERE work_req_no IN ({placeholders})
        """, ids)
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"Follow SOP 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/safe-workplace/delete', methods=['POST'])
def delete_safe_workplace():
    """선택한 Safe Workplace를 소프트 삭제"""
    try:
        data = request.json
        ids = data.get('ids', [])  # safeplace_no 목록

        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE safe_workplace
            SET is_deleted = 1
            WHERE safeplace_no IN ({placeholders})
        """, ids)

        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"Safe Workplace 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/full-process/delete', methods=['POST'])
def delete_full_process():
    """선택한 Full Process들을 소프트 삭제"""
    try:
        data = request.json
        ids = data.get('ids', [])  # 실제로는 fullprocess_number들

        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # full_process 테이블에서 소프트 삭제 (fullprocess_number 기준)
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE full_process 
            SET is_deleted = 1 
            WHERE fullprocess_number IN ({placeholders})
        """, ids)

        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"Full Process 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
def _update_final_check_bulk(conn, *, board_type: str, id_field: str, ids: List[str], status_value: str, tables: List[str]):
    """공통 JSONB 업데이트 로직"""

    cursor = conn.cursor()

    def _row_value(row, key, index):
        if isinstance(row, dict):
            return row.get(key)
        getter = getattr(row, 'get', None)
        if callable(getter):
            value = getter(key)
            if value is not None:
                return value
        try:
            return row[index]
        except Exception:
            return getattr(row, key, None)

    try:
        cursor.execute(
            """
            SELECT option_code, option_value
            FROM dropdown_option_codes_v2
            WHERE board_type = %s
              AND column_key = %s
              AND COALESCE(is_active, 1) = 1
            ORDER BY display_order
            """,
            (board_type, 'final_check_yn'),
        )
        option_rows = cursor.fetchall()
    except Exception:
        option_rows = []

    status_code = None
    status_label = status_value

    if option_rows:
        matched_row = None
        for row in option_rows:
            value = _row_value(row, 'option_value', 1)
            if isinstance(value, str) and value.strip() == status_value:
                matched_row = row
                break
        if matched_row is None:
            matched_row = option_rows[0]

        status_code = _row_value(matched_row, 'option_code', 0)
        raw_label = _row_value(matched_row, 'option_value', 1)
        if isinstance(raw_label, str) and raw_label.strip():
            status_label = raw_label.strip()

    payload_dict = {
        'final_check_yn': status_code or status_value,
        'final_check_yn_label': status_label,
    }
    payload_json = json.dumps(payload_dict, ensure_ascii=False)

    placeholders = ','.join(['%s'] * len(ids))

    def _table_exists(table_name: str) -> bool:
        cursor.execute("SELECT to_regclass(%s)", (f"public.{table_name}",))
        row = cursor.fetchone()
        if isinstance(row, dict):
            value = next(iter(row.values()), None)
        else:
            value = row[0] if row else None
        return value is not None

    updated_count = 0
    primary_count_recorded = False

    for table_name in tables:
        if not _table_exists(table_name):
            continue

        update_sql = (
            f"UPDATE {table_name} "
            "SET custom_data = jsonb_set("  # outer jsonb_set to update label
            "    jsonb_set("  # inner jsonb_set to update code
            "        COALESCE(custom_data::jsonb, '{}'::jsonb),"
            "        '{final_check_yn}', to_jsonb(%s::text), true"
            "    ),"
            "    '{final_check_yn_label}', to_jsonb(%s::text), true"
            ") "
            f"WHERE {id_field} IN ({placeholders})"
        )
        cursor.execute(update_sql, [status_code or status_value, status_label, *ids])
        if not primary_count_recorded:
            updated_count = cursor.rowcount
            primary_count_recorded = True

    return updated_count, status_label


@app.route('/api/full-process/final-check', methods=['POST'])
def mark_full_process_final_check():
    """선택 항목의 최종 검토 상태를 일괄 업데이트"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])

    if not ids:
        return jsonify({"success": False, "message": "처리할 항목을 선택해주세요."}), 400

    status_value = (data.get('status') or '검토완료').strip() or '검토완료'

    conn = None
    try:
        conn = get_db_connection()
        if not getattr(conn, 'is_postgres', False):
            raise RuntimeError('PostgreSQL backend is required for final check updates.')

        unique_ids = list(dict.fromkeys(str(x).strip() for x in ids if str(x).strip()))
        if not unique_ids:
            return jsonify({"success": False, "message": "유효한 항목이 없습니다."}), 400

        updated_count, status_label = _update_final_check_bulk(
            conn,
            board_type='full_process',
            id_field='fullprocess_number',
            ids=unique_ids,
            status_value=status_value,
            tables=['full_process', 'full_process_cache', 'fullprocess_cache'],
        )

        conn.commit()

        message_suffix = f" '{status_label}'" if status_label else ''
        return jsonify({
            "success": True,
            "updated_count": updated_count,
            "message": f"{updated_count}건의 상태가{message_suffix}로 변경되었습니다.",
        })
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        logging.error(f"Full Process 최종 검토 업데이트 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/follow-sop/final-check', methods=['POST'])
def mark_follow_sop_final_check():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])

    if not ids:
        return jsonify({"success": False, "message": "처리할 항목을 선택해주세요."}), 400

    status_value = (data.get('status') or '검토완료').strip() or '검토완료'

    conn = None
    try:
        conn = get_db_connection()
        if not getattr(conn, 'is_postgres', False):
            raise RuntimeError('PostgreSQL backend is required for final check updates.')

        unique_ids = list(dict.fromkeys(str(x).strip() for x in ids if str(x).strip()))
        if not unique_ids:
            return jsonify({"success": False, "message": "유효한 항목이 없습니다."}), 400

        updated_count, status_label = _update_final_check_bulk(
            conn,
            board_type='follow_sop',
            id_field='work_req_no',
            ids=unique_ids,
            status_value=status_value,
            tables=['follow_sop', 'follow_sop_cache'],
        )

        conn.commit()

        message_suffix = f" '{status_label}'" if status_label else ''
        return jsonify({
            "success": True,
            "updated_count": updated_count,
            "message": f"{updated_count}건의 상태가{message_suffix}로 변경되었습니다.",
        })
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        logging.error(f"Follow SOP 최종 검토 업데이트 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/safe-workplace/final-check', methods=['POST'])
def mark_safe_workplace_final_check():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])

    if not ids:
        return jsonify({"success": False, "message": "처리할 항목을 선택해주세요."}), 400

    status_value = (data.get('status') or '검토완료').strip() or '검토완료'

    conn = None
    try:
        conn = get_db_connection()
        if not getattr(conn, 'is_postgres', False):
            raise RuntimeError('PostgreSQL backend is required for final check updates.')

        unique_ids = list(dict.fromkeys(str(x).strip() for x in ids if str(x).strip()))
        if not unique_ids:
            return jsonify({"success": False, "message": "유효한 항목이 없습니다."}), 400

        updated_count, status_label = _update_final_check_bulk(
            conn,
            board_type='safe_workplace',
            id_field='safeplace_no',
            ids=unique_ids,
            status_value=status_value,
            tables=['safe_workplace'],
        )

        conn.commit()

        message_suffix = f" '{status_label}'" if status_label else ''
        return jsonify({
            "success": True,
            "updated_count": updated_count,
            "message": f"{updated_count}건의 상태가{message_suffix}로 변경되었습니다.",
        })
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        logging.error(f"Safe Workplace 최종 검토 업데이트 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/accidents/restore', methods=['POST'])
def restore_accidents():
    """삭제된 사고들을 복구"""
    try:
        data = request.json
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({"success": False, "message": "복구할 항목이 없습니다."}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 선택한 사고들을 복구 (is_deleted = 0)
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE accidents_cache 
            SET is_deleted = 0 
            WHERE id IN ({placeholders})
        """, ids)
        
        restored_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "restored_count": restored_count,
            "message": f"{restored_count}건이 복구되었습니다."
        })
    except Exception as e:
        logging.error(f"사고 복구 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/partners/restore', methods=['POST'])
def restore_partners():
    """삭제된 협력사들을 복구"""
    try:
        data = request.json
        business_numbers = data.get('business_numbers', [])
        
        if not business_numbers:
            return jsonify({"success": False, "message": "복구할 항목이 없습니다."}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 선택한 협력사들을 복구 (is_deleted = 0)
        placeholders = ','.join(['%s'] * len(business_numbers))
        cursor.execute(f"""
            UPDATE partners_cache 
            SET is_deleted = 0 
            WHERE business_number IN ({placeholders})
        """, business_numbers)
        
        restored_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "restored_count": restored_count,
            "message": f"{restored_count}개의 협력사가 복구되었습니다."
        })
    except Exception as e:
        logging.error(f"협력사 복구 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/buildings/search', methods=['GET'])
def search_buildings():
    """건물 검색 API"""
    try:
        search_term = request.args.get('q', '').strip()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if search_term:
            cursor.execute("""
                SELECT building_code, building_name
                FROM buildings_cache
                WHERE building_name LIKE %s OR building_code LIKE %s
                ORDER BY building_name
                LIMIT 50
            """, (f'%{search_term}%', f'%{search_term}%'))
        else:
            cursor.execute("""
                SELECT building_code, building_name
                FROM buildings_cache
                ORDER BY building_name
                LIMIT 50
            """)
        
        buildings = []
        for row in cursor.fetchall():
            buildings.append({
                'building_code': row[0],
                'building_name': row[1]
            })
        
        conn.close()
        return jsonify({'success': True, 'buildings': buildings})
    except Exception as e:
        logging.error(f"건물 검색 중 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/departments/search', methods=['GET'])
def search_departments():
    """부서 검색 API"""
    try:
        search_term = request.args.get('q', '').strip()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if search_term:
            cursor.execute("""
                SELECT d.dept_code, d.dept_name, 
                       p.dept_name as parent_name
                FROM departments_cache d
                LEFT JOIN departments_cache p ON d.parent_dept_code = p.dept_code
                WHERE d.dept_name LIKE %s OR d.dept_code LIKE %s
                ORDER BY d.dept_name
                LIMIT 50
            """, (f'%{search_term}%', f'%{search_term}%'))
        else:
            cursor.execute("""
                SELECT d.dept_code, d.dept_name, 
                       p.dept_name as parent_name
                FROM departments_cache d
                LEFT JOIN departments_cache p ON d.parent_dept_code = p.dept_code
                ORDER BY d.dept_name
                LIMIT 50
            """)
        
        departments = []
        for row in cursor.fetchall():
            departments.append({
                'dept_code': row[0],
                'dept_name': row[1],
                'parent_name': row[2] or '',
                'parent_dept_code': d.parent_dept_code or ''
            })
        
        conn.close()
        return jsonify({'success': True, 'departments': departments})
    except Exception as e:
        logging.error(f"부서 검색 중 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/search', methods=['GET'])
def api_search():
    """범용 검색 API"""
    try:
        search_type = request.args.get('type', 'person')
        search_value = request.args.get('value', 'person')
        search_term = request.args.get('q', '').strip()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        results = []
        
        if search_type == 'person':
            # 담당자 검색 (더미 데이터)
            sample_persons = [
                {'employee_id': '10001', 'name': '김철수', 'department': '안전보건팀'},
                {'employee_id': '10002', 'name': '이영희', 'department': '생산1팀'},
                {'employee_id': '10003', 'name': '박민수', 'department': '품질관리팀'},
                {'employee_id': '10004', 'name': '정수진', 'department': '인사팀'},
                {'employee_id': '10005', 'name': '최동욱', 'department': '총무팀'}
            ]
            
            if search_term:
                if search_value == 'employee_id':
                    results = [p for p in sample_persons if search_term in p['employee_id']]
                else:
                    results = [p for p in sample_persons if search_term in p['name'] or search_term in p['department']]
            else:
                results = sample_persons
                
        elif search_type == 'company':
            # 협력사 검색
            if search_term:
                if search_value == 'business_number':
                    cursor.execute("""
                        SELECT business_number, company_name, representative, 
                               business_type_major, NULL as phone
                        FROM partners_cache
                        WHERE business_number LIKE %s AND is_deleted = 0
                        LIMIT 50
                    """, (f'%{search_term}%',))
                else:
                    cursor.execute("""
                        SELECT business_number, company_name, representative, 
                               business_type_major, NULL as phone
                        FROM partners_cache
                        WHERE company_name LIKE %s AND is_deleted = 0
                        LIMIT 50
                    """, (f'%{search_term}%',))
            else:
                cursor.execute("""
                    SELECT business_number, company_name, representative, 
                           business_type_major, NULL as phone
                    FROM partners_cache
                    WHERE is_deleted = 0
                    LIMIT 50
                """)
            
            for row in cursor.fetchall():
                results.append({
                    'company_business_number': row[0],
                    'company_name': row[1],
                    'representative_name': row[2] or '',
                    'business_type': row[3] or '',
                    'company_phone': row[4] or ''
                })
                
        elif search_type == 'building':
            # 건물 검색 - 외부 DB에서 조회
            try:
                from database_config import execute_SQL
                
                # config.ini에서 BUILDING_QUERY 가져오기
                building_query = db_config.config.get('MASTER_DATA_QUERIES', 'BUILDING_QUERY')
                
                # 검색어 조건 추가
                if search_term:
                    if search_value == 'building_code':
                        building_query += f" AND building_code LIKE '%{search_term}%'"
                    else:
                        building_query += f" AND building_name LIKE '%{search_term}%'"
                
                building_query += " ORDER BY building_name LIMIT 50"
                
                # 외부 DB에서 실행
                df = execute_SQL(building_query)
                
                for _, row in df.iterrows():
                    results.append({
                        'building_code': row.get('building_code', ''),
                        'building_name': row.get('building_name', '')
                    })
                    
            except Exception as e:
                logging.warning(f"외부 DB 건물 조회 실패, 로컬 DB 사용: {e}")
                # 외부 DB 실패 시 로컬 DB 사용
                if search_term:
                    if search_value == 'building_code':
                        cursor.execute("""
                            SELECT building_code, building_name
                            FROM buildings_cache
                            WHERE building_code LIKE %s
                            ORDER BY building_name
                            LIMIT 50
                        """, (f'%{search_term}%',))
                    else:
                        cursor.execute("""
                            SELECT building_code, building_name
                            FROM buildings_cache
                            WHERE building_name LIKE %s
                            ORDER BY building_name
                            LIMIT 50
                        """, (f'%{search_term}%',))
                else:
                    cursor.execute("""
                        SELECT building_code, building_name
                        FROM buildings_cache
                        ORDER BY building_name
                        LIMIT 50
                    """)
                
                for row in cursor.fetchall():
                    results.append({
                        'building_code': row[0],
                        'building_name': row[1]
                    })
                
        elif search_type == 'department':
            # 부서 검색 - 외부 DB에서 조회
            try:
                from database_config import execute_SQL
                
                # config.ini에서 DEPARTMENT_QUERY 가져오기
                department_query = db_config.config.get('MASTER_DATA_QUERIES', 'DEPARTMENT_QUERY')
                
                # 검색어 조건 추가
                if search_term:
                    if search_value == 'dept_code':
                        department_query += f" AND dept_code LIKE '%{search_term}%'"
                    else:
                        department_query += f" AND dept_name LIKE '%{search_term}%'"
                
                department_query += " ORDER BY dept_level, dept_name LIMIT 50"
                
                # 외부 DB에서 실행
                df = execute_SQL(department_query)
                
                for _, row in df.iterrows():
                    results.append({
                        'dept_code': row.get('dept_code', ''),
                        'dept_name': row.get('dept_name', ''),
                        'parent_name': row.get('parent_dept_code', ''),  # 부모 부서코드를 이름으로 사용
                        'dept_level': row.get('dept_level', 0)
                    })
                    
            except Exception as e:
                logging.warning(f"외부 DB 부서 조회 실패, 로컬 DB 사용: {e}")
                # 외부 DB 실패 시 로컬 DB 사용
                if search_term:
                    if search_value == 'dept_code':
                        cursor.execute("""
                            SELECT d.dept_code, d.dept_name, 
                                   p.dept_name as parent_name, d.dept_level
                            FROM departments_cache d
                            LEFT JOIN departments_cache p ON d.parent_dept_code = p.dept_code
                            WHERE d.dept_code LIKE %s
                            ORDER BY d.dept_name
                            LIMIT 50
                        """, (f'%{search_term}%',))
                    else:
                        cursor.execute("""
                            SELECT d.dept_code, d.dept_name, 
                                   p.dept_name as parent_name, d.dept_level
                            FROM departments_cache d
                            LEFT JOIN departments_cache p ON d.parent_dept_code = p.dept_code
                            WHERE d.dept_name LIKE %s
                            ORDER BY d.dept_name
                            LIMIT 50
                        """, (f'%{search_term}%',))
                else:
                    cursor.execute("""
                        SELECT d.dept_code, d.dept_name, 
                               p.dept_name as parent_name, d.dept_level
                        FROM departments_cache d
                        LEFT JOIN departments_cache p ON d.parent_dept_code = p.dept_code
                        ORDER BY d.dept_name
                        LIMIT 50
                    """)
                
                for row in cursor.fetchall():
                    results.append({
                        'dept_code': row[0],
                        'dept_name': row[1],
                        'parent_name': row[2] or '',
                        'parent_dept_code': d.parent_dept_code or ''
                    })
                
        elif search_type == 'contractor':
            # 협력사 근로자 검색 - 외부 DB에서 조회
            try:
                from database_config import execute_SQL
                
                # config.ini에서 CONTRACTOR_QUERY 가져오기
                contractor_query = db_config.config.get('MASTER_DATA_QUERIES', 'CONTRACTOR_QUERY')
                
                # 검색어 조건 추가
                if search_term:
                    if search_value == 'worker_id':
                        contractor_query += f" AND worker_id LIKE '%{search_term}%'"
                    else:
                        contractor_query += f" AND worker_name LIKE '%{search_term}%'"
                
                contractor_query += " ORDER BY worker_name LIMIT 50"
                
                # 외부 DB에서 실행
                df = execute_SQL(contractor_query)
                
                for _, row in df.iterrows():
                    results.append({
                        'worker_id': row.get('worker_id', ''),
                        'worker_name': row.get('worker_name', ''),
                        'company_name': row.get('company_name', ''),
                        'business_number': row.get('business_number', '')
                    })
                    
            except Exception as e:
                logging.warning(f"외부 DB 협력사 근로자 조회 실패, 더미 데이터 사용: {e}")
                # 외부 DB 실패 시 더미 데이터 사용
                sample_contractors = [
                    {'worker_id': 'C001', 'worker_name': '김민수', 'company_name': '삼성건설', 'business_number': '1248100998'},
                    {'worker_id': 'C002', 'worker_name': '이철호', 'company_name': '대림산업', 'business_number': '1108114055'},
                    {'worker_id': 'C003', 'worker_name': '박영진', 'company_name': 'GS건설', 'business_number': '1048145271'},
                    {'worker_id': 'C004', 'worker_name': '최성훈', 'company_name': '현대건설', 'business_number': '1018116293'},
                    {'worker_id': 'C005', 'worker_name': '정미경', 'company_name': '롯데건설', 'business_number': '2148111745'},
                    {'worker_id': 'C006', 'worker_name': '홍길동', 'company_name': '포스코건설', 'business_number': '5068151224'},
                    {'worker_id': 'C007', 'worker_name': '김수연', 'company_name': 'SK건설', 'business_number': '1018143363'},
                    {'worker_id': 'C008', 'worker_name': '장민호', 'company_name': '두산건설', 'business_number': '1028144723'}
                ]
                
                if search_term:
                    if search_value == 'worker_id':  # ID로 검색
                        results = [c for c in sample_contractors if search_term.upper() in c['worker_id'].upper()]
                    else:  # name (이름으로 검색)
                        results = [c for c in sample_contractors if search_term in c['worker_name']]
                else:
                    results = sample_contractors
        
        conn.close()
        return jsonify({'success': True, 'data': results})
    except Exception as e:
        logging.error(f"검색 API 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/accidents/permanent-delete', methods=['POST'])
def permanent_delete_accidents():
    """선택한 사고들을 영구 삭제"""
    try:
        data = request.json
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 선택한 사고들을 영구 삭제
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            DELETE FROM accidents_cache 
            WHERE id IN ({placeholders})
        """, ids)
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 영구 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"사고 영구 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/safety-instruction-columns", methods=["GET"])
def get_safety_instruction_columns():
    """환경안전 지시서 페이지 동적 컬럼 설정 조회"""
    try:
        column_service = ColumnConfigService('safety_instruction', DB_PATH)
        columns = column_service.list_columns()
        return jsonify(columns)
    except Exception as e:
        logging.error(f"환경안전 지시서 컬럼 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/safety-instruction-columns", methods=["POST"])
def add_safety_instruction_column():
    """환경안전 지시서 페이지 동적 컬럼 추가"""
    try:
        column_service = ColumnConfigService('safety_instruction', DB_PATH)
        result = column_service.add_column(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"환경안전 지시서 컬럼 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/safety-instruction-columns/<int:column_id>", methods=["PUT"])
def update_safety_instruction_column(column_id):
    """환경안전 지시서 페이지 동적 컬럼 수정"""
    try:
        column_service = ColumnConfigService('safety_instruction', DB_PATH)
        result = column_service.update_column(column_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"환경안전 지시서 컬럼 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/safety-instruction-columns/<int:column_id>", methods=["DELETE"])
def delete_safety_instruction_column(column_id):
    """환경안전 지시서 페이지 동적 컬럼 삭제 (비활성화)"""
    try:
        column_service = ColumnConfigService('safety_instruction', DB_PATH)
        result = column_service.delete_column(column_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"환경안전 지시서 컬럼 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ===== Admin utility: force delete safety-instruction columns by keys =====
@app.route("/admin/safety-instruction-columns/force-delete-keys", methods=["POST"])
@require_admin_auth
def admin_force_delete_si_columns():
    """강제 컬럼 삭제(soft delete). 관리자 전용.

    Request JSON: {"keys": ["attachments","notes","note"]}
    """
    try:
        data = request.get_json(force=True) or {}
        keys = data.get('keys') or []
        if not keys:
            return jsonify({"success": False, "message": "keys is required"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        placeholders = ','.join(['%s'] * len(keys))
        cursor.execute(
            f"UPDATE safety_instruction_column_config SET is_deleted = 1 WHERE LOWER(column_key) IN ({placeholders})",
            [k.lower() for k in keys]
        )
        affected = cursor.rowcount
        conn.commit()
        conn.close()
        return jsonify({"success": True, "deleted": affected})
    except Exception as e:
        logging.error(f"force delete si columns error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ============= 섹션 관리 API =============
@app.route("/api/safety-instruction-sections", methods=["GET"])
def get_safety_instruction_sections():
    """환경안전 지시서 섹션 목록 조회"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safety_instruction', DB_PATH)
        sections = section_service.get_sections()
        return jsonify({"success": True, "sections": sections})
    except Exception as e:
        logging.error(f"섹션 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/safety-instruction-sections", methods=["POST"])
def add_safety_instruction_section():
    """환경안전 지시서 섹션 추가"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safety_instruction', DB_PATH)
        result = section_service.add_section(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"섹션 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/safety-instruction-sections/<int:section_id>", methods=["PUT"])
def update_safety_instruction_section(section_id):
    """환경안전 지시서 섹션 수정"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safety_instruction', DB_PATH)
        result = section_service.update_section(section_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"섹션 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/safety-instruction-sections/<int:section_id>", methods=["DELETE"])
def delete_safety_instruction_section(section_id):
    """환경안전 지시서 섹션 삭제"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safety_instruction', DB_PATH)
        result = section_service.delete_section(section_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"섹션 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/safety-instruction-sections/reorder", methods=["POST"])
def reorder_safety_instruction_sections():
    """환경안전 지시서 섹션 순서 변경"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('safety_instruction', DB_PATH)
        result = section_service.reorder_sections(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"섹션 순서 변경 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ============= 사고게시판 섹션 관리 API =============
@app.route("/api/accident-sections", methods=["GET"])
def get_accident_sections():
    """사고게시판 섹션 목록 조회"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('accident', DB_PATH)
        sections = section_service.get_sections()
        return jsonify({"success": True, "sections": sections})
    except Exception as e:
        logging.error(f"사고 섹션 조회 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-sections", methods=["POST"])
def add_accident_section():
    """사고게시판 섹션 추가"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('accident', DB_PATH)
        result = section_service.add_section(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"사고 섹션 추가 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-sections/<int:section_id>", methods=["PUT"])
def update_accident_section(section_id):
    """사고게시판 섹션 수정"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('accident', DB_PATH)
        result = section_service.update_section(section_id, request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"사고 섹션 수정 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-sections/<int:section_id>", methods=["DELETE"])
def delete_accident_section(section_id):
    """사고게시판 섹션 삭제"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('accident', DB_PATH)
        result = section_service.delete_section(section_id)
        return jsonify(result)
    except Exception as e:
        logging.error(f"사고 섹션 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accident-sections/reorder", methods=["POST"])
def reorder_accident_sections():
    """사고게시판 섹션 순서 변경"""
    try:
        from section_service import SectionConfigService
        section_service = SectionConfigService('accident', DB_PATH)
        result = section_service.reorder_sections(request.json)
        return jsonify(result)
    except Exception as e:
        logging.error(f"사고 섹션 순서 변경 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/accident-export")
def export_accidents_excel():
    """사고 데이터 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        import io
        
        # 검색 조건 가져오기
        company_name = request.args.get('company_name', '')
        business_number = request.args.get('business_number', '')
        accident_date_start = request.args.get('accident_date_start', '')
        accident_date_end = request.args.get('accident_date_end', '')
        
        # DB 연결
        conn = get_db_connection()
        
        # 섹션 정보 가져오기
        section_sql = f"""
            SELECT section_key, section_name, section_order
            FROM section_config
            WHERE board_type = 'accident'
              AND {sql_is_active_true('is_active', conn)}
              AND {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY section_order
        """
        try:
            sections = [dict(row) for row in conn.execute(section_sql).fetchall()]
        except:
            # 섹션 테이블이 없으면 기본 섹션 사용
            sections = []

        # 동적 컬럼 정보 가져오기 (활성+미삭제)
        where_c_active = sql_is_active_true('is_active', conn)
        where_c_notdel = sql_is_deleted_false('is_deleted', conn)
        dyn_sql = f"""
            SELECT * FROM accident_column_config
            WHERE {where_c_active}
              AND {where_c_notdel}
            ORDER BY column_order
        """
        dynamic_columns_rows = conn.execute(dyn_sql).fetchall()
        dynamic_columns_all = [dict(row) for row in dynamic_columns_rows]

        # 섹션별로 컬럼 그룹핑 (섹션 순서 -> 섹션 내 컬럼 순서)
        dynamic_columns = []
        if sections:
            # 섹션 순서대로 컬럼 추가
            for section in sections:
                section_columns = [col for col in dynamic_columns_all if col.get('tab') == section['section_key']]
                dynamic_columns.extend(section_columns)
            # 섹션이 없는 컬럼들 추가
            no_section_columns = [col for col in dynamic_columns_all if not col.get('tab') or not any(s['section_key'] == col.get('tab') for s in sections)]
            dynamic_columns.extend(no_section_columns)
        else:
            # 섹션 정보가 없으면 기존 순서 사용
            dynamic_columns = dynamic_columns_all
        
        # 사고 데이터 조회 (partner_accident 함수와 동일한 로직)
        # 삭제되지 않은 데이터만 조회
        query = f"""
            SELECT * FROM accidents_cache 
            WHERE {sql_is_deleted_false('is_deleted', conn)}
        """
        params = []
        
        # company_name과 business_number 필터링은 제거 (responsible_company 관련)
        
        if accident_date_start:
            query += " AND accident_date >= %s"
            params.append(accident_date_start)
        
        if accident_date_end:
            query += " AND accident_date <= %s"
            params.append(accident_date_end)
        
        # 등록일 기준 최신순 정렬 (시분초까지 정확히 정렬됨)
        query += " ORDER BY created_at DESC, accident_number DESC"
        
        accidents = conn.execute(query, params).fetchall()
        
        # 디버그: 첫 번째 사고 데이터 로깅
        if accidents:
            first_accident = dict(accidents[0])
            logging.info(f"First accident data: {first_accident}")
            if first_accident.get('custom_data'):
                logging.info(f"Custom data: {first_accident['custom_data']}")
        
        # 디버그: 동적 컬럼 정보 로깅  
        logging.info(f"Dynamic columns count: {len(dynamic_columns)}")
        for col in dynamic_columns[:3]:  # 처음 3개만
            logging.info(f"Column: {col['column_key']} - {col['column_name']} ({col['column_type']})")
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "사고 현황"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성: 목록 테이블과 동일 구성 (사고번호, 등록일, 사고명 + 동적 컬럼)
        headers = ['사고번호', '등록일', '사고명']
        # 목록에서 숨기는 기본 키는 제외
        skip_keys = {'accident_number', 'created_at', 'accident_name'}
        custom_columns = [col for col in dynamic_columns if col.get('column_key') not in skip_keys]
        headers.extend([col['column_name'] for col in custom_columns])
        
        # 헤더 쓰기
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # 드롭다운 코드-값 매핑 함수
        def get_display_value(column_key, code_value):
            if not code_value or code_value == '':
                return ''
            try:
                options = get_dropdown_options_for_display('accident', column_key)
                if options:
                    # 리스트 형태의 옵션에서 코드에 해당하는 값 찾기
                    for option in options:
                        if option['code'] == code_value:
                            return option['value']
                return code_value
            except:
                return code_value
        
        # 날짜 형식 정리 함수 (시분초 제거)
        def format_date(date_value):
            if not date_value:
                return ''
            date_str = str(date_value)
            if ' ' in date_str:
                return date_str.split(' ')[0]  # 2025-09-07 0:00:00 → 2025-09-07
            return date_str
        
        # 데이터 쓰기 (목록 화면과 동일한 병합/폴백 적용)
        for row_idx, accident_row in enumerate(accidents, 2):
            rec = dict(accident_row)
            # custom_data 파싱
            custom = {}
            try:
                raw = rec.get('custom_data')
                if isinstance(raw, dict):
                    custom = raw
                elif isinstance(raw, str) and raw:
                    custom = pyjson.loads(raw)
            except Exception:
                custom = {}
            # 사고명 폴백
            if not rec.get('accident_name'):
                nm = custom.get('accident_name')
                if nm and str(nm).strip():
                    rec['accident_name'] = nm
            # 등록일 계산
            acc_no = str(rec.get('accident_number') or '')
            if acc_no.startswith('K'):
                display_created = rec.get('report_date') or rec.get('created_at')
            else:
                display_created = rec.get('created_at')
            # 고정 열
            ws.cell(row=row_idx, column=1, value=rec.get('accident_number', ''))
            ws.cell(row=row_idx, column=2, value=format_date(display_created))
            ws.cell(row=row_idx, column=3, value=rec.get('accident_name', ''))
            # 동적 열: 화면과 동일 키 순서로, custom 우선 → 상위값
            start_col = 4
            for offset, col in enumerate(custom_columns):
                key = col['column_key']
                value = custom.get(key, rec.get(key, ''))

                # 빈 리스트 처리를 먼저 수행
                if isinstance(value, list):
                    if col.get('column_type') == 'list':
                        try:
                            # 빈 리스트는 빈 문자열로, 내용이 있으면 JSON 문자열로
                            value = pyjson.dumps(value, ensure_ascii=False) if value else ''
                        except Exception:
                            value = ''
                    else:
                        # list 타입이 아닌데 리스트 값이면 빈 문자열로
                        value = ''
                elif isinstance(value, dict):
                    value = value.get('name') or str(value)
                elif col.get('column_type') == 'dropdown' and value:
                    value = get_display_value(key, value)
                elif col.get('column_type') in ['date','datetime'] and value:
                    value = format_date(value)

                # None이나 빈 리스트인 경우 빈 문자열로 보장
                if value is None or value == [] or value == {}:
                    value = ''

                ws.cell(row=row_idx, column=start_col + offset, value=value)
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일을 메모리에 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        conn.close()
        
        # 파일명 생성
        filename = f"accident_list_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 다운로드 응답
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logging.error(f"엑셀 다운로드 중 오류: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500

# ===== 엑셀 임포트 API =====
@app.route('/api/accident-import', methods=['POST'])
def import_accidents():
    try:
        import openpyxl
        # json already imported globally
        from datetime import datetime
        import re
        
        # 파일 확인
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "파일이 없습니다."}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "message": "파일이 선택되지 않았습니다."}), 400
            
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"success": False, "message": "엑셀 파일만 업로드 가능합니다."}), 400
        
        # 옵션 확인
        skip_duplicates = request.form.get('skip_duplicates') == 'on'
        validate_data = request.form.get('validate_data') == 'on'
        
        # 엑셀 파일 읽기
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # DB 연결
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # accident_columns 테이블 확인 및 생성
        cursor.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='accident_column_config'
        """)
        if not cursor.fetchone():
            # 테이블이 없으면 빈 리스트로 처리
            dynamic_columns = []
            logging.info("accident_columns 테이블이 없어서 동적 컬럼 없이 처리합니다.")
        else:
            # 동적 컬럼 조회
            _wa = sql_is_active_true('is_active', conn)
            _wd = sql_is_deleted_false('is_deleted', conn)
            cursor.execute(f"""
                SELECT column_key, column_name, column_type, dropdown_options
                FROM accident_column_config 
                WHERE {_wa} AND {_wd}
                ORDER BY column_order
            """)
            dynamic_columns = cursor.fetchall()
        
        # 헤더 매핑 (한글 헤더명 -> DB 컬럼명)
        # 주의: 사고번호는 자동 생성하므로 매핑에서 제외
        header_mapping = {
            '사고명': 'accident_name', 
            '재해날짜': 'accident_date',
            '시간': 'accident_time',
            '사고등급': 'accident_level',
            '사고분류': 'accident_classification',
            '재해유형': 'disaster_type',
            '재해형태': 'disaster_form',
            '사업장': 'workplace',
            '건물': 'building',
            '층': 'floor',
            '세부위치': 'location_detail',
            '요일': 'day_of_week',
            '처리상태': 'processing_status',
            '조치사항': 'measures',
            '재발방지대책': 'prevention_measures',
            '담당부서': 'department',
            '담당자': 'manager',
            '완료예정일': 'completion_date',
            '원인분석': 'cause_analysis',
            '첨부문서': 'attachment',
            '발생위치': 'occurrence_location'
        }
        
        # 동적 컬럼 매핑 추가
        for col in dynamic_columns:
            header_mapping[col['column_name']] = col['column_key']
        
        # 첫 번째 행에서 헤더 읽기
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else '')
        
        success_count = 0
        error_count = 0
        errors = []
        
        # 헤더 정보 로그
        logging.info(f"엑셀 헤더: {headers}")
        logging.info(f"헤더 매핑: {header_mapping}")
        
        # 데이터 행 처리 (2행부터)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # 빈 행 건너뛰기
                if not any(row):
                    continue
                
                logging.info(f"처리 중인 행 {row_idx}: {row}")
                
                # 데이터 매핑
                data = {}
                custom_data = {}
                
                for col_idx, cell_value in enumerate(row):
                    if col_idx >= len(headers):
                        break
                        
                    header = headers[col_idx]
                    if not header or not cell_value:
                        continue
                        
                    # 문자열로 변환
                    str_value = str(cell_value).strip()
                    if not str_value:
                        continue
                    
                    # 헤더 매핑
                    if header in header_mapping:
                        db_column = header_mapping[header]
                        
                        # 기본 컬럼인지 동적 컬럼인지 확인 (accident_number는 자동 생성하므로 제외)
                        if db_column in ['accident_name', 'accident_date', 'accident_time', 
                                       'accident_level', 'accident_classification', 'disaster_type', 'disaster_form',
                                       'workplace', 'building', 'floor', 'location_detail', 'day_of_week',
                                       'processing_status', 'measures', 
                                       'prevention_measures', 'department', 'manager', 'completion_date',
                                       'cause_analysis', 'attachment', 'occurrence_location']:
                            data[db_column] = str_value
                        else:
                            # 동적 컬럼
                            custom_data[db_column] = str_value
                
                # 사고번호는 항상 자동 생성 (사용자 입력 무시)
                if data.get('accident_date'):
                    # 재해날짜를 기준으로 ACCYYMMDD 형식으로 생성
                    try:
                        accident_date = data['accident_date']
                        if isinstance(accident_date, str):
                            # 날짜 문자열을 파싱
                            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
                                try:
                                    dt = datetime.strptime(accident_date, fmt)
                                    break
                                except ValueError:
                                    continue
                            else:
                                # 파싱 실패시 현재 날짜 사용
                                dt = get_korean_time()
                        else:
                            dt = datetime.now()
                        
                        # ACCYYMMDD 기본 형식으로 생성
                        base_number = dt.strftime('ACC%y%m%d')
                        
                        # 같은 날짜에 이미 있는 사고 수 확인 (cache 기준)
                        cursor.execute("""
                            SELECT COUNT(*) FROM accidents_cache 
                            WHERE accident_number LIKE %s
                        """, (base_number + '%',))
                        count = cursor.fetchone()[0]
                        
                        # 일련번호 추가 (01, 02, 03...)
                        sequence = str(count + 1).zfill(2)
                        data['accident_number'] = base_number + sequence
                        
                    except Exception as e:
                        # 오류 발생시 기본 자동 생성 방식 사용
                        data['accident_number'] = generate_manual_accident_number(cursor)
                else:
                    # 재해날짜가 없으면 기본 자동 생성 방식 사용
                    data['accident_number'] = generate_manual_accident_number(cursor)
                
                # 중복 확인 (cache 기준)
                if skip_duplicates and data.get('accident_number'):
                    try:
                        cursor.execute("SELECT COUNT(*) FROM accidents_cache WHERE accident_number = %s", (data['accident_number'],))
                        if (cursor.fetchone() or [0])[0] > 0:
                            continue
                    except Exception:
                        pass
                
                # 날짜 형식 처리 - 간단화
                if data.get('accident_date'):
                    date_str = str(data['accident_date']).strip()
                    if date_str and date_str != 'None':
                        data['accident_date'] = date_str
                    else:
                        # 날짜가 없으면 오늘 날짜로 설정
                        data['accident_date'] = get_korean_time().strftime('%Y-%m-%d')
                
                logging.info(f"매핑된 데이터: {data}")
                logging.info(f"동적 컬럼 데이터: {custom_data}")
                
                # 최소 필수 데이터 확인
                if not data.get('accident_number'):
                    logging.error(f"행 {row_idx}: 사고번호가 생성되지 않음")
                    continue
                
                # DB 저장 - SOT: accidents_cache
                try:
                    # 표준 키 정규화
                    accident_grade = data.pop('accident_level', None)
                    major_category = data.pop('accident_classification', None)
                    injury_type = data.pop('disaster_type', None)
                    injury_form = data.pop('disaster_form', None)

                    top = {
                        'accident_number': data.get('accident_number'),
                        'accident_name': data.get('accident_name', ''),
                        'workplace': data.get('workplace', ''),
                        'accident_grade': accident_grade or data.get('accident_grade', ''),
                        'major_category': major_category or data.get('major_category', ''),
                        'injury_form': injury_form or data.get('injury_form', ''),
                        'injury_type': injury_type or data.get('injury_type', ''),
                        'accident_date': data.get('accident_date', get_korean_time().strftime('%Y-%m-%d')),
                        'day_of_week': data.get('day_of_week', ''),
                        'report_date': data.get('accident_date', get_korean_time().strftime('%Y-%m-%d')),
                        'created_at': get_korean_time().strftime('%Y-%m-%d'),
                        'building': data.get('building', ''),
                        'floor': data.get('floor', ''),
                        'location_category': data.get('location_category', ''),
                        'location_detail': data.get('location_detail', ''),
                    }

                    # 나머지 키는 custom_data에 병합
                    extra = {k: v for k, v in data.items() if k not in top}
                    # 기존 dynamic custom_data와 합치기
                    full_custom = {}
                    full_custom.update(extra)
                    full_custom.update(custom_data)

                    cursor.execute(
                        """
                        INSERT INTO accidents_cache (
                            accident_number,
                            accident_name,
                            workplace,
                            accident_grade,
                            major_category,
                            injury_form,
                            injury_type,
                            accident_date,
                            day_of_week,
                            report_date,
                            created_at,
                            building,
                            floor,
                            location_category,
                            location_detail,
                            custom_data
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            top['accident_number'],
                            top['accident_name'],
                            top['workplace'],
                            top['accident_grade'],
                            top['major_category'],
                            top['injury_form'],
                            top['injury_type'],
                            top['accident_date'],
                            top['day_of_week'],
                            top['report_date'],
                            top['created_at'],
                            top['building'],
                            top['floor'],
                            top['location_category'],
                            top['location_detail'],
                            full_custom,
                        )
                    )

                except Exception as sql_error:
                    logging.error(f"accidents_cache INSERT 오류: {sql_error}")
                    raise sql_error
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"행 {row_idx}: {str(e)}")
                continue
        
        conn.commit()
        conn.close()
        
        result = {
            "success": True,
            "success_count": success_count,
            "error_count": error_count
        }
        
        if errors:
            result["errors"] = errors[:10]  # 최대 10개 오류만 반환
            
        return jsonify(result)
        
    except Exception as e:
        logging.error(f"엑셀 임포트 중 오류: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500

# ===== Follow SOP 엑셀 다운로드 API =====
@app.route('/api/follow-sop-export')
def export_follow_sop_excel():
    """Follow SOP 데이터 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        # DB 연결
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 섹션 정보 가져오기
        section_sql = f"""
            SELECT section_key, section_name, section_order
            FROM follow_sop_sections
            WHERE {sql_is_active_true('is_active', conn)}
              AND {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY section_order
        """
        try:
            cursor.execute(section_sql)
            sections = [dict(row) for row in cursor.fetchall()]
        except:
            # 섹션 테이블이 없으면 기본 섹션 사용
            sections = []

        # 동적 컬럼 정보 (활성+미삭제)
        where_c_active = sql_is_active_true('is_active', conn)
        where_c_notdel = sql_is_deleted_false('is_deleted', conn)
        dyn_sql = f"""
            SELECT * FROM follow_sop_column_config
            WHERE {where_c_active}
              AND {where_c_notdel}
            ORDER BY column_order
        """
        cursor.execute(dyn_sql)
        dynamic_columns_all = [dict(row) for row in cursor.fetchall()]

        # 섹션별로 컬럼 그룹핑 (섹션 순서 -> 섹션 내 컬럼 순서)
        dynamic_columns = []
        if sections:
            # 섹션 순서대로 컬럼 추가
            for section in sections:
                section_columns = [col for col in dynamic_columns_all if col.get('tab') == section['section_key']]
                dynamic_columns.extend(section_columns)
            # 섹션이 없는 컬럼들 추가
            no_section_columns = [col for col in dynamic_columns_all if not col.get('tab') or not any(s['section_key'] == col.get('tab') for s in sections)]
            dynamic_columns.extend(no_section_columns)
        else:
            # 섹션 정보가 없으면 기존 순서 사용
            dynamic_columns = dynamic_columns_all
        
        # Follow SOP 데이터 조회
        data_sql = f"""
            SELECT * FROM follow_sop
            WHERE {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY created_at DESC
        """
        cursor.execute(data_sql)
        data = [dict(row) for row in cursor.fetchall()]
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "Follow SOP"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        col_idx = 1
        
        # 기본 필드 (follow_sop 테이블에 맞게)
        basic_headers = ['점검번호', '등록일', '작성자']
        for header in basic_headers:
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            col_idx += 1
        
        # 채점 항목을 개별 컬럼으로 확장한 목록 구성
        def _expand_scoring_columns(_cols):
            out = []
            import json as _json
            for c in _cols:
                if c.get('column_type') == 'scoring':
                    sc = c.get('scoring_config')
                    if sc and isinstance(sc, str):
                        try: sc = _json.loads(sc)
                        except Exception: sc = {}
                    items = (sc or {}).get('items') or []
                    for it in items:
                        iid = it.get('id')
                        label = it.get('label') or iid
                        if not iid:
                            continue
                        out.append({
                            'column_key': f"{c['column_key']}__{iid}",
                            'column_name': f"{c.get('column_name', c.get('column_key'))} - {label}",
                            'column_type': 'number',
                            '_virtual': 1,
                            '_source_scoring_key': c['column_key'],
                            '_source_item_id': iid
                        })
                else:
                    out.append(dict(c))
            return out

        dyn_cols_list = [dict(x) for x in dynamic_columns]
        expanded_columns = _expand_scoring_columns(dyn_cols_list)

        # 스코어 총점 계산 준비
        import json as _json
        scoring_cols = [dict(c) for c in dyn_cols_list if dict(c).get('column_type') == 'scoring']
        score_total_cols = [dict(c) for c in dyn_cols_list if dict(c).get('column_type') == 'score_total']

        # 동적 컬럼 헤더
        for col in expanded_columns:
            cell = ws.cell(row=1, column=col_idx, value=col['column_name'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            col_idx += 1
        
        # 데이터 작성
        for row_idx, row in enumerate(data, 2):
            # 먼저 row를 dict로 변환
            row_dict = dict(row)
            col_idx = 1
            
            # 기본 필드 - 실제 컬럼명 사용
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('work_req_no', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('created_at', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('created_by', ''))
            col_idx += 1
            
            # custom_data 파싱 (PostgreSQL 전용)
            custom_data = row_dict.get('custom_data', {})
            if not isinstance(custom_data, dict):
                custom_data = {}

            # 동적 컬럼 데이터 - 드롭다운/가상 채점 항목 포함
            # 드롭다운 매핑 지원
            def _map_value(col, value):
                # 팝업형 값(dict) → name
                if isinstance(value, dict):
                    return value.get('name', str(value))
                # 리스트 ⇒ JSON 문자열 (모든 리스트 처리)
                if isinstance(value, list):
                    if not value:  # 빈 리스트는 빈 문자열로
                        return ''
                    try:
                        return json.dumps(value, ensure_ascii=False)
                    except Exception:
                        return str(value)
                # 드롭다운 코드 → 표시값
                if col['column_type'] == 'dropdown' and value:
                    opts = get_dropdown_options_for_display('follow_sop', col['column_key'])
                    if opts:
                        for opt in opts:
                            if opt['code'] == value:
                                return opt['value']
                return value if value is not None else ''

            for col in expanded_columns:
                if col.get('_virtual') == 1:
                    src = col.get('_source_scoring_key')
                    iid = col.get('_source_item_id')
                    group_obj = custom_data.get(src, {})
                    if isinstance(group_obj, str):
                        try:
                            group_obj = json.loads(group_obj)
                        except Exception:
                            group_obj = {}
                    v = 0
                    if isinstance(group_obj, dict):
                        v = group_obj.get(iid, 0)
                    ws.cell(row=row_idx, column=col_idx, value=v)
                else:
                    if col.get('column_type') == 'score_total':
                        # 총점 계산: include_keys 우선, 없으면 total_key 기준
                        try:
                            stc = col
                            conf = stc.get('scoring_config')
                            if conf and isinstance(conf, str):
                                try: conf = _json.loads(conf)
                                except Exception: conf = {}
                            base = (conf or {}).get('base_score', 100)
                            total = base
                            include_keys = (conf or {}).get('include_keys') or []
                            if include_keys:
                                for key in include_keys:
                                    sc_col = next((c for c in scoring_cols if c.get('column_key') == key), None)
                                    if not sc_col:
                                        continue
                                    sconf = sc_col.get('scoring_config')
                                    if sconf and isinstance(sconf, str):
                                        try: sconf = _json.loads(sconf)
                                        except Exception: sconf = {}
                                    items_cfg = (sconf or {}).get('items') or []
                                    group_obj = custom_data.get(key, {})
                                    if isinstance(group_obj, str):
                                        try: group_obj = _json.loads(group_obj)
                                        except Exception: group_obj = {}
                                    for it in items_cfg:
                                        iid = it.get('id')
                                        delta = float(it.get('per_unit_delta') or 0)
                                        cnt = 0
                                        if isinstance(group_obj, dict) and iid in group_obj:
                                            try: cnt = int(group_obj.get(iid) or 0)
                                            except Exception: cnt = 0
                                        total += cnt * delta
                            else:
                                total_key = (conf or {}).get('total_key') or 'default'
                                for sc_col in scoring_cols:
                                    sconf = sc_col.get('scoring_config')
                                    if sconf and isinstance(sconf, str):
                                        try: sconf = _json.loads(sconf)
                                        except Exception: sconf = {}
                                    if ((sconf or {}).get('total_key') or 'default') != total_key:
                                        continue
                                    items_cfg = (sconf or {}).get('items') or []
                                    group_obj = custom_data.get(sc_col.get('column_key'), {})
                                    if isinstance(group_obj, str):
                                        try: group_obj = _json.loads(group_obj)
                                        except Exception: group_obj = {}
                                    for it in items_cfg:
                                        iid = it.get('id')
                                        delta = float(it.get('per_unit_delta') or 0)
                                        cnt = 0
                                        if isinstance(group_obj, dict) and iid in group_obj:
                                            try: cnt = int(group_obj.get(iid) or 0)
                                            except Exception: cnt = 0
                                        total += cnt * delta
                            ws.cell(row=row_idx, column=col_idx, value=total)
                        except Exception:
                            ws.cell(row=row_idx, column=col_idx, value='')
                    else:
                        v = custom_data.get(col['column_key'], '')
                        ws.cell(row=row_idx, column=col_idx, value=_map_value(col, v))
                col_idx += 1
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"follow_sop_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        conn.close()
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        
    except Exception as e:
        import traceback
        logging.error(f"Follow SOP 엑셀 다운로드 중 오류: {e}")
        logging.error(traceback.format_exc())
        if conn:
            conn.close()
        return jsonify({"success": False, "message": str(e)}), 500

# ===== Safe Workplace 엑셀 다운로드 API =====
@app.route('/api/safe-workplace-export')
def export_safe_workplace_excel():
    """Safe Workplace 데이터 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io

        conn = get_db_connection()
        cursor = conn.cursor()

        section_sql = f"""
            SELECT section_key, section_name, section_order
            FROM safe_workplace_sections
            WHERE {sql_is_active_true('is_active', conn)}
              AND {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY section_order
        """
        try:
            cursor.execute(section_sql)
            sections = [dict(row) for row in cursor.fetchall()]
        except Exception:
            sections = []

        where_c_active = sql_is_active_true('is_active', conn)
        where_c_notdel = sql_is_deleted_false('is_deleted', conn)
        cursor.execute(f"""
            SELECT * FROM safe_workplace_column_config
            WHERE {where_c_active}
              AND {where_c_notdel}
            ORDER BY column_order
        """)
        dynamic_columns_all = [dict(row) for row in cursor.fetchall()]

        dynamic_columns = []
        if sections:
            for section in sections:
                section_columns = [col for col in dynamic_columns_all if col.get('tab') == section['section_key']]
                dynamic_columns.extend(section_columns)
            no_section_columns = [col for col in dynamic_columns_all if not col.get('tab') or not any(s['section_key'] == col.get('tab') for s in sections)]
            dynamic_columns.extend(no_section_columns)
        else:
            dynamic_columns = dynamic_columns_all

        cursor.execute(f"""
            SELECT * FROM safe_workplace
            WHERE {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY created_at DESC
        """)
        data = [dict(row) for row in cursor.fetchall()]

        wb = Workbook()
        ws = wb.active
        ws.title = "Safe Workplace"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        col_idx = 1
        basic_headers = ['점검번호', '등록일', '작성자']
        for header in basic_headers:
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            col_idx += 1

        def _expand_scoring_columns(_cols):
            out = []
            import json as _json
            for c in _cols:
                if c.get('column_type') == 'scoring':
                    sc = c.get('scoring_config')
                    if sc and isinstance(sc, str):
                        try:
                            sc = _json.loads(sc)
                        except Exception:
                            sc = {}
                    items = (sc or {}).get('items') or []
                    for it in items:
                        iid = it.get('id')
                        label = it.get('label') or iid
                        if not iid:
                            continue
                        out.append({
                            'column_key': f"{c['column_key']}__{iid}",
                            'column_name': f"{c.get('column_name', c.get('column_key'))} - {label}",
                            'column_type': 'number',
                            '_virtual': 1,
                            '_source_scoring_key': c['column_key'],
                            '_source_item_id': iid
                        })
                else:
                    out.append(dict(c))
            return out

        dyn_cols_list = [dict(x) for x in dynamic_columns]
        expanded_columns = _expand_scoring_columns(dyn_cols_list)

        import json as _json
        scoring_cols = [dict(c) for c in dyn_cols_list if dict(c).get('column_type') == 'scoring']
        score_total_cols = [dict(c) for c in dyn_cols_list if dict(c).get('column_type') == 'score_total']

        for col in expanded_columns:
            cell = ws.cell(row=1, column=col_idx, value=col['column_name'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            col_idx += 1

        for row_idx, row in enumerate(data, 2):
            row_dict = dict(row)
            col_idx = 1

            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('safeplace_no', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('created_at', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('created_by', ''))
            col_idx += 1

            custom_data = row_dict.get('custom_data', {})
            if not isinstance(custom_data, dict):
                try:
                    custom_data = _json.loads(custom_data) if custom_data else {}
                except Exception:
                    custom_data = {}

            def _map_value(col, value):
                if isinstance(value, dict):
                    return value.get('name') or str(value)
                if isinstance(value, list):
                    if not value:
                        return ''
                    try:
                        return json.dumps(value, ensure_ascii=False)
                    except Exception:
                        return str(value)
                if col['column_type'] == 'dropdown' and value:
                    opts = get_dropdown_options_for_display('safe_workplace', col['column_key'])
                    if opts:
                        for opt in opts:
                            if opt['code'] == value:
                                return opt['value']
                return value if value is not None else ''

            for col in expanded_columns:
                if col.get('_virtual') == 1:
                    src = col.get('_source_scoring_key')
                    iid = col.get('_source_item_id')
                    group_obj = custom_data.get(src, {})
                    if isinstance(group_obj, str):
                        try:
                            group_obj = _json.loads(group_obj)
                        except Exception:
                            group_obj = {}
                    value = 0
                    if isinstance(group_obj, dict):
                        value = group_obj.get(iid, 0)
                    ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    if col.get('column_type') == 'score_total':
                        try:
                            conf = col.get('scoring_config')
                            if conf and isinstance(conf, str):
                                conf = _json.loads(conf)
                            base = (conf or {}).get('base_score', 100)
                            include_keys = (conf or {}).get('include_keys') or []
                            total = base
                            if include_keys:
                                for key in include_keys:
                                    sc_col = next((c for c in scoring_cols if c.get('column_key') == key), None)
                                    if not sc_col:
                                        continue
                                    sconf = sc_col.get('scoring_config')
                                    if sconf and isinstance(sconf, str):
                                        try:
                                            sconf = _json.loads(sconf)
                                        except Exception:
                                            sconf = {}
                                    items_cfg = (sconf or {}).get('items') or []
                                    group_obj = custom_data.get(key, {})
                                    if isinstance(group_obj, str):
                                        try:
                                            group_obj = _json.loads(group_obj)
                                        except Exception:
                                            group_obj = {}
                                    for it in items_cfg:
                                        iid = it.get('id')
                                        delta = float(it.get('per_unit_delta') or 0)
                                        cnt = 0
                                        if isinstance(group_obj, dict) and iid in group_obj:
                                            try:
                                                cnt = int(group_obj.get(iid) or 0)
                                            except Exception:
                                                cnt = 0
                                        total += cnt * delta
                            ws.cell(row=row_idx, column=col_idx, value=total)
                        except Exception:
                            ws.cell(row=row_idx, column=col_idx, value='')
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=_map_value(col, custom_data.get(col['column_key'], '')))
                col_idx += 1

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"safe_workplace_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"

        conn.close()

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

    except Exception as e:
        import traceback
        logging.error(f"Safe Workplace 엑셀 다운로드 중 오류: {e}")
        logging.error(traceback.format_exc())
        if conn:
            conn.close()
        return jsonify({"success": False, "message": str(e)}), 500

# ===== Full Process 엑셀 다운로드 API =====
@app.route('/api/full-process-export')
def export_full_process_excel():
    """Full Process 데이터 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        # DB 연결
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 섹션 정보 가져오기
        section_sql = f"""
            SELECT section_key, section_name, section_order
            FROM full_process_sections
            WHERE {sql_is_active_true('is_active', conn)}
              AND {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY section_order
        """
        try:
            cursor.execute(section_sql)
            sections = [dict(row) for row in cursor.fetchall()]
        except:
            # 섹션 테이블이 없으면 기본 섹션 사용
            sections = []

        # 동적 컬럼 정보 (활성+미삭제)
        where_c_active = sql_is_active_true('is_active', conn)
        where_c_notdel = sql_is_deleted_false('is_deleted', conn)
        dyn_sql = f"""
            SELECT * FROM full_process_column_config
            WHERE {where_c_active}
              AND {where_c_notdel}
            ORDER BY column_order
        """
        cursor.execute(dyn_sql)
        dynamic_columns_all = [dict(row) for row in cursor.fetchall()]

        # 섹션별로 컬럼 그룹핑 (섹션 순서 -> 섹션 내 컬럼 순서)
        dynamic_columns = []
        if sections:
            # 섹션 순서대로 컬럼 추가
            for section in sections:
                section_columns = [col for col in dynamic_columns_all if col.get('tab') == section['section_key']]
                dynamic_columns.extend(section_columns)
            # 섹션이 없는 컬럼들 추가
            no_section_columns = [col for col in dynamic_columns_all if not col.get('tab') or not any(s['section_key'] == col.get('tab') for s in sections)]
            dynamic_columns.extend(no_section_columns)
        else:
            # 섹션 정보가 없으면 기존 순서 사용
            dynamic_columns = dynamic_columns_all
        
        # Full Process 데이터 조회
        data_sql = f"""
            SELECT * FROM full_process
            WHERE {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY created_at DESC
        """
        cursor.execute(data_sql)
        data = [dict(row) for row in cursor.fetchall()]
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "Full Process"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        col_idx = 1
        
        # 기본 필드 (full_process 테이블에 맞게)
        basic_headers = ['프로세스 번호', '작성일', '작성자']
        for header in basic_headers:
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            col_idx += 1
        
        # 동적 컬럼 헤더 (채점 항목 확장)
        import json as _json
        def _expand_scoring_columns(_cols):
            out = []
            for c in _cols:
                if c.get('column_type') == 'scoring':
                    sc = c.get('scoring_config')
                    if sc and isinstance(sc, str):
                        try: sc = _json.loads(sc)
                        except Exception: sc = {}
                    items = (sc or {}).get('items') or []
                    for it in items:
                        iid = it.get('id')
                        label = it.get('label') or iid
                        if not iid:
                            continue
                        out.append({
                            'column_key': f"{c['column_key']}__{iid}",
                            'column_name': f"{c.get('column_name', c.get('column_key'))} - {label}",
                            'column_type': 'number',
                            '_virtual': 1,
                            '_source_scoring_key': c['column_key'],
                            '_source_item_id': iid
                        })
                else:
                    out.append(dict(c))
            return out

        dyn_cols_list = [dict(x) for x in dynamic_columns]
        expanded_columns = _expand_scoring_columns(dyn_cols_list)
        scoring_cols = [dict(c) for c in dyn_cols_list if dict(c).get('column_type') == 'scoring']
        score_total_cols = [dict(c) for c in dyn_cols_list if dict(c).get('column_type') == 'score_total']

        for col in expanded_columns:
            cell = ws.cell(row=1, column=col_idx, value=col['column_name'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            col_idx += 1
        
        # 데이터 작성
        for row_idx, row in enumerate(data, 2):
            # 먼저 row를 dict로 변환
            row_dict = dict(row)
            col_idx = 1
            
            # 기본 필드 - 실제 컬럼명 사용
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('fullprocess_number', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('created_at', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('created_by', ''))
            col_idx += 1
            
            # custom_data 파싱 (PostgreSQL 전용)
            custom_data = row_dict.get('custom_data', {})
            if not isinstance(custom_data, dict):
                custom_data = {}
            
            # 동적 컬럼 데이터 (확장 포함)
            # 드롭다운 매핑 지원
            def _map_value(col, value):
                if isinstance(value, dict):
                    return value.get('name', str(value))
                if col['column_type'] == 'list' and isinstance(value, list):
                    if not value:  # 빈 리스트는 빈 문자열로
                        return ''
                    try:
                        return json.dumps(value, ensure_ascii=False)
                    except Exception:
                        return str(value)
                if col['column_type'] == 'dropdown' and value:
                    opts = get_dropdown_options_for_display('full_process', col['column_key'])
                    if opts:
                        for opt in opts:
                            if opt['code'] == value:
                                return opt['value']
                return value

            for col in expanded_columns:
                if col.get('_virtual') == 1:
                    src = col.get('_source_scoring_key')
                    iid = col.get('_source_item_id')
                    group_obj = custom_data.get(src, {})
                    if isinstance(group_obj, str):
                        try:
                            group_obj = json.loads(group_obj)
                        except Exception:
                            group_obj = {}
                    v = 0
                    if isinstance(group_obj, dict):
                        v = group_obj.get(iid, 0)
                    ws.cell(row=row_idx, column=col_idx, value=v)
                else:
                    if col.get('column_type') == 'score_total':
                        try:
                            stc = col
                            conf = stc.get('scoring_config')
                            if conf and isinstance(conf, str):
                                try: conf = _json.loads(conf)
                                except Exception: conf = {}
                            base = (conf or {}).get('base_score', 100)
                            total = base
                            include_keys = (conf or {}).get('include_keys') or []
                            if include_keys:
                                for key in include_keys:
                                    sc_col = next((c for c in scoring_cols if c.get('column_key') == key), None)
                                    if not sc_col:
                                        continue
                                    sconf = sc_col.get('scoring_config')
                                    if sconf and isinstance(sconf, str):
                                        try: sconf = _json.loads(sconf)
                                        except Exception: sconf = {}
                                    items_cfg = (sconf or {}).get('items') or []
                                    group_obj = custom_data.get(key, {})
                                    if isinstance(group_obj, str):
                                        try: group_obj = _json.loads(group_obj)
                                        except Exception: group_obj = {}
                                    for it in items_cfg:
                                        iid = it.get('id')
                                        delta = float(it.get('per_unit_delta') or 0)
                                        cnt = 0
                                        if isinstance(group_obj, dict) and iid in group_obj:
                                            try: cnt = int(group_obj.get(iid) or 0)
                                            except Exception: cnt = 0
                                        total += cnt * delta
                            else:
                                total_key = (conf or {}).get('total_key') or 'default'
                                for sc_col in scoring_cols:
                                    sconf = sc_col.get('scoring_config')
                                    if sconf and isinstance(sconf, str):
                                        try: sconf = _json.loads(sconf)
                                        except Exception: sconf = {}
                                    if ((sconf or {}).get('total_key') or 'default') != total_key:
                                        continue
                                    items_cfg = (sconf or {}).get('items') or []
                                    group_obj = custom_data.get(sc_col.get('column_key'), {})
                                    if isinstance(group_obj, str):
                                        try: group_obj = _json.loads(group_obj)
                                        except Exception: group_obj = {}
                                    for it in items_cfg:
                                        iid = it.get('id')
                                        delta = float(it.get('per_unit_delta') or 0)
                                        cnt = 0
                                        if isinstance(group_obj, dict) and iid in group_obj:
                                            try: cnt = int(group_obj.get(iid) or 0)
                                            except Exception: cnt = 0
                                        total += cnt * delta
                            ws.cell(row=row_idx, column=col_idx, value=total)
                        except Exception:
                            ws.cell(row=row_idx, column=col_idx, value='')
                    else:
                        v = custom_data.get(col['column_key'], '')
                        ws.cell(row=row_idx, column=col_idx, value=_map_value(col, v))
                col_idx += 1
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"full_process_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        conn.close()
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        
    except Exception as e:
        import traceback
        logging.error(f"Full Process 엑셀 다운로드 중 오류: {e}")
        logging.error(traceback.format_exc())
        if conn:
            conn.close()
        return jsonify({"success": False, "message": str(e)}), 500

# ===== Safety Instruction 엑셀 다운로드 API =====
@app.route('/api/safety-instruction-export')
def export_safety_instruction_excel():
    """Safety Instruction 데이터 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        # DB 연결
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 섹션 정보 가져오기
        section_sql = f"""
            SELECT section_key, section_name, section_order
            FROM safety_instruction_sections
            WHERE {sql_is_active_true('is_active', conn)}
              AND {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY section_order
        """
        try:
            cursor.execute(section_sql)
            sections = [dict(row) for row in cursor.fetchall()]
        except:
            # 섹션 테이블이 없거나 section_config 테이블 확인
            try:
                section_sql = f"""
                    SELECT section_key, section_name, section_order
                    FROM section_config
                    WHERE board_type = 'safety_instruction'
                      AND {sql_is_active_true('is_active', conn)}
                      AND {sql_is_deleted_false('is_deleted', conn)}
                    ORDER BY section_order
                """
                cursor.execute(section_sql)
                sections = [dict(row) for row in cursor.fetchall()]
            except:
                sections = []

        # 동적 컬럼 정보 (활성+미삭제)
        where_c_active = sql_is_active_true('is_active', conn)
        where_c_notdel = sql_is_deleted_false('is_deleted', conn)
        dyn_sql = f"""
            SELECT * FROM safety_instruction_column_config
            WHERE {where_c_active}
              AND {where_c_notdel}
            ORDER BY column_order
        """
        cursor.execute(dyn_sql)
        dynamic_columns_all = [dict(row) for row in cursor.fetchall()]

        # 섹션별로 컬럼 그룹핑 (섹션 순서 -> 섹션 내 컬럼 순서)
        dynamic_columns = []
        if sections:
            # 섹션 순서대로 컬럼 추가
            for section in sections:
                section_columns = [col for col in dynamic_columns_all if col.get('tab') == section['section_key']]
                dynamic_columns.extend(section_columns)
            # 섹션이 없는 컬럼들 추가
            no_section_columns = [col for col in dynamic_columns_all if not col.get('tab') or not any(s['section_key'] == col.get('tab') for s in sections)]
            dynamic_columns.extend(no_section_columns)
        else:
            # 섹션 정보가 없으면 기존 순서 사용
            dynamic_columns = dynamic_columns_all
        
        # Safety Instruction 데이터 조회 - 메인 테이블 사용
        data_sql = f"""
            SELECT * FROM safety_instructions
            WHERE {sql_is_deleted_false('is_deleted', conn)}
            ORDER BY created_at DESC, issue_number DESC
        """
        cursor.execute(data_sql)
        data = [dict(row) for row in cursor.fetchall()]
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "Safety Instructions"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        col_idx = 1
        
        # 기본 필드 (safety_instructions 테이블의 주요 컬럼)
        basic_headers = ['발부번호', '발부자', '위반일자', '징계일자', '피징계자']
        for header in basic_headers:
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            col_idx += 1
        
        # 동적 컬럼 - 삭제되지 않은 것만 (기본 필드 제외)
        basic_column_keys = ['issue_number', 'issuer', 'violation_date', 'discipline_date', 'disciplined_person']
        for col in dynamic_columns:
            if col['column_key'] not in basic_column_keys:
                cell = ws.cell(row=1, column=col_idx, value=col['column_name'])
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                col_idx += 1
        
        # 데이터를 목록 페이지와 동일한 방식으로 처리
        data_list = []
        for row in data:
            row_dict = dict(row)
            
            # custom_data 파싱 및 플래튼 (목록 페이지와 동일)
            if row_dict.get('custom_data'):
                try:
                    import json as pyjson
                    raw = row_dict.get('custom_data')
                    
                    # dict/str 분기 처리
                    if isinstance(raw, dict):
                        custom_data = raw
                    elif isinstance(raw, str):
                        custom_data = pyjson.loads(raw) if raw else {}
                    else:
                        custom_data = {}
                    
                    # 기본 필드를 보호하면서 custom_data 병합
                    BASE_FIELDS = {'issue_number', 'created_at', 'updated_at', 'is_deleted', 'synced_at'}
                    for k, v in custom_data.items():
                        if k not in BASE_FIELDS:
                            row_dict[k] = v
                except Exception as e:
                    logging.error(f"Custom data parsing error: {e}")
            
            data_list.append(row_dict)
        
        # smart_apply_mappings 적용 (드롭다운 코드를 라벨로 변환)
        from common_mapping import smart_apply_mappings
        if data_list:
            data_list = smart_apply_mappings(
                data_list, 
                'safety_instruction', 
                [dict(col) for col in dynamic_columns],
                DB_PATH
            )
        
        # 데이터 작성
        for row_idx, row_dict in enumerate(data_list, 2):
            col_idx = 1

            # 기본 필드
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('issue_number', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('issuer', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('violation_date', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('discipline_date', ''))
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get('disciplined_person', ''))
            col_idx += 1

            # 동적 컬럼 데이터 - 기본 필드 제외하고 처리
            basic_column_keys = ['issue_number', 'issuer', 'violation_date', 'discipline_date', 'disciplined_person']
            
            for col in dynamic_columns:
                if col['column_key'] not in basic_column_keys:
                    col_key = col['column_key']
                    value = row_dict.get(col_key, '')

                    # 리스트 처리 (column_type과 관계없이 실제 값이 리스트면 처리)
                    if isinstance(value, list):
                        if not value:  # 빈 리스트는 빈 문자열로
                            value = ''
                        else:
                            try:
                                value = json.dumps(value, ensure_ascii=False)
                            except Exception:
                                value = str(value)
                    # 딕셔너리 처리
                    elif isinstance(value, dict):
                        if not value:  # 빈 딕셔너리도 빈 문자열로
                            value = ''
                        else:
                            value = value.get('name', str(value))
                    # None 처리
                    elif value is None:
                        value = ''

                    ws.cell(row=row_idx, column=col_idx, value=value)
                    col_idx += 1
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"safety_instruction_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        conn.close()
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        
    except Exception as e:
        import traceback
        logging.error(f"Safety Instruction 엑셀 다운로드 중 오류: {e}")
        logging.error(traceback.format_exc())
        if conn:
            conn.close()
        return jsonify({"success": False, "message": str(e)}), 500

# ===== 기준정보 변경요청 엑셀 다운로드 API =====
@app.route('/api/change-requests/export')
def export_change_requests_excel():
    """기준정보 변경요청 데이터 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        import io
        
        # 검색 조건 가져오기
        company_name = request.args.get('company_name', '')
        business_number = request.args.get('business_number', '')
        status = request.args.get('status', '')
        created_date_start = request.args.get('created_date_start', '')
        created_date_end = request.args.get('created_date_end', '')
        
        # DB 연결
        conn = get_db_connection()
        
        # 섹션 정보 가져오기
        section_sql = f"""
            SELECT section_key, section_name, section_order
            FROM section_config
            WHERE board_type = 'change_request'
              AND {sql_is_active_true('is_active', conn)}
            ORDER BY section_order
        """
        sections = conn.execute(section_sql).fetchall()
        
        # 동적 컬럼 정보 (활성+미삭제)
        where_c_active = sql_is_active_true('is_active', conn)
        where_c_notdel = sql_is_deleted_false('is_deleted', conn)
        dyn_sql = f"""
            SELECT * FROM change_request_column_config
            WHERE {where_c_active}
              AND {where_c_notdel}
            ORDER BY column_order
        """
        dynamic_columns_rows = conn.execute(dyn_sql).fetchall()
        dynamic_columns_all = [dict(row) for row in dynamic_columns_rows]
        
        # 섹션 기반으로 컬럼 정렬
        dynamic_columns = []
        if sections:
            # 각 섹션에 속하는 컬럼들을 순서대로 추가
            for section in sections:
                section_columns = [col for col in dynamic_columns_all 
                                 if col.get('tab') == section['section_key']]
                # 섹션 내에서 column_order로 정렬
                section_columns.sort(key=lambda x: x.get('column_order', 0))
                dynamic_columns.extend(section_columns)
            
            # 섹션이 없는 컬럼들을 마지막에 추가
            no_section_columns = [col for col in dynamic_columns_all 
                                if not col.get('tab') or 
                                not any(section['section_key'] == col.get('tab') for section in sections)]
            no_section_columns.sort(key=lambda x: x.get('column_order', 0))
            dynamic_columns.extend(no_section_columns)
        else:
            # 섹션이 없으면 원래 순서대로
            dynamic_columns = dynamic_columns_all
        
        # 변경요청 데이터 조회
        query = f"""
            SELECT * FROM partner_change_requests 
            WHERE {sql_is_deleted_false('is_deleted', conn)}
        """
        params = []
        
        if company_name:
            query += " AND company_name LIKE %s"
            params.append(f"%{company_name}%")
        
        if business_number:
            query += " AND business_number LIKE %s"
            params.append(f"%{business_number}%")
            
        if status:
            query += " AND status = %s"
            params.append(status)
        
        if created_date_start:
            query += " AND DATE(created_at) >= %s"
            params.append(created_date_start)
        
        if created_date_end:
            query += " AND DATE(created_at) <= %s"
            params.append(created_date_end)
        
        # 등록일 기준 최신순 정렬
        query += " ORDER BY created_at DESC, id DESC"
        
        change_requests = conn.execute(query, params).fetchall()
        
        # 드롭다운 코드-값 매핑 함수
        def get_display_value(column_key, code_value):
            if not code_value or code_value == '':
                return ''
            try:
                options = get_dropdown_options_for_display('change_request', column_key)
                if options:
                    # 리스트 형태의 옵션에서 코드에 해당하는 값 찾기
                    for option in options:
                        if option['code'] == code_value:
                            return option['value']
                return code_value
            except:
                return code_value
        
        # 날짜 형식 정리 함수 (시분초 제거)
        def format_date(date_value):
            if not date_value:
                return ''
            date_str = str(date_value)
            if ' ' in date_str:
                return date_str.split(' ')[0]  # 2025-09-07 0:00:00 → 2025-09-07
            return date_str
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "기준정보 변경요청"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # 기본 헤더
        headers = [
            '요청번호', '회사명', '사업자번호', '상태', '등록일', '수정일'
        ]
        
        # 동적 컬럼 헤더 추가
        for col in dynamic_columns:
            headers.append(col['column_name'])
        
        # 헤더 쓰기
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # 데이터 쓰기
        for row_idx, request_row in enumerate(change_requests, 2):
            request_data = dict(request_row)
            
            # 기본 필드 쓰기
            ws.cell(row=row_idx, column=1, value=request_data.get('request_number', ''))
            ws.cell(row=row_idx, column=2, value=request_data.get('company_name', ''))
            ws.cell(row=row_idx, column=3, value=request_data.get('business_number', ''))
            
            # 상태 코드를 실제 값으로 변환
            status_value = request_data.get('status', '')
            status_display = get_display_value('status', status_value) if status_value else ''
            ws.cell(row=row_idx, column=4, value=status_display)
            
            ws.cell(row=row_idx, column=5, value=format_date(request_data.get('created_at', '')))
            ws.cell(row=row_idx, column=6, value=format_date(request_data.get('updated_at', '')))
            
            # 동적 컬럼 데이터 쓰기
            custom_data = {}
            if request_data.get('custom_data'):
                try:
                    if isinstance(request_data['custom_data'], str):
                        custom_data = pyjson.loads(request_data['custom_data'])
                    else:
                        custom_data = request_data['custom_data']  # PostgreSQL JSONB
                except:
                    custom_data = {}
            
            for col_idx, col in enumerate(dynamic_columns, 7):  # 7번째 컬럼부터
                value = custom_data.get(col['column_key'], '')
                
                # popup 타입 데이터 처리
                if isinstance(value, dict):
                    if 'name' in value:
                        value = value['name']
                    else:
                        value = str(value)

                # 리스트 타입 처리
                elif isinstance(value, list):
                    if not value:  # 빈 리스트는 빈 문자열로
                        value = ''
                    else:
                        try:
                            value = pyjson.dumps(value, ensure_ascii=False)
                        except Exception:
                            value = str(value)

                # 드롭다운 타입인 경우 코드를 실제 값으로 변환
                elif col['column_type'] == 'dropdown' and value:
                    value = get_display_value(col['column_key'], value)
                
                # 날짜 타입인 경우 시분초 제거
                elif col['column_type'] in ['date', 'datetime'] and value:
                    value = format_date(value)
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 메모리에서 엑셀 파일 생성
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 파일명 생성 (한국시간 기준)
        korean_time = get_korean_time()
        filename = f"기준정보_변경요청_{korean_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        conn.close()
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logging.error(f"변경요청 엑셀 다운로드 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ===== 협력사 엑셀 다운로드 API =====
@app.route('/api/partners/export')
def export_partners_to_excel():
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        import io
        
        # 검색 조건 가져오기
        company_name = request.args.get('company_name', '')
        business_number = request.args.get('business_number', '')
        business_type_major = request.args.get('business_type_major', '')
        business_type_minor = request.args.get('business_type_minor', '')
        workers_min = request.args.get('workers_min', '')
        workers_max = request.args.get('workers_max', '')
        
        # 협력사 데이터 조회 (partner_standards 함수와 동일한 로직)
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # 쿼리 구성 (삭제되지 않은 데이터만)
            query = f"SELECT * FROM partners_cache WHERE {sql_is_deleted_false('is_deleted', conn)}"
            params = []
            
            if company_name:
                query += " AND company_name LIKE %s"
                params.append(f'%{company_name}%')
            
            if business_number:
                query += " AND business_number LIKE %s"
                params.append(f'%{business_number}%')
                
            if business_type_major:
                query += " AND business_type_major = %s"
                params.append(business_type_major)
                
            if business_type_minor:
                query += " AND business_type_minor = %s"
                params.append(business_type_minor)
                
            if workers_min:
                try:
                    min_val = int(workers_min)
                    query += " AND permanent_workers >= %s"
                    params.append(min_val)
                except ValueError:
                    pass
                    
            if workers_max:
                try:
                    max_val = int(workers_max)
                    query += " AND permanent_workers <= %s"
                    params.append(max_val)
                except ValueError:
                    pass
            
            query += " ORDER BY company_name"
            
            partners = cursor.execute(query, params).fetchall()
            
            # 엑셀 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "협력사 기준정보"
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            # 헤더 작성
            headers = [
                '협력사명', '사업자번호', 'Class', '업종(대분류)', '업종(소분류)',
                '위험작업여부', '대표자성명', '주소', '평균연령', '매출액', 
                '거래차수', '상시근로자'
            ]
            
            # 헤더 쓰기
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            # 데이터 쓰기
            for row_idx, partner_row in enumerate(partners, 2):
                partner = dict(partner_row)
                
                ws.cell(row=row_idx, column=1, value=partner.get('company_name', ''))
                ws.cell(row=row_idx, column=2, value=partner.get('business_number', ''))
                ws.cell(row=row_idx, column=3, value=partner.get('partner_class', ''))
                ws.cell(row=row_idx, column=4, value=partner.get('business_type_major', ''))
                ws.cell(row=row_idx, column=5, value=partner.get('business_type_minor', ''))
                
                # 위험작업여부 처리
                hazard_work = partner.get('hazard_work_flag', '')
                hazard_text = '예' if hazard_work == 'O' else '아니오' if hazard_work == 'X' else ''
                ws.cell(row=row_idx, column=6, value=hazard_text)
                
                ws.cell(row=row_idx, column=7, value=partner.get('representative', ''))
                ws.cell(row=row_idx, column=8, value=partner.get('address', ''))
                ws.cell(row=row_idx, column=9, value=partner.get('average_age', ''))
                
                # 매출액 처리 (억원 단위)
                revenue = partner.get('annual_revenue')
                if revenue:
                    revenue_text = f"{revenue // 100000000}억원"
                else:
                    revenue_text = ''
                ws.cell(row=row_idx, column=10, value=revenue_text)
                
                ws.cell(row=row_idx, column=11, value=partner.get('transaction_count', ''))
                
                # 상시근로자 처리
                workers = partner.get('permanent_workers')
                workers_text = f"{workers}명" if workers else ''
                ws.cell(row=row_idx, column=12, value=workers_text)
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일을 메모리에 저장
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            conn.close()
            
            # 파일명 생성
            filename = f"partners_list_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # 다운로드 응답
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as db_error:
            logging.error(f"협력사 데이터 조회 중 오류: {db_error}")
            # 더미 데이터로 대체
            wb = Workbook()
            ws = wb.active
            ws.title = "협력사 기준정보"
            
            # 헤더만 있는 빈 파일 생성
            headers = [
                '협력사명', '사업자번호', 'Class', '업종(대분류)', '업종(소분류)',
                '위험작업여부', '대표자성명', '주소', '평균연령', '매출액', 
                '거래차수', '상시근로자'
            ]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            # 샘플 데이터 1행 추가
            sample_data = [
                '샘플 협력사', '123-45-67890', 'A', '제조업', '전자제품',
                '예', '김대표', '서울시 강남구', '35', '100억원', '5', '50명'
            ]
            for col_idx, value in enumerate(sample_data, 1):
                ws.cell(row=2, column=col_idx, value=value)
            
            # 컬럼 너비 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"partners_list_{get_korean_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
    
    except Exception as e:
        logging.error(f"협력사 엑셀 다운로드 중 오류: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500

# ===== 협력사 삭제 API =====
@app.route('/api/partners/delete', methods=['POST'])
def delete_partners():
    try:
        data = request.get_json()
        business_numbers = data.get('business_numbers', [])
        
        if not business_numbers:
            return jsonify({"success": False, "message": "삭제할 협력사가 선택되지 않았습니다."}), 400
        
        # DB 연결
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Soft delete (is_deleted = 1로 설정)
        placeholders = ','.join(['%s'] * len(business_numbers))
        cursor.execute(f"""
            UPDATE partners_cache 
            SET is_deleted = 1 
            WHERE business_number IN ({placeholders})
        """, business_numbers)
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}개의 협력사가 삭제되었습니다.",
            "reload": True  # 페이지 새로고침 필요 플래그 추가
        })
        
    except Exception as e:
        logging.error(f"협력사 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# 새로운 메뉴들의 라우트
@app.route('/work-safety')
def work_safety():
    """작업안전 현황 페이지"""
    return render_template('work-safety.html', menu=menu)

@app.route('/risk-assessment')
def risk_assessment():
    """위험성평가 현황 페이지"""
    return render_template('risk-assessment.html', menu=menu)

@app.route('/qualification-assessment')
def qualification_assessment():
    """적격성평가 현황 페이지"""
    return render_template('qualification-assessment.html', menu=menu)

@app.route('/safety-culture')
def safety_culture():
    """안전문화 현황 페이지"""
    return render_template('safety-culture.html', menu=menu)

@app.after_request
def add_header(response):
    """응답 헤더 추가 - 캐시 무효화"""
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


EXCLUDED_AUDIT_PATHS = {
    '/api/admin/audit-logs',
    '/api/admin/usage-dashboard',
}


@app.after_request
def audit_request_activity(response):
    """요청/응답 단위 감사 로그 기록"""
    try:
        path = request.path or ''
        if not path or path.startswith('/static') or path.startswith('/favicon'):
            return response
        if request.method == 'OPTIONS':
            return response
        if path in EXCLUDED_AUDIT_PATHS:
            return response

        endpoint = request.endpoint or ''
        scope = 'API'
        if path.startswith('/api/admin') or endpoint.startswith('admin_'):
            scope = 'SYSTEM'
        elif 'permission' in endpoint or path.startswith('/api/permission'):
            scope = 'PERMISSION'
        elif not path.startswith('/api'):
            scope = 'MENU'

        slug_parts = path.strip('/').split('/')
        primary = slug_parts[0] if slug_parts else ''
        menu_code = resolve_menu_code(primary) if primary else None
        if scope == 'SYSTEM' and menu_code and menu_code.upper().startswith('API'):
            menu_code = 'SYSTEM'
        elif scope == 'API' and not menu_code:
            menu_code = resolve_menu_code(primary) or 'API'
        elif scope == 'MENU' and not menu_code and primary:
            menu_code = resolve_menu_code(primary) or primary.replace('-', '_').upper()
        elif scope == 'MENU' and not menu_code:
            menu_code = 'HOME'

        lowered_endpoint = endpoint.lower()
        lowered_path = path.lower()
        action_hint = None

        def _match_action(keywords, label):
            nonlocal action_hint
            if action_hint is not None:
                return
            for word in keywords:
                if word in lowered_endpoint or word in lowered_path:
                    action_hint = label
                    return

        _match_action(['register', 'create', 'new'], 'REGISTER')
        _match_action(['detail', 'view'], 'DETAIL')
        _match_action(['update', 'modify', 'edit'], 'UPDATE')
        _match_action(['delete', 'remove'], 'DELETE')
        _match_action(['list', 'index'], 'LIST')

        if action_hint is None:
            method = request.method.upper()
            if method == 'POST':
                action_hint = 'CREATE'
            elif method in ('PUT', 'PATCH'):
                action_hint = 'UPDATE'
            elif method == 'DELETE':
                action_hint = 'DELETE'
            elif method == 'GET':
                action_hint = 'VIEW'
            else:
                action_hint = method

        status_code = response.status_code
        success = status_code < 400
        details = {
            'method': request.method,
            'endpoint': endpoint,
            'status_code': status_code,
        }
        if request.args:
            details['query'] = {
                key: request.args.getlist(key) if len(request.args.getlist(key)) > 1 else request.args.get(key)
                for key in request.args
            }
        if request.blueprint:
            details['blueprint'] = request.blueprint
        json_payload = None
        try:
            json_payload = request.get_json(silent=True)
        except Exception:
            json_payload = None
        if isinstance(json_payload, dict):
            details['json_keys'] = list(json_payload.keys())[:10]
        elif json_payload is not None:
            details['json_type'] = type(json_payload).__name__

        record_audit_log(
            action_scope=scope,
            action_type=action_hint,
            action=action_hint,
            menu_code=menu_code,
            request_path=path,
            permission_result='SUCCESS' if success else 'FAILED',
            success=success,
            details=details,
        )
    except Exception as exc:
        app.logger.debug('Audit logging skipped: %s', exc)
    return response


@app.route('/api/partner-change-request', methods=['POST'])
def create_partner_change_request():
    """기준정보 변경요청 등록 API"""
    try:
        data = request.get_json()
        
        # 필수 필드 검증
        required_fields = ['requester_name', 'requester_department', 'company_name', 
                          'business_number', 'change_type', 'current_value', 'new_value', 'change_reason']
        
        for field in required_fields:
            if not data.get(field):
                return jsonify({"success": False, "message": f"{field} 필드가 필요합니다."}), 400
        
        # DB 연결 및 테이블 생성
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # PostgreSQL/SQLite 호환 처리
        # 운영환경에서는 테이블이 이미 생성되어 있어야 함
        
        # request_number 생성 (CR-YYYYMM-SEQ 형식)
        current_month = get_korean_time().strftime('%Y%m')
        
        # 이번 달 최대 시퀀스 번호 조회
        cursor.execute("""
            SELECT MAX(CAST(SUBSTR(request_number, -2) AS INTEGER)) as max_seq
            FROM partner_change_requests
            WHERE request_number LIKE %s
        """, (f'CR-{current_month}-%',))
        
        result = cursor.fetchone()
        max_seq = result[0] if result[0] else 0
        new_seq = max_seq + 1
        request_number = f'CR-{current_month}-{new_seq:02d}'
        
        # status 값 결정 (requested로 설정)
        status = data.get('status', 'requested')
        
        # 변경요청 데이터 삽입 (request_number, custom_data, status 포함)
        cursor.execute("""
            INSERT INTO partner_change_requests
            (request_number, requester_name, requester_department, company_name, business_number,
             change_type, current_value, new_value, change_reason, custom_data, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            request_number,  # 자동 생성된 request_number
            data['requester_name'],
            data['requester_department'],
            data['company_name'],
            data['business_number'],
            data['change_type'],
            data['current_value'],
            data['new_value'],
            data['change_reason'],
            pyjson.dumps(data.get('custom_data', {})),  # custom_data 추가
            status  # status 추가
        ))
        
        request_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "request_id": request_id,
            "message": "변경요청이 성공적으로 등록되었습니다.",
            "reload": True
        })
        
    except Exception as e:
        logging.error(f"변경요청 등록 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/change-requests/delete', methods=['POST'])
def delete_change_requests():
    """선택한 변경요청들을 소프트 삭제 (공통 서비스 활용)"""
    try:
        from board_services import ItemService
        
        data = request.json
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({"success": False, "message": "삭제할 항목이 없습니다."}), 400
        
        # partner_change_requests 테이블에 is_deleted 컬럼 추가 (없으면)
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # PostgreSQL: information_schema를 통해 컬럼 정보 조회
        cursor.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'partner_change_requests'
        """)
        columns = [col[0] for col in cursor.fetchall()]
        
        if 'is_deleted' not in columns:
            cursor.execute("""
                ALTER TABLE partner_change_requests 
                ADD COLUMN is_deleted INTEGER DEFAULT 0
            """)
            conn.commit()
        
        # 소프트 삭제 실행
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            UPDATE partner_change_requests 
            SET is_deleted = 1 
            WHERE id IN ({placeholders})
        """, ids)
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "deleted_count": deleted_count,
            "message": f"{deleted_count}건이 삭제되었습니다."
        })
    except Exception as e:
        logging.error(f"변경요청 삭제 중 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/partner-change-requests', methods=['GET'])
def get_partner_change_requests():
    """기준정보 변경요청 목록 조회 API"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, requester_name, requester_department, company_name, business_number,
                   change_type, current_value, new_value, change_reason, status, created_at
            FROM partner_change_requests
            ORDER BY created_at DESC
        """)
        
        requests = []
        for row in cursor.fetchall():
            requests.append({
                'id': row[0],
                'requester_name': row[1],
                'requester_department': row[2],
                'company_name': row[3],
                'business_number': row[4],
                'change_type': row[5],
                'current_value': row[6],
                'new_value': row[7],
                'change_reason': row[8],
                'status': row[9],
                'created_at': row[10]
            })
        
        conn.close()
        return jsonify({'success': True, 'requests': requests})
        
    except Exception as e:
        logging.error(f"변경요청 목록 조회 중 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# 변경요청 컬럼 관리 API

@app.route('/api/person-master', methods=['GET'])
def api_get_person_master():
    """담당자 마스터 목록 조회"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, name, department, position, company_name, phone, email, is_active
            FROM person_master
            WHERE is_active = 1
            ORDER BY name
        """)
        
        persons = []
        for row in cursor.fetchall():
            persons.append({
                'id': row[0],
                'name': row[1],
                'department': row[2],
                'position': row[3],
                'company_name': row[4],
                'phone': row[5],
                'email': row[6],
                'is_active': row[7]
            })
        
        conn.close()
        return jsonify({'success': True, 'persons': persons})
        
    except Exception as e:
        logging.error(f"담당자 마스터 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/person-master/<int:person_id>', methods=['GET'])
def api_get_person_detail(person_id):
    """담당자 상세 정보 조회"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, name, department, position, company_name, phone, email
            FROM person_master
            WHERE id = %s AND is_active = 1
        """, (person_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            person = {
                'id': row[0],
                'name': row[1],
                'department': row[2],
                'position': row[3],
                'company_name': row[4],
                'phone': row[5],
                'email': row[6]
            }
            return jsonify({'success': True, 'person': person})
        else:
            return jsonify({'success': False, 'message': '담당자를 찾을 수 없습니다.'}), 404
            
    except Exception as e:
        logging.error(f"담당자 상세 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/person-master', methods=['POST'])
@require_admin_auth
def api_create_person():
    """담당자 추가"""
    try:
        data = request.get_json()
        
        if not data.get('name'):
            return jsonify({'success': False, 'message': '이름은 필수 입력 항목입니다.'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute_with_returning_id("""
            INSERT INTO person_master (name, department, position, company_name, phone, email)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            data.get('name'),
            data.get('department', ''),
            data.get('position', ''),
            data.get('company_name', ''),
            data.get('phone', ''),
            data.get('email', '')
        ))
        
        person_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'person_id': person_id, 'message': '담당자가 추가되었습니다.'})
        
    except Exception as e:
        logging.error(f"담당자 추가 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/person-master/<int:person_id>', methods=['PUT'])
@require_admin_auth
def api_update_person(person_id):
    """담당자 수정"""
    try:
        data = request.get_json()
        
        if not data.get('name'):
            return jsonify({'success': False, 'message': '이름은 필수 입력 항목입니다.'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE person_master 
            SET name = %s, department = %s, position = %s, company_name = %s, phone = %s, email = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND is_active = 1
        """, (
            data.get('name'),
            data.get('department', ''),
            data.get('position', ''),
            data.get('company_name', ''),
            data.get('phone', ''),
            data.get('email', ''),
            person_id
        ))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': '담당자를 찾을 수 없습니다.'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '담당자가 수정되었습니다.'})
        
    except Exception as e:
        logging.error(f"담당자 수정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/person-master/<int:person_id>', methods=['DELETE'])
@require_admin_auth
def api_delete_person(person_id):
    """담당자 삭제 (소프트 삭제)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE person_master 
            SET is_active = 0, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND is_active = 1
        """, (person_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': '담당자를 찾을 수 없습니다.'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '담당자가 삭제되었습니다.'})
        
    except Exception as e:
        logging.error(f"담당자 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# 표준화된 API 엔드포인트 (/api/<board>/...)
# ============================================================================

@app.route("/api/<board>/columns", methods=["GET", "POST"])
def board_columns_api(board):
    """보드별 컬럼 관리 API"""
    board_type = board.replace('-', '_')
    
    try:
        repository = ColumnConfigRepository(DB_PATH, board_type)

        if request.method == "GET":
            columns = repository.list()
            return jsonify(columns)

        elif request.method == "POST":
            data = request.json or {}
            column_id = repository.add(data)
            return jsonify({"success": True, "id": column_id})

    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board columns API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/columns/<int:column_id>", methods=["PUT", "DELETE"])
def board_column_api(board, column_id):
    """보드별 개별 컬럼 관리 API"""
    board_type = board.replace('-', '_')
    
    try:
        repository = ColumnConfigRepository(DB_PATH, board_type)

        if request.method == "PUT":
            data = request.json or {}
            repository.update(column_id, data)
            return jsonify({"success": True})

        elif request.method == "DELETE":
            repository.delete(column_id)
            return jsonify({"success": True})

    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board column API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/columns/order", methods=["PUT"])
def board_columns_order_api(board):
    """보드별 컬럼 순서 변경 API"""
    board_type = board.replace('-', '_')
    
    try:
        repository = ColumnConfigRepository(DB_PATH, board_type)
        items = request.json or []
        repository.reorder(items)
        return jsonify({"success": True})

    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board columns order API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/dropdown-codes", methods=["GET", "POST"])
def board_dropdown_codes_api(board):
    """보드별 드롭다운 코드 관리 API"""
    board_type = board.replace('-', '_')
    column_key = request.args.get('column_key')
    
    if not column_key:
        return jsonify({"error": "column_key is required"}), 400
    
    try:
        service = CodeService(board_type, DB_PATH)
        
        if request.method == "GET":
            codes = service.list(column_key)
            return jsonify(codes)
        
        elif request.method == "POST":
            codes = request.json
            service.save(column_key, codes)
            return jsonify({"success": True})
            
    except Exception as e:
        logging.error(f"Board dropdown codes API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/dropdown-codes/<int:code_id>", methods=["DELETE"])
def board_dropdown_code_api(board, code_id):
    """보드별 개별 드롭다운 코드 삭제 API"""
    board_type = board.replace('-', '_')
    
    try:
        service = CodeService(board_type, DB_PATH)
        service.delete(code_id)
        return jsonify({"success": True})
        
    except Exception as e:
        logging.error(f"Board dropdown code API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/dropdown-codes/history", methods=["GET"])
def board_dropdown_history(board):
    """보드별 드롭다운 코드 변경 이력 (컬럼 기준)"""
    board_type = board.replace('-', '_')
    column_key = request.args.get('column_key')
    if not column_key:
        return jsonify({"error": "column_key is required"}), 400
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # board_type 컬럼 존재 여부 확인
        # PostgreSQL: information_schema를 통해 컬럼 정보 조회
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'dropdown_code_audit'
        """)
        cols = [r[0] for r in cur.fetchall()]
        if 'board_type' in cols:
            history = conn.execute(
                """
                SELECT * FROM dropdown_code_audit
                WHERE board_type = %s AND column_key = %s
                ORDER BY changed_at DESC, id DESC
                """,
                (board_type, column_key),
            ).fetchall()
        else:
            history = conn.execute(
                """
                SELECT * FROM dropdown_code_audit
                WHERE column_key = %s
                ORDER BY changed_at DESC, id DESC
                """,
                (column_key,),
            ).fetchall()
        conn.close()
        return jsonify({"history": [dict(h) for h in history]})
    except Exception as e:
        logging.error(f"Board dropdown history error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/dropdown-codes/audit-summary", methods=["GET"])
def board_dropdown_audit_summary(board):
    """보드별 드롭다운 감사 요약"""
    board_type = board.replace('-', '_')
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # PostgreSQL: information_schema를 통해 컬럼 정보 조회
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'dropdown_code_audit'
        """)
        cols = [r[0] for r in cur.fetchall()]
        if 'board_type' in cols:
            recent = conn.execute(
                """
                SELECT DATE(changed_at) as date,
                       COUNT(*) as total_changes,
                       COUNT(DISTINCT column_key) as columns_changed
                FROM dropdown_code_audit
                WHERE board_type = %s AND changed_at >= datetime('now','-7 days')
                GROUP BY DATE(changed_at)
                ORDER BY date DESC
                """,
                (board_type,),
            ).fetchall()
            most = conn.execute(
                """
                SELECT column_key, COUNT(*) as change_count, MAX(changed_at) as last_changed
                FROM dropdown_code_audit
                WHERE board_type = %s
                GROUP BY column_key
                ORDER BY change_count DESC
                LIMIT 5
                """,
                (board_type,),
            ).fetchall()
        else:
            # board_type 미도입 환경: 전역 기준 요약
            recent = conn.execute(
                """
                SELECT DATE(changed_at) as date,
                       COUNT(*) as total_changes,
                       COUNT(DISTINCT column_key) as columns_changed
                FROM dropdown_code_audit
                WHERE changed_at >= datetime('now','-7 days')
                GROUP BY DATE(changed_at)
                ORDER BY date DESC
                """,
            ).fetchall()
            most = conn.execute(
                """
                SELECT column_key, COUNT(*) as change_count, MAX(changed_at) as last_changed
                FROM dropdown_code_audit
                GROUP BY column_key
                ORDER BY change_count DESC
                LIMIT 5
                """,
            ).fetchall()
        conn.close()
        return jsonify({
            "recent_changes": [dict(r) for r in recent],
            "most_changed_columns": [dict(m) for m in most],
        })
    except Exception as e:
        logging.error(f"Board audit summary error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/items", methods=["GET"])
def board_items_api(board):
    """보드별 아이템 목록 API"""
    board_type = board.replace('-', '_')
    
    try:
        service = ItemService(board_type, DB_PATH)
        
        filters = {
            'company_name': request.args.get('company_name'),
            'business_number': request.args.get('business_number'),
            'date_start': request.args.get('date_start'),
            'date_end': request.args.get('date_end'),
        }
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        
        result = service.list(filters, page, per_page)
        return jsonify(result)
        
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board items API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/items/<int:item_id>", methods=["GET"])
def board_item_detail_api(board, item_id):
    """보드별 아이템 상세 API"""
    board_type = board.replace('-', '_')
    
    try:
        service = ItemService(board_type, DB_PATH)
        item = service.detail(item_id)
        
        if item:
            return jsonify(item)
        else:
            return jsonify({"error": "Item not found"}), 404
            
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board item detail API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/items/register", methods=["POST"])
def board_item_register_api(board):
    """보드별 아이템 등록 API"""
    board_type = board.replace('-', '_')
    
    try:
        service = ItemService(board_type, DB_PATH)
        data = request.json
        result = service.register(data)
        return jsonify({"success": True, **result})
        
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board item register API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/items/update/<int:item_id>", methods=["POST"])
def board_item_update_api(board, item_id):
    """보드별 아이템 수정 API"""
    board_type = board.replace('-', '_')
    
    try:
        service = ItemService(board_type, DB_PATH)
        data = request.json
        service.update(item_id, data)
        return jsonify({"success": True})
        
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board item update API error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/<board>/items/delete", methods=["POST"])
def board_items_delete_api(board):
    """보드별 아이템 삭제 API"""
    board_type = board.replace('-', '_')
    
    try:
        service = ItemService(board_type, DB_PATH)
        item_ids = request.json.get('ids', [])
        hard_delete = request.json.get('hard_delete', False)
        service.delete(item_ids, hard_delete)
        return jsonify({"success": True})
        
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logging.error(f"Board items delete API error: {e}")
        return jsonify({"error": str(e)}), 500


# ============= CMS Catch-all Route (모든 라우트 후에 배치) =============
@app.route("/<path:url>")
def page_view(url):
    """일반 페이지 체크 (catch-all 라우트) - 모든 다른 라우트 후에 실행"""
    # 실제 라우트로 리다이렉트
    route_map = {
        'accident': 'accident_route',
        # 구경로 호환: partner-accident는 accident로 리다이렉트
        'partner-accident': 'accident_route',
        'safety-instruction': 'safety_instruction.safety_instruction_route',
        'follow-sop': 'follow_sop.follow_sop_route',
        'full-process': 'full_process.full_process_route',
        'safe-workplace': 'safe_workplace.safe_workplace_route',
        'partner-standards': 'partner_standards_route',
        # 구 라우트 호환: /change-request -> /partner-change-request
        'change-request': 'partner_change_request_route',
    }
    
    if url in route_map:
        endpoint = route_map[url]
        if endpoint != 'page_view':
            # 엔드포인트가 미등록이면 BuildError를 유발하지 않도록 가드
            if endpoint in app.view_functions:
                return redirect(url_for(endpoint))
            else:
                try:
                    import logging as _logging
                    _logging.error(f"[page_view] 미등록 엔드포인트: {endpoint} (url={url})")
                except Exception:
                    pass
                # 미등록이면 CMS 페이지 탐색으로 폴백 (아래 로직 수행)
    
    conn = get_db_connection()
    # Ensure pages table exists even if init hook didn't run yet
    try:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS pages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT UNIQUE,
                title TEXT,
                content TEXT
            )
        ''')
    except Exception as _e:
        try:
            import logging as _logging
            _logging.debug(f"pages ensure failed: {_e}")
        except Exception:
            pass
    page = conn.execute("SELECT * FROM pages WHERE url = %s", (url,)).fetchone()
    conn.close()
    
    if not page:
        return "Page not found", 404
    
    return render_template("page.html",
                         page={'url': page[1], 'title': page[2], 'content': page[3]},
                         menu=MENU_CONFIG)

# =====================
# 콘텐츠 업로드 라우트 (if __name__ 위로 이동)
# =====================
@app.route('/upload-inline-image', methods=['POST'])
def upload_inline_image():
    try:
        file = request.files.get('upload') or request.files.get('file')
        if not file or not file.filename:
            return jsonify({ 'error': { 'message': 'No file uploaded' } }), 400

        # 확장자/크기 검증
        allowed = {'png','jpg','jpeg','gif','webp','bmp'}
        ext = os.path.splitext(file.filename)[1].lower().lstrip('.')
        if ext not in allowed:
            return jsonify({ 'error': { 'message': 'Unsupported file type' } }), 400
        file.seek(0, os.SEEK_END)
        size = file.tell()
        file.seek(0)
        if size > 10 * 1024 * 1024:  # 10MB
            return jsonify({ 'error': { 'message': 'File too large' } }), 400

        # 저장 경로
        upload_dir = os.path.join('uploads', 'content')
        os.makedirs(upload_dir, exist_ok=True)

        # 안전한 파일명
        import time
        basename = sanitize_filename(file.filename, fallback_prefix='inline')
        fname = f"{int(time.time())}_{basename}"
        path = os.path.join(upload_dir, fname)
        file.save(path)

        url = f"/uploads/content/{fname}"
        return jsonify({ 'url': url })
    except Exception as e:
        logging.error(f"Inline image upload failed: {e}")
        return jsonify({ 'error': { 'message': 'Upload failed' } }), 500

@app.route('/uploads/content/<path:filename>')
def serve_uploaded_content(filename):
    directory = os.path.join('uploads', 'content')
    return send_from_directory(directory, filename)

# =====================
# 메뉴/권한 관련 헬퍼
# =====================

@app.route('/api/menus')
@require_admin_auth
def api_menus():
    """상단 메뉴 구성을 반환"""
    flattened = []
    for section in MENU_CONFIG:
        group = section.get('title')
        for item in section.get('submenu', []):
            slug = item.get('url')
            code = resolve_menu_code(slug)
            flattened.append({
                'group': group,
                'code': code,
                'slug': slug,
                'name': item.get('title')
            })
    return jsonify(flattened)


@app.route('/api/admin/audit-logs')
@require_admin_auth
def api_admin_audit_logs():
    """감사 로그 리스트"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = max(1, min(per_page, 200))

    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    action_type_param = request.args.get('action_type', '')
    scope_param = request.args.get('scope', '')
    menu_code_param = request.args.get('menu_code', '')
    result_param = request.args.get('result', '')
    success_param = request.args.get('success', '')
    user_param = request.args.get('user', '')

    normalized_action = normalize_action(action_type_param) if action_type_param else ''
    normalized_scope = normalize_scope(scope_param) if scope_param else ''
    normalized_result = normalize_result(result_param) if result_param else ''

    success_filter = None
    if success_param:
        lowered = success_param.strip().lower()
        if lowered in ('true', '1', 'yes', 'y'):
            success_filter = True
        elif lowered in ('false', '0', 'no', 'n'):
            success_filter = False

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            SELECT
                al.created_at,
                al.emp_id,
                al.login_id,
                COALESCE(u_login.user_name, u_emp.user_name) AS user_name,
                COALESCE(u_login.dept_name, u_emp.dept_name) AS dept_name,
                al.action_scope,
                al.action_type,
                al.action,
                al.menu_code,
                COALESCE(mn.menu_name, al.menu_code) AS menu_name,
                al.request_path,
                al.object_type,
                al.object_id,
                al.object_name,
                al.success,
                al.permission_result,
                al.details,
                al.ip_address,
                al.error_message
            FROM access_audit_log al
            LEFT JOIN system_users u_login ON al.login_id = u_login.login_id
            LEFT JOIN system_users u_emp ON al.emp_id = u_emp.emp_id
            LEFT JOIN menu_names mn ON mn.menu_code = al.menu_code
            WHERE 1=1
        """
        params = []

        if start_date:
            query += " AND al.created_at >= %s"
            params.append(start_date)
        if end_date:
            query += " AND al.created_at <= %s"
            params.append(f"{end_date} 23:59:59")
        if normalized_scope:
            query += " AND al.action_scope = %s"
            params.append(normalized_scope)
        if normalized_action:
            query += " AND al.action_type = %s"
            params.append(normalized_action)
        if menu_code_param:
            query += " AND al.menu_code = %s"
            params.append(menu_code_param)
        if normalized_result:
            query += " AND al.permission_result = %s"
            params.append(normalized_result)
        if success_filter is not None:
            query += " AND al.success = %s"
            params.append(success_filter)
        if user_param:
            query += " AND (al.login_id ILIKE %s OR al.emp_id ILIKE %s OR u_login.user_name ILIKE %s OR u_emp.user_name ILIKE %s)"
            like_value = f"%{user_param}%"
            params.extend([like_value, like_value, like_value, like_value])

        query += " ORDER BY al.created_at DESC LIMIT %s OFFSET %s"
        params.extend([per_page, (page - 1) * per_page])
        cursor.execute(query, params)
        logs = cursor.fetchall()

        def _cursor_description(cur):
            desc = getattr(cur, 'description', None)
            if desc is None and hasattr(cur, '_cursor'):
                desc = getattr(cur._cursor, 'description', None)
            return desc or []

        select_columns = [desc[0] for desc in _cursor_description(cursor) if desc]

        count_query = "SELECT COUNT(*) FROM access_audit_log al WHERE 1=1"
        count_params = []
        if start_date:
            count_query += " AND al.created_at >= %s"
            count_params.append(start_date)
        if end_date:
            count_query += " AND al.created_at <= %s"
            count_params.append(f"{end_date} 23:59:59")
        if normalized_scope:
            count_query += " AND al.action_scope = %s"
            count_params.append(normalized_scope)
        if normalized_action:
            count_query += " AND al.action_type = %s"
            count_params.append(normalized_action)
        if menu_code_param:
            count_query += " AND al.menu_code = %s"
            count_params.append(menu_code_param)
        if normalized_result:
            count_query += " AND al.permission_result = %s"
            count_params.append(normalized_result)
        if success_filter is not None:
            count_query += " AND al.success = %s"
            count_params.append(success_filter)
        if user_param:
            count_query += " AND (al.login_id ILIKE %s OR al.emp_id ILIKE %s OR EXISTS (SELECT 1 FROM system_users su WHERE su.login_id = al.login_id AND su.user_name ILIKE %s) OR EXISTS (SELECT 1 FROM system_users su WHERE su.emp_id = al.emp_id AND su.user_name ILIKE %s))"
            like_value = f"%{user_param}%"
            count_params.extend([like_value, like_value, like_value, like_value])
        cursor.execute(count_query, count_params)
        total_count = cursor.fetchone()[0] if cursor.rowcount != -1 else 0

        def _safe_detail(value):
            if value is None:
                return None
            if isinstance(value, (dict, list)):
                return value
            try:
                import json
                return json.loads(value)
            except Exception:
                return value

        def _row_to_dict(row_tuple):
            mapping = {}
            if select_columns:
                for idx, col_name in enumerate(select_columns):
                    try:
                        mapping[col_name] = row_tuple[idx]
                    except Exception:
                        continue
            elif hasattr(row_tuple, 'keys') and callable(row_tuple.keys):
                try:
                    mapping = {k: row_tuple[k] for k in row_tuple.keys()}
                except Exception:
                    mapping = {}
            return mapping

        log_payload = []
        for log in logs:
            row_map = _row_to_dict(log)

            def _field(name, index):
                if row_map and name in row_map:
                    return row_map[name]
                try:
                    return log[index]
                except Exception:
                    return None

            created_val = _field('created_at', 0)
            if isinstance(created_val, datetime):
                created_at_iso = created_val.isoformat()
            elif created_val is not None:
                created_at_iso = str(created_val)
            else:
                created_at_iso = None

            details_val = _field('details', 16)

            log_payload.append({
                'created_at': created_at_iso,
                'login_id': _field('login_id', 2),
                'user_name': _field('user_name', 3),
                'dept_name': _field('dept_name', 4),
                'action_scope': _field('action_scope', 5),
                'action_type': _field('action_type', 6),
                'action': _field('action', 7),
                'menu_code': _field('menu_code', 8),
                'menu_name': _field('menu_name', 9),
                'request_path': _field('request_path', 10),
                'object_type': _field('object_type', 11),
                'object_id': _field('object_id', 12),
                'object_name': _field('object_name', 13),
                'success': _field('success', 14),
                'permission_result': _field('permission_result', 15),
                'details': _safe_detail(details_val),
                'ip_address': _field('ip_address', 17),
                'error_message': _field('error_message', 18),
            })

        return jsonify({
            'logs': log_payload,
            'total': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_count + per_page - 1) // per_page if per_page else 0,
        })
    except Exception as exc:
        logging.exception('Failed to fetch audit logs: %s', exc)
        return jsonify({'success': False, 'message': str(exc)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.route('/api/admin/usage-dashboard')
@require_admin_auth
def api_admin_usage_dashboard():
    """사용 현황 대시보드 데이터"""

    def _extract(row, key, index):
        if hasattr(row, 'keys'):
            try:
                if hasattr(row, 'get'):
                    value = row.get(key)
                    if value is not None:
                        return value
            except Exception:
                pass
            try:
                if key in row.keys():
                    return row[key]
            except Exception:
                pass
        try:
            return row[index]
        except Exception:
            return None

    def _fetch_trend(option_name):
        query = db_config.config.get('USAGE_DASHBOARD', option_name, fallback='').strip()
        if not query:
            return []

        records = []
        df = None
        try:
            df = execute_SQL(query)
        except Exception as exc:  # noqa: BLE001
            logging.debug("Usage dashboard iqadb 조회 실패 (%s): %s", option_name, exc)

        if df is not None:
            try:
                records = [{k.lower(): v for k, v in row.items()} for row in df.to_dict(orient='records')]
            except Exception as exc:  # noqa: BLE001
                logging.debug("Usage dashboard iqadb 결과 변환 실패 (%s): %s", option_name, exc)
                records = []

        if not records:
            conn = None
            cursor = None
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                try:
                    cursor.execute("SET search_path TO iqadb, public")
                except Exception as exc:  # noqa: BLE001
                    logging.debug("Usage dashboard search_path 설정 실패 (계속 진행): %s", exc)

                cursor.execute(query)
                fetched = cursor.fetchall() or []
                column_names = [desc[0].lower() for desc in cursor.description] if cursor.description else []

                for row in fetched:
                    if hasattr(row, 'keys'):
                        lowered = {str(key).lower(): row[key] for key in row.keys()}
                    else:
                        lowered = {column_names[idx]: row[idx] for idx in range(min(len(column_names), len(row)))}
                    records.append(lowered)
            except Exception as exc:  # noqa: BLE001
                logging.error("Usage dashboard query 실패 (%s): %s", option_name, exc)
                return []
            finally:
                if cursor:
                    try:
                        cursor.close()
                    except Exception:
                        pass
                if conn:
                    conn.close()

        if not records:
            return []

        trend = []
        for record in records:
            raw_date = record.get('date') or record.get('usage_date')
            raw_count = record.get('count') or record.get('total') or record.get('value')

            if isinstance(raw_date, datetime):
                date_str = raw_date.strftime('%Y-%m-%d')
            else:
                date_str = str(raw_date) if raw_date is not None else ''

            try:
                count_val = int(raw_count) if raw_count is not None else 0
            except (TypeError, ValueError):
                try:
                    count_val = int(float(raw_count))
                except Exception:
                    count_val = 0

            if date_str:
                trend.append({'date': date_str, 'count': count_val})

        return trend

    portal_trend = _fetch_trend('portal_trend_query')
    chatbot_trend = _fetch_trend('chatbot_trend_query')

    response = {
        'portal_trend': portal_trend,
        'chatbot_trend': chatbot_trend,
        'portal_total': sum(item.get('count', 0) or 0 for item in portal_trend),
        'chatbot_total': sum(item.get('count', 0) or 0 for item in chatbot_trend),
        'updated_at': datetime.utcnow().isoformat() + 'Z'
    }

    return jsonify(response)

# =====================
# SSO 샘플 라우트 추가
# =====================

def _get_config_value(cfg: configparser.ConfigParser, section: str, key: str, *fallback_keys, default: str = '') -> str:
    """Try multiple keys (for legacy compatibility) and return first available value."""
    # Primary
    if cfg.has_option(section, key):
        return cfg.get(section, key)
    # Fallbacks (flat, legacy-style)
    for fk in fallback_keys:
        if cfg.has_option(section, fk):
            return cfg.get(section, fk)
    return default

def _compute_redirect_uri(cfg: configparser.ConfigParser) -> str:
    # config.ini에서 설정된 URL 우선 사용, 없으면 자동 감지
    try:
        # 1. config.ini의 명시적 설정값 확인
        config_url = _get_config_value(cfg, 'SSO', 'sp_redirect_url', default='')
        if config_url and config_url.strip():
            # config에 값이 있으면 그대로 사용
            return config_url.strip()

        # 2. 설정이 없으면 현재 요청 기반으로 자동 생성
        # 개발환경 (localhost)
        if 'localhost' in request.host or '127.0.0.1' in request.host:
            port = request.host.split(':')[1] if ':' in request.host else '44369'
            return f"https://localhost:{port}/acs"

        # 3. 운영환경 - X-Forwarded 헤더 확인
        forwarded_host = request.headers.get('X-Forwarded-Host')
        forwarded_proto = request.headers.get('X-Forwarded-Proto', 'https')

        if forwarded_host:
            return f"{forwarded_proto}://{forwarded_host}/acs"

        # 4. 기본값: 현재 호스트 사용
        scheme = request.headers.get('X-Forwarded-Proto') or request.scheme
        if scheme == 'http' and 'localhost' not in request.host:
            scheme = 'https'  # 운영환경은 HTTPS 가정

        return f"{scheme}://{request.host}/acs"

    except Exception as e:
        print(f"[SSO] redirect_uri 계산 실패: {e}")
        # 최종 폴백: config의 기본값 또는 localhost
        return config_url if config_url else 'https://localhost:44369/acs'

def _urlencode(params: dict) -> str:
    from urllib.parse import urlencode
    return urlencode(params)


def _load_sso_config():
    cfg = configparser.ConfigParser()
    cfg.read('config.ini', encoding='utf-8')
    return cfg


def _sso_dev_flags(cfg: configparser.ConfigParser):
    dev_mode = cfg.getboolean('SSO', 'dev_mode', fallback=False)
    dev_flow = cfg.getboolean('SSO', 'dev_simulate_flow', fallback=False)
    return dev_mode and dev_flow, dev_mode, dev_flow


@app.route('/sso/dev-login', methods=['GET', 'POST'])
def sso_dev_login():
    """개발용 SSO 시뮬레이터."""
    cfg = _load_sso_config()
    dev_enabled, dev_mode, dev_flow = _sso_dev_flags(cfg)

    if not dev_enabled:
        abort(404)

    defaults = {
        'login_id': cfg.get('SSO', 'dev_user_id', fallback='dev_user'),
        'user_name': cfg.get('SSO', 'dev_user_name', fallback='개발자'),
        'dept_id': cfg.get('SSO', 'dev_department', fallback='DEV001'),
        'dept_name': cfg.get('SSO', 'dev_department_name', fallback=cfg.get('SSO', 'dev_department', fallback='개발팀')),
        'email': cfg.get('SSO', 'dev_user_email', fallback='dev@example.com'),
        'emp_id': cfg.get('SSO', 'dev_emp_id', fallback='DEVUSER'),
        'company_id': cfg.get('SSO', 'dev_company_id', fallback='COMP001'),
        'grade': cfg.get('SSO', 'dev_grade', fallback='과장'),
    }

    super_admin_users = []
    try:
        from permission_helpers import SUPER_ADMIN_USERS
        super_admin_users = SUPER_ADMIN_USERS
    except Exception:
        super_admin_users = []

    if request.method == 'POST':
        form = request.form
        login_id = form.get('login_id', '').strip() or defaults['login_id']
        user_name = form.get('user_name', '').strip() or defaults['user_name']
        emp_id = form.get('emp_id', '').strip() or login_id
        dept_id = form.get('dept_id', '').strip() or defaults['dept_id']
        dept_name = form.get('dept_name', '').strip() or defaults['dept_name']
        email = form.get('email', '').strip() or defaults['email']
        company_id = form.get('company_id', '').strip() or defaults['company_id']
        grade = form.get('grade', '').strip() or defaults['grade']

        session['user_id'] = login_id
        session['loginid'] = login_id
        session['emp_id'] = emp_id
        session['userid'] = emp_id
        session['user_name'] = user_name
        session['deptid'] = dept_id
        session['deptname'] = dept_name
        session['mail'] = email
        session['compid'] = company_id
        session['grade'] = grade
        session['authenticated'] = True
        session['sso_simulated'] = True

        if login_id in super_admin_users:
            session['role'] = 'super_admin'
        else:
            session['role'] = session.get('role', 'user')

        next_url = session.pop('next_url', '/') if 'next_url' in session else '/'
        print(f"[SSO DEV] Simulated login for {login_id}, redirect -> {next_url}")
        return redirect(next_url)

    return render_template(
        'sso-dev-login.html',
        defaults=defaults,
        super_admins=super_admin_users,
        dev_mode=dev_mode,
        dev_flow=dev_flow
    )

@app.route('/SSO')
def sso():
    """SSO 인증 시작"""
    cfg = configparser.ConfigParser()
    cfg.read('config.ini', encoding='utf-8')

    # State/Nonce for CSRF + replay protection
    state_val = uuid.uuid4().hex
    nonce_val = uuid.uuid4().hex
    session['sso_state'] = state_val
    session['sso_nonce'] = nonce_val

    # IDP authorize endpoint (config name varies in samples)
    idp_url = _get_config_value(
        cfg, 'SSO', 'idp_authorize_url',
        'idp_entity_id', 'Idp.EntityID', default=''
    )
    client_id = _get_config_value(cfg, 'SSO', 'idp_client_id', 'Idp.ClientID', default='')
    redirect_uri = _compute_redirect_uri(cfg)

    if not idp_url or not client_id:
        print(f"[SSO] Missing idp_url/client_id. idp_url='{idp_url}', client_id='{client_id}'")
    
    params = {
        'client_id': client_id,
        'redirect_uri': redirect_uri,
        'response_mode': 'form_post',
        'response_type': 'id_token',
        'scope': 'openid profile',
        'nonce': nonce_val,
        'state': state_val,
    }
    auth_url = idp_url + ('?' if '?' not in idp_url else '&') + _urlencode(params)
    try:
        # 개발 편의: /SSO?debug=1 로 호출 시 리다이렉트 대신 내용을 그대로 보여줌
        if request.args.get('debug') == '1':
            headers = {
                'X-Forwarded-Proto': request.headers.get('X-Forwarded-Proto'),
                'X-Forwarded-Host': request.headers.get('X-Forwarded-Host'),
                'Host': request.headers.get('Host'),
            }
            html = f"""
            <h2>SSO Debug</h2>
            <pre>
            idp_url: {idp_url}
            client_id: {client_id}
            redirect_uri: {redirect_uri}
            state: {state_val}
            nonce: {nonce_val}
            auth_url: {auth_url}
            headers: {headers}
            scheme: {request.scheme}
            url_root: {request.url_root}
            </pre>
            <a href="{auth_url}">Go to Auth URL</a>
            """
            return html
        print(f"[SSO] Auth URL: {auth_url}")
        return redirect(auth_url, code=302)
    except Exception as e:
        return f"SSO error building auth URL: {e}", 500

def _load_public_key_from_cert_bytes(cert_bytes: bytes):
    """Try PEM then DER to load a certificate and return its public key."""
    try:
        cert_obj = x509.load_pem_x509_certificate(cert_bytes, default_backend())
        return cert_obj.public_key()
    except Exception:
        try:
            cert_obj = x509.load_der_x509_certificate(cert_bytes, default_backend())
            return cert_obj.public_key()
        except Exception as e:
            raise e

def _read_cert_bytes(cfg: configparser.ConfigParser) -> bytes:
    # Accept multiple config shapes and case variants
    cert_dir = _get_config_value(cfg, 'SSO', 'cert_file_path', 'CertFile_Path', default='templates/Cert/')
    cert_name = _get_config_value(cfg, 'SSO', 'cert_file_name', 'CertFile_Name', default='Idp.cer')
    # Normalize path separators
    base = os.getcwd()
    # 윈도우 경로 구분자 섞임 방지
    cert_dir_norm = cert_dir.replace('\\', os.sep).replace('/', os.sep)
    candidates = [
        os.path.join(base, cert_dir_norm, cert_name),
        os.path.join(base, 'templates', 'Cert', cert_name),
        os.path.join(base, 'Templates', 'Cert', cert_name),
    ]
    print(f"[SSO][IDP CERT] CWD: {base}")
    print(f"[SSO][IDP CERT] Declared dir/name: '{cert_dir}' / '{cert_name}'")
    tried = []
    for p in candidates:
        exists = os.path.exists(p)
        print(f"[SSO][IDP CERT] Try: {p} exists={exists}")
        tried.append((p, exists))
        if not exists:
            continue
        try:
            with open(p, 'rb') as f:
                print(f"[SSO] Using cert: {p}")
                return f.read()
        except Exception as e:
            print(f"[SSO][IDP CERT] Read fail: {p} error={e}")
            continue
    raise FileNotFoundError("Certificate not found. Tried: " + ", ".join([f"{p} (exists={ex})" for p, ex in tried]))

@app.route('/acs', methods=['GET', 'POST'])
def acs():
    """SSO 콜백 처리"""
    isLoad = False
    isError = False
    Error_MSG = ''
    claim_val = ''

    if request.method == 'POST':
        try:
            cfg = configparser.ConfigParser()
            cfg.read('config.ini', encoding='utf-8')
            cert_bytes = _read_cert_bytes(cfg)
            public_key = _load_public_key_from_cert_bytes(cert_bytes)

            id_token_val = request.form['id_token']
            # Optional state check if response_mode=form_post includes state
            state_in = request.form.get('state')
            if state_in and session.get('sso_state') and state_in != session.get('sso_state'):
                raise Exception('Invalid SSO state')

            # Decode
            decoded = jwt.decode(
                id_token_val,
                key=public_key,
                algorithms=['RS256'],
                options={'verify_signature': True, 'verify_exp': True, 'verify_aud': False}
            )
            json_str = json.dumps(decoded)
            claim_val = json.loads(json_str)

            # Optional nonce check
            nonce_expected = session.get('sso_nonce')
            nonce_received = claim_val.get('nonce') if isinstance(claim_val, dict) else None
            if nonce_expected and nonce_received and nonce_expected != nonce_received:
                print(f"[SSO] Nonce mismatch: expected={nonce_expected}, received={nonce_received}")

            # 세션에 사용자 정보 저장
            if claim_val:
                # SSO에서 받은 사용자 정보 추출 (정확한 필드명만 사용)
                emp_id = claim_val.get('userid', '')  # 시스템 고유 ID
                login_id = claim_val.get('loginid', '')  # 로그인 ID

                sso_fields = {
                    'user_id': login_id,  # 기존 코드 호환성 유지
                    'emp_id': emp_id,     # 권한 시스템용 추가
                    'user_name': claim_val.get('username', ''),
                    'userid': emp_id,     # 원본 필드도 유지
                    'compid': claim_val.get('compid', ''),
                    'deptid': claim_val.get('deptid', ''),
                    'mail': claim_val.get('mail', ''),
                    'deptname': claim_val.get('deptname', ''),
                    'mobile': claim_val.get('mobile', '')
                }

                # 세션에 저장
                for key, value in sso_fields.items():
                    if value:  # 값이 있는 경우만 저장
                        session[key] = value

                # 인증 플래그 설정
                session['authenticated'] = True

                # 권한 시스템: DB에 사용자 정보 저장/업데이트
                if emp_id and login_id:
                    try:
                        conn = get_db_connection()
                        cursor = conn.cursor()

                        # system_users 테이블 업데이트
                        cursor.execute("""
                            INSERT INTO system_users
                            (emp_id, login_id, user_name, dept_id, dept_name,
                             company_id, email, last_login_at, is_active)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, TRUE)
                            ON CONFLICT (emp_id) DO UPDATE SET
                                login_id = EXCLUDED.login_id,
                                user_name = EXCLUDED.user_name,
                                dept_id = EXCLUDED.dept_id,
                                dept_name = EXCLUDED.dept_name,
                                company_id = EXCLUDED.company_id,
                                email = EXCLUDED.email,
                                last_login_at = CURRENT_TIMESTAMP,
                                updated_at = CURRENT_TIMESTAMP
                        """, (
                            emp_id,
                            login_id,
                            claim_val.get('username', ''),
                            claim_val.get('deptid', ''),
                            claim_val.get('deptname', ''),
                            claim_val.get('compid', ''),
                            claim_val.get('mail', '')
                        ))

                        # 기본 역할(user) 할당 (없으면)
                        cursor.execute("""
                            INSERT INTO user_role_mapping (emp_id, role_code)
                            VALUES (%s, 'user')
                            ON CONFLICT (emp_id, role_code) DO NOTHING
                        """, (emp_id,))

                        conn.commit()
                        cursor.close()
                        conn.close()

                        print(f"[SSO] User {emp_id}/{login_id} updated in DB")

                    except Exception as e:
                        print(f"[SSO] DB update error: {e}")
                        # DB 오류가 있어도 로그인은 진행

                # 디버깅 출력 (간결하게)
                print("[SSO] Session updated:")
                for key in ['user_id', 'emp_id', 'user_name', 'deptname', 'mail']:
                    if key in session:
                        print(f"  {key}: '{session.get(key)}'")
                print("="*50)

        except jwt.ExpiredSignatureError:
            isError = True
            Error_MSG = 'Authentication Token has expired.'
        except jwt.InvalidTokenError:
            isError = True
            Error_MSG = 'Authentication Token is not valid.'
        except Exception as e:
            isError = True
            Error_MSG = f'Error: {str(e)}'

        # 성공 시 원래 가려던 페이지로 리다이렉트
        if not isError and claim_val and session.get('user_name'):
            next_url = session.pop('next_url', '/')
            print(f"[SSO SUCCESS] Redirecting to {next_url}")
            return redirect(next_url)

    # 에러 또는 GET 요청 시 SSO 페이지 표시 (간단한 디버그 정보 포함)
    try:
        return render_template('sso_index.html', isLoad=isLoad, isError=isError, Error_MSG=Error_MSG, Claims=claim_val)
    except Exception:
        # 템플릿이 없더라도 최소한 텍스트 표시
        return f"SSO {'OK' if not isError else 'ERROR'}: {Error_MSG}", (200 if not isError else 400)

@app.route('/slo')
def slo():
    """SSO 로그아웃"""
    # 세션 클리어
    session.clear()

    # config.ini에서 로그아웃 URL 읽기
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    idp_url = config.get('SSO', 'idp_signout_url', fallback='example')
    return redirect(idp_url, code=302)


# SSO 자동 인증 미들웨어
@app.before_request
def auto_sso_redirect():
    """세션이 없으면 자동으로 SSO 시작"""

    # 제외 경로 (SSO, 정적 파일, API 등)
    excluded_paths = [
        '/SSO', '/sso', '/sso/diagnostics', '/sso/dev-login', '/acs', '/slo', '/static', '/uploads', '/api',
        '/admin/', '/admin/login', '/debug-session',
        # allow automation endpoints without SSO session
        '/admin/sync-now', '/admin/sync-now/', '/admin/cache-counts'
    ]

    cfg = None
    try:
        cfg = _load_sso_config()
    except Exception:
        cfg = None

    enabled_flag = True
    sso_enabled_flag = True
    dev_enabled = False
    dev_mode_flag = False
    dev_flow_flag = False
    if cfg:
        try:
            enabled_flag = cfg.getboolean('SSO', 'enabled', fallback=True)
            sso_enabled_flag = cfg.getboolean('SSO', 'sso_enabled', fallback=True)
            dev_enabled, dev_mode_flag, dev_flow_flag = _sso_dev_flags(cfg)
        except Exception:
            enabled_flag = True
            sso_enabled_flag = True
            dev_enabled = False

    if not (enabled_flag and sso_enabled_flag) and not dev_enabled:
        return None

    # 동기화 토큰이 있는 요청은 항상 통과 (자동화/스크립트 호출용)
    expected = ''
    if cfg:
        try:
            expected = cfg.get('ADMIN', 'SYNC_TOKEN', fallback='')
        except Exception:
            expected = ''
    if expected and request.headers.get('X-Sync-Token') == expected:
        return None

    # 제외 경로는 체크 안 함
    for path in excluded_paths:
        if request.path.startswith(path):
            return None

    # 세션에 user_name이 없으면 자동 SSO 또는 개발용 로그인
    if not session.get('user_name'):
        # 원래 가려던 URL 저장
        session['next_url'] = request.url
        if dev_enabled and not (enabled_flag and sso_enabled_flag):
            print(f"[AUTO SSO] Dev simulate enabled, redirecting to /sso/dev-login from {request.path}")
            return redirect('/sso/dev-login')
        print(f"[AUTO SSO] No session, redirecting to /SSO from {request.path}")
        return redirect('/SSO')

    return None

# 수동 SSO 테스트용 디버그 라우트
@app.route('/debug-session')
def debug_session():
    """세션 확인용 디버그 페이지"""
    return f"""
    <h1>Session Debug</h1>
    <pre>
    user_id: {session.get('user_id', 'NOT SET')}
    user_name: {session.get('user_name', 'NOT SET')}
    authenticated: {session.get('authenticated', False)}

    All session keys: {list(session.keys())}
    </pre>
    <hr>
    <a href="/SSO">Manual SSO Login</a> |
    <a href="/slo">Logout</a> |
    <a href="/">Home</a>
    """

# SSO 진단 페이지: 경로/설정/파일 존재 여부 표시
@app.route('/sso/diagnostics')
def sso_diagnostics():
    try:
        cfg = configparser.ConfigParser()
        cfg.read('config.ini', encoding='utf-8')

        # 기본 경로 정보
        cwd = os.getcwd()
        base_dir = os.path.dirname(os.path.abspath(__file__))

        # 서버 TLS 인증서(HTTPS 포트용)
        ssl_cert_cfg = cfg.get('SSO', 'ssl_cert_file', fallback='Templates/Cert/cert.pem')
        ssl_key_cfg = cfg.get('SSO', 'ssl_key_file', fallback='Templates/Cert/key.pem')
        ssl_cert_abs = ssl_cert_cfg if os.path.isabs(ssl_cert_cfg) else os.path.join(base_dir, ssl_cert_cfg)
        ssl_key_abs = ssl_key_cfg if os.path.isabs(ssl_key_cfg) else os.path.join(base_dir, ssl_key_cfg)

        # IdP 공개키 인증서(토큰 검증용)
        cert_dir = _get_config_value(cfg, 'SSO', 'cert_file_path', 'CertFile_Path', default='templates/Cert/')
        cert_name = _get_config_value(cfg, 'SSO', 'cert_file_name', 'CertFile_Name', default='Idp.cer')
        idp_candidates = [
            os.path.join(cwd, cert_dir.replace('\\', os.sep).replace('/', os.sep), cert_name),
            os.path.join(cwd, 'templates', 'Cert', cert_name),
            os.path.join(cwd, 'Templates', 'Cert', cert_name),
        ]

        # 후보 경로 HTML 생성
        li_items = []
        for p in idp_candidates:
            li_items.append(f"<li>{p} — exists={os.path.exists(p)}</li>")
        li_html = '\n'.join(li_items)

        # SSO 파라미터들
        idp_url = _get_config_value(cfg, 'SSO', 'idp_authorize_url', 'idp_entity_id', 'Idp.EntityID', default='')
        client_id = _get_config_value(cfg, 'SSO', 'idp_client_id', 'Idp.ClientID', default='')
        redirect_uri = _compute_redirect_uri(cfg)
        sp_redirect_url_cfg = _get_config_value(cfg, 'SSO', 'sp_redirect_url', default='')

        html = f"""
        <h2>SSO Diagnostics</h2>
        <h3>Paths</h3>
        <pre>CWD: {cwd}
BASE_DIR: {base_dir}</pre>
        <h4>Server TLS (HTTPS)</h4>
        <pre>ssl_cert_file (cfg): {ssl_cert_cfg}
ssl_key_file  (cfg): {ssl_key_cfg}
ssl_cert_abs (resolved): {ssl_cert_abs} exists={os.path.exists(ssl_cert_abs)}
ssl_key_abs  (resolved): {ssl_key_abs} exists={os.path.exists(ssl_key_abs)}</pre>
        <h4>IdP Signing Cert (Token Verification)</h4>
        <pre>cert_file_path (cfg): {cert_dir}
cert_file_name (cfg): {cert_name}</pre>
        <ul>
        {li_html}
        </ul>
        <h3>SSO Params</h3>
        <pre>idp_authorize_url: {idp_url}
client_id: {client_id}
sp_redirect_url (cfg): {sp_redirect_url_cfg}
computed redirect_uri: {redirect_uri}
</pre>
        <h3>Request</h3>
        <pre>scheme: {request.scheme}
Host: {request.headers.get('Host')}
X-Forwarded-Proto: {request.headers.get('X-Forwarded-Proto')}
X-Forwarded-Host: {request.headers.get('X-Forwarded-Host')}
</pre>
        """

        return html
    except Exception as e:
        return f"Diagnostics error: {e}", 500

# ===== Day 3 에러 핸들러 =====
@app.errorhandler(403)
def forbidden(e):
    """403 권한 없음 처리"""
    return render_template('errors/403.html'), 403

@app.errorhandler(401)
def unauthorized(e):
    """401 인증 필요"""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Authentication required'}), 401
    return redirect(url_for('login', next=request.url))

# ===== Day 2 권한 관리 라우트 =====
from permission_utils import check_permission, get_user_menus, clear_user_cache

@app.route('/admin/permissions')
@check_permission('permission_admin', 'view')
def permission_dashboard():
    """권한 관리 대시보드"""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # 통계 데이터 조회
        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM system_users) as user_count,
                (SELECT COUNT(*) FROM system_roles) as role_count,
                (SELECT COUNT(*) FROM user_role_mapping) as mapping_count,
                (SELECT COUNT(*) FROM access_audit_log
                 WHERE created_at > CURRENT_DATE) as today_access
        """)

        stats = cursor.fetchone()
        return render_template('admin/permission_dashboard.html',
                             stats=stats, menu=MENU_CONFIG)
    finally:
        cursor.close()
        conn.close()

@app.route('/admin/permissions/users')
@check_permission('permission_admin', 'view')
def permission_users():
    """사용자별 권한 관리"""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT u.emp_id, u.login_id, u.user_name, u.dept_name,
                   STRING_AGG(r.role_name, ', ') as roles
            FROM system_users u
            LEFT JOIN user_role_mapping um ON u.emp_id = um.emp_id
            LEFT JOIN system_roles r ON um.role_code = r.role_code
            GROUP BY u.emp_id, u.login_id, u.user_name, u.dept_name
            ORDER BY u.user_name
        """)

        users = cursor.fetchall()
        return render_template('admin/permission_users.html',
                             users=users, menu=MENU_CONFIG)
    finally:
        cursor.close()
        conn.close()

@app.route('/admin/permissions/audit')
@check_permission('permission_admin', 'view')
def permission_audit():
    """접근 로그 조회"""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT a.emp_id, a.login_id, a.action, a.menu_code,
                   a.resource_id, a.ip_address, a.success, a.error_message,
                   a.created_at, u.user_name
            FROM access_audit_log a
            LEFT JOIN system_users u ON a.emp_id = u.emp_id
            ORDER BY a.created_at DESC
            LIMIT 100
        """)

        logs = cursor.fetchall()
        return render_template('admin/permission_audit.html',
                             logs=logs, menu=MENU_CONFIG)
    finally:
        cursor.close()
        conn.close()

@app.route('/api/permissions/grant', methods=['POST'])
@check_permission('permission_admin', 'create')
def api_grant_permission():
    """권한 부여 API"""
    data = request.json
    emp_id = data.get('emp_id')
    menu_code = data.get('menu_code')
    permissions = data.get('permissions', {})

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO user_menu_permissions
            (emp_id, menu_code, can_view, can_create, can_edit, can_delete, data_scope, granted_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (emp_id, menu_code) DO UPDATE SET
                can_view = EXCLUDED.can_view,
                can_create = EXCLUDED.can_create,
                can_edit = EXCLUDED.can_edit,
                can_delete = EXCLUDED.can_delete,
                data_scope = EXCLUDED.data_scope,
                granted_at = CURRENT_TIMESTAMP
        """, (
            emp_id, menu_code,
            permissions.get('can_view', False),
            permissions.get('can_create', False),
            permissions.get('can_edit', False),
            permissions.get('can_delete', False),
            permissions.get('data_scope', 'own'),
            session.get('emp_id')
        ))

        conn.commit()

        # 캐시 클리어
        clear_user_cache(emp_id)

        return jsonify({'success': True})

    except Exception as e:
        conn.rollback()
        logging.error(f"권한 부여 실패: {e}")
        return jsonify({'success': False, 'error': str(e)})
    finally:
        cursor.close()
        conn.close()

@app.route('/api/permissions/cache/clear', methods=['POST'])
@check_permission('permission_admin', 'edit')
def api_clear_cache():
    """권한 캐시 클리어"""
    emp_id = request.json.get('emp_id') if request.json else None

    if emp_id:
        clear_user_cache(emp_id)
    else:
        clear_user_cache()  # 전체 캐시 클리어

    return jsonify({'success': True})

if __name__ == "__main__":
    print("Flask 앱 시작 중...", flush=True)
    
    # 데이터베이스 초기화 및 동기화 (서버 시작 시 한 번만 실행)
    print("데이터베이스 초기화 중...", flush=True)
    init_db()
    
    # 일일 동기화 체크 (24시간 경과 시만 동기화)
    from database_config import maybe_daily_sync
    if db_config.external_db_enabled:
        try:
            print("일일 동기화 체크 중...", flush=True)
            maybe_daily_sync()
        except Exception as e:
            print(f"일일 동기화 체크 중 오류: {e}", flush=True)
    
    # JSON 컬럼 설정 동기화 (조건부 - GPT 제안대로 단순화)
    if db_config.config.getboolean('COLUMNS', 'SYNC_ON_STARTUP', fallback=False):
        try:
            print("JSON 컬럼 설정 동기화 시작... (config: SYNC_ON_STARTUP=true)", flush=True)
            sync_service = ColumnSyncService(DB_PATH, 'columns')
            sync_results = sync_service.sync_all_boards()
            
            for board, count in sync_results.items():
                if count >= 0:
                    print(f"  - {board}: {count}개 컬럼 동기화 완료", flush=True)
                else:
                    print(f"  - {board}: 동기화 실패", flush=True)
                    
            print("JSON 컬럼 설정 동기화 완료!", flush=True)
        except Exception as e:
            print(f"컬럼 동기화 중 오류 발생: {e}", flush=True)
            # 동기화 실패해도 앱은 실행되도록 함
    else:
        print("JSON 동기화 건너뜀 (config: SYNC_ON_STARTUP=false)", flush=True)
        print("DB의 컬럼 설정을 그대로 사용합니다.", flush=True)
    
    # 스케줄러는 실제 메인 프로세스에서만 시작(리로더 중복 방지)
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        # 스케줄 등록: 매일 07:00에 maybe_daily_sync() 호출
        schedule.clear()  # 혹시 있을지 모르는 기존 잡 제거
        schedule.every().day.at("07:00").do(maybe_daily_sync)

        scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        scheduler_thread.start()

        logging.info("=" * 50)
        logging.info("자동 동기화 스케줄러 시작")
        logging.info("매일 오전 7시에 자동으로 데이터를 동기화합니다.")
        logging.info("=" * 50)
    
    print(f"partner-accident 라우트 등록됨: {'/partner-accident' in [rule.rule for rule in app.url_map.iter_rules()]}", flush=True)

    # SSL 설정 (config.ini에서 읽기)
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')

    ssl_cert = config.get('SSO', 'ssl_cert_file', fallback='Templates/Cert/cert.pem')
    ssl_key = config.get('SSO', 'ssl_key_file', fallback='Templates/Cert/key.pem')
    ssl_port = config.getint('SSO', 'ssl_port', fallback=44369)
    http_port = config.getint('SSO', 'http_port', fallback=5000)

    # SSL 인증서가 있으면 HTTPS, 없으면 HTTP
    ssl_context = ssl.SSLContext(ssl.PROTOCOL_TLS)
    try:
        # 상대경로일 경우 앱 루트 기준으로 보정
        base_dir = os.path.dirname(os.path.abspath(__file__))
        cert_path = ssl_cert if os.path.isabs(ssl_cert) else os.path.join(base_dir, ssl_cert)
        key_path = ssl_key if os.path.isabs(ssl_key) else os.path.join(base_dir, ssl_key)
        print(f"[SSO] Server TLS cert: {cert_path}")
        print(f"[SSO] Server TLS key : {key_path}")
        ssl_context.load_cert_chain(certfile=cert_path, keyfile=key_path)
        print(f"HTTPS 모드로 실행: https://localhost:{ssl_port}")
        app.run(host="0.0.0.0", port=ssl_port, ssl_context=ssl_context, debug=app.debug)
    except (FileNotFoundError, Exception) as e:
        print(f"SSL 인증서 문제: {e}")
        print(f"HTTP 모드로 실행: http://localhost:{http_port}")
        app.run(host="0.0.0.0", port=http_port, debug=app.debug)
